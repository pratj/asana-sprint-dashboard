#!/usr/bin/env python3
"""
Asana Daily Compliance Report Generator
========================================
Generates a daily report checking ticket compliance against team standards:

MANDATORY TICKET ATTRIBUTES:
- EPIC
- Sprint
- Type
- Story Points
- Severity
- Due Dates
- Comprehensive Description (ACs, Context, Specific Asks)

DAILY PROGRESS UPDATES (In Progress/Review/QA):
- Current Status comments
- Blockers identified
- Scope changes documented

COMPLETION PROTOCOL:
- Proper remarks on completed tickets
- Contributors tagged

Usage:
    python asana_daily_report.py [--output-dir /path/to/output] [--format markdown|html|json]

Requirements:
    pip install asana python-dotenv rich

Environment Variables:
    ASANA_ACCESS_TOKEN - Your Asana Personal Access Token
"""

import os
import sys
import json
import argparse
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional
from dataclasses import dataclass, field, asdict
from collections import defaultdict

try:
    import asana
    from asana.rest import ApiException
except ImportError:
    print("Error: asana package not installed. Run: pip install asana")
    sys.exit(1)

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    from rich.console import Console
    from rich.table import Table
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# Configuration
# =============================================================================

@dataclass
class Config:
    """Configuration for the Asana report generator.

    All GIDs can be overridden via environment variables.
    """
    workspace_gid: str = field(
        default_factory=lambda: os.environ.get("ASANA_WORKSPACE_GID", "1198498469382287")
    )
    project_gid: str = field(
        default_factory=lambda: os.environ.get("ASANA_PROJECT_GID", "1210472897337973")
    )

    # Custom field GIDs (all configurable via environment variables)
    sprint_field_gid: str = field(
        default_factory=lambda: os.environ.get("ASANA_SPRINT_FIELD_GID", "1210724429562701")
    )
    progress_field_gid: str = field(
        default_factory=lambda: os.environ.get("ASANA_PROGRESS_FIELD_GID", "1210724432100811")
    )
    epic_field_gid: str = field(
        default_factory=lambda: os.environ.get("ASANA_EPIC_FIELD_GID", "1210724432100829")
    )
    type_field_gid: str = field(
        default_factory=lambda: os.environ.get("ASANA_TYPE_FIELD_GID", "1210724432100822")
    )
    severity_field_gid: str = field(
        default_factory=lambda: os.environ.get("ASANA_SEVERITY_FIELD_GID", "1211033986599429")
    )
    points_field_gid: str = field(
        default_factory=lambda: os.environ.get("ASANA_POINTS_FIELD_GID", "1210724432100820")
    )

    # Progress statuses requiring daily updates
    active_statuses: tuple = ("In Progress", "Review", "QA")

    # Statuses to exclude from compliance checks entirely
    excluded_statuses: tuple = ("Backlog",)

    # Statuses that don't require daily updates but need other compliance checks
    pending_statuses: tuple = ("To Do",)

    # Thresholds
    min_description_length: int = 100  # Minimum chars for description
    hours_without_update: int = 24     # Hours before flagging no update

    # Valid Fibonacci story points (0 allowed for epics/parent tasks, max 13)
    valid_story_points: tuple = (0, 1, 2, 3, 5, 8, 13)

    # Output
    output_dir: str = "./reports"
    output_format: str = "markdown"

    # Validation rules: types that should NOT have story points
    types_without_points: tuple = ("Epic", "Bug")


# =============================================================================
# Data Models
# =============================================================================

@dataclass
class TaskCompliance:
    """Compliance analysis of a single task."""
    gid: str
    name: str
    url: str
    assignee: str
    assignee_gid: Optional[str]
    created_at: str

    # Current state
    progress: Optional[str] = None
    due_on: Optional[str] = None
    completed_at: Optional[str] = None  # ISO datetime when task was completed

    # Field values
    epic: Optional[str] = None
    sprint: Optional[str] = None
    task_type: Optional[str] = None
    story_points: Optional[str] = None
    severity: Optional[str] = None
    description_length: int = 0

    # Comments/Updates
    last_comment_date: Optional[str] = None
    last_comment_author: Optional[str] = None
    hours_since_update: Optional[float] = None
    total_comments: int = 0

    # Compliance flags - Mandatory Attributes
    missing_epic: bool = False
    missing_sprint: bool = False
    missing_type: bool = False
    missing_points: bool = False
    invalid_points: bool = False  # Story points not a Fibonacci number
    missing_severity: bool = False
    missing_due_date: bool = False
    missing_description: bool = False

    # Compliance flags - Daily Updates
    needs_daily_update: bool = False
    missing_daily_update: bool = False

    # Compliance flags - Completion
    missing_completion_remarks: bool = False

    # Rule violations
    rule_violations: list[str] = field(default_factory=list)

    # Quick Wins - Computed fields for overdue/due soon tracking
    is_overdue: bool = False
    days_until_due: Optional[int] = None  # Negative if overdue
    task_age_days: int = 0  # Days since created

    @property
    def mandatory_missing(self) -> list[str]:
        """List of missing or invalid mandatory attributes."""
        missing = []
        if self.missing_epic:
            missing.append("Epic")
        if self.missing_sprint:
            missing.append("Sprint")
        if self.missing_type:
            missing.append("Type")
        if self.missing_points:
            missing.append("Story Points")
        if self.invalid_points:
            missing.append("Invalid Points (non-Fibonacci)")
        if self.missing_severity:
            missing.append("Severity")
        if self.missing_due_date:
            missing.append("Due Date")
        if self.missing_description:
            missing.append("Description/ACs")
        return missing

    @property
    def mandatory_count(self) -> int:
        return len(self.mandatory_missing)

    @property
    def total_issues(self) -> int:
        """Total compliance issues including rule violations."""
        return self.mandatory_count + len(self.rule_violations)

    @property
    def is_compliant(self) -> bool:
        """Check if task meets all compliance requirements."""
        return (
            self.mandatory_count == 0 and
            len(self.rule_violations) == 0 and
            not self.missing_daily_update and
            not self.missing_completion_remarks
        )

    @property
    def compliance_score(self) -> int:
        """Score from 0-100 based on compliance."""
        total_checks = 7  # mandatory fields
        passed = 7 - self.mandatory_count

        if self.needs_daily_update:
            total_checks += 1
            if not self.missing_daily_update:
                passed += 1

        # Rule violations count as failed checks
        if self.rule_violations:
            total_checks += len(self.rule_violations)
            # passed stays the same (violations = 0 passed)

        return int((passed / total_checks) * 100) if total_checks > 0 else 100

    @property
    def is_todo(self) -> bool:
        """Check if task is in To Do status."""
        return self.progress == "To Do"

    @property
    def status_label(self) -> str:
        """Get a display label for the task status."""
        return self.progress or "Unknown"


@dataclass
class ReportSummary:
    """Summary statistics for compliance report."""
    total_tasks: int = 0
    compliant_tasks: int = 0
    compliance_rate: float = 0.0

    # Mandatory attributes missing
    missing_epic: int = 0
    missing_sprint: int = 0
    missing_type: int = 0
    missing_points: int = 0
    invalid_points: int = 0  # Non-Fibonacci story points
    missing_severity: int = 0
    missing_due_date: int = 0
    missing_description: int = 0

    # Rule violations
    rule_violations: int = 0  # Tasks with rule violations (e.g., Epic/Bug with story points)

    # Daily updates
    tasks_needing_updates: int = 0
    tasks_missing_updates: int = 0

    # Status counts
    tasks_todo: int = 0
    tasks_active: int = 0

    # By assignee
    by_assignee: dict = field(default_factory=dict)

    # Report metadata
    report_date: str = ""      # Date of the report (YYYY-MM-DD)
    generated_at: str = ""     # Full timestamp when report was generated

    # Quick Wins metrics
    overdue_tasks: int = 0
    due_this_week: int = 0
    overdue_points: float = 0
    due_this_week_points: float = 0


# =============================================================================
# Asana Client
# =============================================================================

class AsanaClient:
    """Client for Asana API interactions."""

    def __init__(self, access_token: str, config: Config):
        self.config = config
        configuration = asana.Configuration()
        configuration.access_token = access_token
        self.api_client = asana.ApiClient(configuration)
        self.tasks_api = asana.TasksApi(self.api_client)
        self.stories_api = asana.StoriesApi(self.api_client)

    def get_tasks(
        self,
        completed: bool = False,
        completed_since: Optional[str] = None,
        modified_since: Optional[str] = None
    ) -> list[dict]:
        """Fetch tasks from project.

        Args:
            completed: If True, fetch completed tasks. If False, fetch incomplete tasks.
            completed_since: ISO date string to filter completed tasks (only for completed=True)
            modified_since: ISO date string to filter by modification date
        """
        tasks = []
        opt_fields = [
            "name", "assignee", "assignee.name", "due_on", "notes",
            "completed", "completed_at", "created_at", "modified_at",
            "custom_fields", "custom_fields.name", "custom_fields.display_value",
            "custom_fields.number_value", "permalink_url"
        ]

        try:
            opts = {
                "opt_fields": ",".join(opt_fields),
            }

            # Build search params separately (Asana SDK requires this format)
            search_params = {
                "projects.any": self.config.project_gid,
                "completed": completed,
            }

            # Add date filters if provided
            if completed and completed_since:
                search_params["completed_at.after"] = completed_since
            if modified_since:
                search_params["modified_at.after"] = modified_since

            # Merge search params into opts for the API call
            opts.update(search_params)

            result = self.tasks_api.search_tasks_for_workspace(
                self.config.workspace_gid,
                opts=opts
            )
            for task in result:
                task_dict = task.to_dict() if hasattr(task, 'to_dict') else dict(task)
                tasks.append(task_dict)
        except ApiException as e:
            print(f"Error fetching tasks: {e}")
            raise

        return tasks

    def get_completed_tasks(self, since_days: int = 30) -> list[dict]:
        """Fetch recently completed tasks.

        Args:
            since_days: Number of days to look back for completed tasks
        """
        since_date = (datetime.now(timezone.utc) - timedelta(days=since_days)).isoformat()
        return self.get_tasks(completed=True, completed_since=since_date)

    def get_task_comments(self, task_gid: str, limit: int = 10) -> list[dict]:
        """Fetch recent comments/stories for a task."""
        comments = []
        try:
            result = self.stories_api.get_stories_for_task(
                task_gid,
                opts={
                    "opt_fields": "created_at,created_by,created_by.name,text,resource_subtype",
                    "limit": limit
                }
            )
            for story in result:
                story_dict = story.to_dict() if hasattr(story, 'to_dict') else dict(story)
                # Only include actual comments, not system stories
                if story_dict.get('resource_subtype') == 'comment_added':
                    comments.append(story_dict)
        except ApiException as e:
            print(f"Error fetching comments for task {task_gid}: {e}")

        return comments


# =============================================================================
# Compliance Analyzer
# =============================================================================

class ComplianceAnalyzer:
    """Analyzes tasks for compliance with team standards."""

    def __init__(self, config: Config, client: AsanaClient):
        self.config = config
        self.client = client

    def analyze_task(self, task: dict, fetch_comments: bool = True) -> TaskCompliance:
        """Analyze a single task for compliance."""
        gid = task.get('gid', '')
        name = task.get('name', '(unnamed)')

        # Assignee
        assignee_data = task.get('assignee') or {}
        assignee = assignee_data.get('name', 'Unassigned') if assignee_data else 'Unassigned'
        assignee_gid = assignee_data.get('gid') if assignee_data else None

        # Basic fields
        due_on = task.get('due_on')
        notes = task.get('notes', '') or ''
        created_at = task.get('created_at', '')
        modified_at = task.get('modified_at', '')  # ISO datetime when task was last modified
        completed_at = task.get('completed_at')  # ISO datetime when task was completed
        url = task.get('permalink_url', f'https://app.asana.com/0/{self.config.project_gid}/{gid}')

        # Extract custom fields
        custom_fields = task.get('custom_fields', []) or []
        sprint = epic = progress = task_type = severity = story_points = None

        for cf in custom_fields:
            if not cf:
                continue
            cf_gid = cf.get('gid', '')
            display_value = cf.get('display_value')
            number_value = cf.get('number_value')

            if cf_gid == self.config.sprint_field_gid:
                sprint = display_value
            elif cf_gid == self.config.epic_field_gid:
                epic = display_value
            elif cf_gid == self.config.progress_field_gid:
                progress = display_value
            elif cf_gid == self.config.type_field_gid:
                task_type = display_value
            elif cf_gid == self.config.severity_field_gid:
                severity = display_value
            elif cf_gid == self.config.points_field_gid:
                story_points = str(number_value) if number_value is not None else None

        # Create compliance record
        compliance = TaskCompliance(
            gid=gid,
            name=name,
            url=url,
            assignee=assignee,
            assignee_gid=assignee_gid,
            created_at=created_at,
            progress=progress,
            due_on=due_on,
            completed_at=completed_at,
            epic=epic,
            sprint=sprint,
            task_type=task_type,
            story_points=story_points,
            severity=severity,
            description_length=len(notes),
        )

        # Calculate overdue and due soon (Quick Wins)
        from datetime import date
        today = date.today()

        if due_on:
            try:
                due_date = datetime.strptime(due_on, '%Y-%m-%d').date()
                compliance.days_until_due = (due_date - today).days
                compliance.is_overdue = compliance.days_until_due < 0 and progress != "Done"
            except (ValueError, TypeError):
                pass

        # Calculate task age (days since created)
        if created_at:
            try:
                created = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                compliance.task_age_days = (datetime.now(timezone.utc) - created).days
            except (ValueError, TypeError):
                pass

        # Check mandatory attributes
        compliance.missing_epic = not epic or epic.strip() == ''
        compliance.missing_sprint = not sprint or sprint.strip() == ''
        compliance.missing_type = not task_type or task_type.strip() == ''

        # Story points: only required for types NOT in types_without_points (e.g., Bugs/Epics don't need points)
        if task_type in self.config.types_without_points:
            compliance.missing_points = False  # Bugs/Epics don't need story points
        else:
            compliance.missing_points = story_points is None or story_points == 'None'

        # Check if story points are valid Fibonacci numbers (0, 1, 2, 3, 5, 8, 13)
        # Skip validation for types that shouldn't have points
        if not compliance.missing_points and task_type not in self.config.types_without_points:
            try:
                points_value = float(story_points)
                # Check if it's a whole number and in the valid list
                if points_value != int(points_value) or int(points_value) not in self.config.valid_story_points:
                    compliance.invalid_points = True
            except (ValueError, TypeError):
                compliance.invalid_points = True

        compliance.missing_severity = not severity or severity.strip() == ''
        compliance.missing_due_date = due_on is None
        compliance.missing_description = len(notes) < self.config.min_description_length

        # Check if task needs daily updates
        compliance.needs_daily_update = progress in self.config.active_statuses

        # Validate rules: check for types that should not have story points
        if task_type in self.config.types_without_points and story_points is not None and story_points != 'None':
            compliance.rule_violations.append(f"{task_type} should not have story points")

        # Check for daily updates on active tasks
        if fetch_comments and compliance.needs_daily_update:
            now = datetime.now(timezone.utc)
            last_activity_time = None

            # Check modified_at (captures status changes, field updates, etc.)
            if modified_at:
                try:
                    mod_time = datetime.fromisoformat(modified_at.replace('Z', '+00:00'))
                    last_activity_time = mod_time
                except (ValueError, TypeError):
                    pass

            # Fetch comments
            comments = self.client.get_task_comments(gid, limit=5)
            compliance.total_comments = len(comments)

            if comments:
                # Get most recent comment
                latest = comments[0]
                compliance.last_comment_date = latest.get('created_at')
                created_by = latest.get('created_by') or {}
                compliance.last_comment_author = created_by.get('name', 'Unknown')

                # Check if comment is more recent than modified_at
                if compliance.last_comment_date:
                    try:
                        comment_time = datetime.fromisoformat(
                            compliance.last_comment_date.replace('Z', '+00:00')
                        )
                        if last_activity_time is None or comment_time > last_activity_time:
                            last_activity_time = comment_time
                    except (ValueError, TypeError):
                        pass

            # Calculate hours since last activity (comment OR modification)
            if last_activity_time:
                delta = now - last_activity_time
                compliance.hours_since_update = delta.total_seconds() / 3600

                if compliance.hours_since_update > self.config.hours_without_update:
                    compliance.missing_daily_update = True
            else:
                # No activity tracked at all
                compliance.missing_daily_update = True

        return compliance

    def analyze_all(
        self,
        tasks: list[dict],
        fetch_comments: bool = True,
        include_done: bool = False
    ) -> list[TaskCompliance]:
        """Analyze all tasks for compliance.

        Args:
            tasks: List of task dictionaries from Asana API
            fetch_comments: Whether to fetch comments for active tasks
            include_done: If True, include Done tasks (useful for burndown charts).
                         If False (default), skip Done tasks for compliance analysis.
        """
        results = []
        total = len(tasks)
        skipped_done = 0
        skipped_backlog = 0

        for i, task in enumerate(tasks, 1):
            # Get progress status
            progress = None
            for cf in (task.get('custom_fields') or []):
                if cf and cf.get('gid') == self.config.progress_field_gid:
                    progress = cf.get('display_value')
                    break

            # Skip Done tasks (unless include_done is True)
            if progress == 'Done' and not include_done:
                skipped_done += 1
                continue

            # Skip Backlog tasks (excluded from compliance checks)
            if progress in self.config.excluded_statuses:
                skipped_backlog += 1
                continue

            if i % 10 == 0:
                print(f"  Analyzing task {i}/{total}...")

            compliance = self.analyze_task(task, fetch_comments=fetch_comments)
            results.append(compliance)

        print(f"  Skipped {skipped_done} Done tasks, {skipped_backlog} Backlog tasks")
        return results

    def get_unique_sprints(self, results: list[TaskCompliance]) -> list[str]:
        """Extract unique sprint values from results, sorted naturally (Sprint 2 before Sprint 10)."""
        import re
        sprints = set()
        for task in results:
            if task.sprint and task.sprint.strip():
                # Split comma-separated sprints (multi-enum values) into individual sprints
                for s in task.sprint.split(","):
                    s = s.strip()
                    if s:
                        sprints.add(s)

        def natural_sort_key(s: str):
            """Sort strings with embedded numbers naturally."""
            return [int(x) if x.isdigit() else x.lower() for x in re.split(r'(\d+)', s)]

        return sorted(sprints, key=natural_sort_key)

    def get_unique_assignees(self, results: list[TaskCompliance]) -> list[str]:
        """Extract unique assignee names from results, sorted."""
        assignees = set()
        for task in results:
            if task.assignee and task.assignee != "Unassigned":
                assignees.add(task.assignee)
        return sorted(assignees)

    def get_unique_statuses(self, results: list[TaskCompliance]) -> list[str]:
        """Extract unique progress statuses from results."""
        statuses = set()
        for task in results:
            if task.progress:
                statuses.add(task.progress)
        # Return in logical order
        status_order = ["To Do", "In Progress", "Review", "QA", "Done", "Backlog"]
        return [s for s in status_order if s in statuses] + sorted(statuses - set(status_order))

    def get_unique_epics(self, results: list[TaskCompliance]) -> list[str]:
        """Extract unique epic values from results, sorted."""
        epics = set()
        for task in results:
            if task.epic and task.epic.strip():
                epics.add(task.epic)
        return sorted(epics)

    def filter_results(
        self,
        results: list[TaskCompliance],
        sprint: Optional[str] = None,
        assignees: Optional[list[str]] = None,
        statuses: Optional[list[str]] = None,
        epics: Optional[list[str]] = None,
        due_date_start: Optional[str] = None,
        due_date_end: Optional[str] = None,
        created_date_start: Optional[str] = None,
        created_date_end: Optional[str] = None,
    ) -> list[TaskCompliance]:
        """Filter results based on various criteria.

        Args:
            results: List of TaskCompliance objects to filter
            sprint: Sprint name to filter by (None or "All" means no filter)
            assignees: List of assignee names to include (None means all)
            statuses: List of progress statuses to include (None means all)
            epics: List of epic names to include (None means all)
            due_date_start: ISO date string for due date range start
            due_date_end: ISO date string for due date range end
            created_date_start: ISO date string for created date range start
            created_date_end: ISO date string for created date range end
        """
        filtered = results

        # Filter by sprint (handles comma-separated sprint values like "Manali, London")
        if sprint and sprint != "All":
            def task_in_sprint(task: TaskCompliance) -> bool:
                if not task.sprint:
                    return False
                task_sprints = [s.strip() for s in task.sprint.split(",")]
                return sprint in task_sprints
            filtered = [t for t in filtered if task_in_sprint(t)]

        # Filter by assignees
        if assignees and len(assignees) > 0:
            filtered = [t for t in filtered if t.assignee in assignees]

        # Filter by statuses
        if statuses and len(statuses) > 0:
            filtered = [t for t in filtered if t.progress in statuses]

        # Filter by epics
        if epics and len(epics) > 0:
            filtered = [t for t in filtered if t.epic in epics]

        # Filter by due date range
        if due_date_start:
            filtered = [t for t in filtered if t.due_on and t.due_on >= due_date_start]
        if due_date_end:
            filtered = [t for t in filtered if t.due_on and t.due_on <= due_date_end]

        # Filter by created date range
        if created_date_start:
            filtered = [
                t for t in filtered
                if t.created_at and t.created_at[:10] >= created_date_start
            ]
        if created_date_end:
            filtered = [
                t for t in filtered
                if t.created_at and t.created_at[:10] <= created_date_end
            ]

        return filtered

    def calculate_sprint_metrics(self, results: list[TaskCompliance]) -> dict:
        """Calculate detailed metrics for sprint analytics.

        Returns dict with:
            - total_points: Total story points
            - completed_points: Completed story points (Done status)
            - remaining_points: Remaining story points
            - points_by_status: Dict of status -> points
            - tasks_by_status: Dict of status -> count
            - points_by_assignee: Dict of assignee -> points
            - avg_points_per_task: Average story points per task
        """
        total_points = 0
        completed_points = 0
        points_by_status = {}
        tasks_by_status = {}
        points_by_assignee = {}

        for task in results:
            # Parse story points
            try:
                points = float(task.story_points) if task.story_points else 0
            except (ValueError, TypeError):
                points = 0

            total_points += points

            # By status
            status = task.progress or "Unknown"
            if status not in points_by_status:
                points_by_status[status] = 0
                tasks_by_status[status] = 0
            points_by_status[status] += points
            tasks_by_status[status] += 1

            # Track completed
            if status == "Done":
                completed_points += points

            # By assignee
            assignee = task.assignee or "Unassigned"
            if assignee not in points_by_assignee:
                points_by_assignee[assignee] = 0
            points_by_assignee[assignee] += points

        remaining_points = total_points - completed_points
        avg_points = total_points / len(results) if results else 0

        return {
            "total_points": total_points,
            "completed_points": completed_points,
            "remaining_points": remaining_points,
            "points_by_status": points_by_status,
            "tasks_by_status": tasks_by_status,
            "points_by_assignee": dict(sorted(
                points_by_assignee.items(),
                key=lambda x: x[1],
                reverse=True
            )),
            "avg_points_per_task": round(avg_points, 1),
        }

    def generate_summary(self, results: list[TaskCompliance]) -> ReportSummary:
        """Generate summary statistics."""
        now = datetime.now()
        summary = ReportSummary(
            total_tasks=len(results),
            report_date=now.strftime('%Y-%m-%d'),
            generated_at=now.isoformat()
        )

        by_assignee = defaultdict(lambda: {"total": 0, "issues": 0})

        for task in results:
            # Count mandatory missing
            if task.missing_epic:
                summary.missing_epic += 1
            if task.missing_sprint:
                summary.missing_sprint += 1
            if task.missing_type:
                summary.missing_type += 1
            if task.missing_points:
                summary.missing_points += 1
            if task.invalid_points:
                summary.invalid_points += 1
            if task.missing_severity:
                summary.missing_severity += 1
            if task.missing_due_date:
                summary.missing_due_date += 1
            if task.missing_description:
                summary.missing_description += 1

            # Count rule violations
            if task.rule_violations:
                summary.rule_violations += 1

            # Count daily updates
            if task.needs_daily_update:
                summary.tasks_needing_updates += 1
                summary.tasks_active += 1
                if task.missing_daily_update:
                    summary.tasks_missing_updates += 1

            # Count To Do tasks
            if task.is_todo:
                summary.tasks_todo += 1

            # Count compliant
            if task.is_compliant:
                summary.compliant_tasks += 1

            # Quick Wins: Count overdue and due this week
            if task.is_overdue:
                summary.overdue_tasks += 1
                try:
                    points = float(task.story_points) if task.story_points else 0
                    summary.overdue_points += points
                except (ValueError, TypeError):
                    pass

            if (task.days_until_due is not None
                and 0 <= task.days_until_due <= 7
                and task.progress != "Done"):
                summary.due_this_week += 1
                try:
                    points = float(task.story_points) if task.story_points else 0
                    summary.due_this_week_points += points
                except (ValueError, TypeError):
                    pass

            # By assignee
            by_assignee[task.assignee]["total"] += 1
            if not task.is_compliant:
                by_assignee[task.assignee]["issues"] += 1

        # Calculate compliance rate
        if summary.total_tasks > 0:
            summary.compliance_rate = (summary.compliant_tasks / summary.total_tasks) * 100

        # Sort by issues descending
        summary.by_assignee = dict(
            sorted(by_assignee.items(), key=lambda x: x[1]["issues"], reverse=True)
        )

        return summary


# =============================================================================
# Report Generators
# =============================================================================

class MarkdownReportGenerator:
    """Generates markdown compliance reports."""

    def __init__(self, config: Config):
        self.config = config

    def generate(self, results: list[TaskCompliance], summary: ReportSummary) -> str:
        """Generate markdown report."""
        lines = []

        # Header
        lines.append("# Asana Ticket Compliance Report")
        lines.append("## Unified Partner Portal - Dev Team")
        lines.append("")
        lines.append(f"**Report Date:** {summary.report_date}")
        lines.append(f"**Generated At:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")

        # Compliance Score
        lines.append("---")
        lines.append("")
        lines.append("## Overall Compliance")
        lines.append("")
        lines.append(f"| Metric | Value |")
        lines.append("|--------|-------|")
        lines.append(f"| **Total Tasks** | {summary.total_tasks} |")
        lines.append(f"| **Compliant Tasks** | {summary.compliant_tasks} |")
        lines.append(f"| **Compliance Rate** | {summary.compliance_rate:.1f}% |")
        lines.append("")

        # Mandatory Attributes Summary
        lines.append("---")
        lines.append("")
        lines.append("## Mandatory Attributes Missing")
        lines.append("")
        lines.append("| Attribute | Missing Count | % of Tasks |")
        lines.append("|-----------|---------------|------------|")

        attrs = [
            ("Epic", summary.missing_epic),
            ("Sprint", summary.missing_sprint),
            ("Type", summary.missing_type),
            ("Story Points", summary.missing_points),
            ("Severity", summary.missing_severity),
            ("Due Date", summary.missing_due_date),
            ("Description/ACs", summary.missing_description),
        ]
        for name, count in attrs:
            pct = (count / summary.total_tasks * 100) if summary.total_tasks > 0 else 0
            lines.append(f"| **{name}** | {count} | {pct:.1f}% |")
        lines.append("")

        # Daily Updates Summary
        lines.append("---")
        lines.append("")
        lines.append("## Daily Progress Updates")
        lines.append("")
        lines.append(f"Tasks in **In Progress/Review/QA** requiring daily updates: **{summary.tasks_needing_updates}**")
        lines.append("")
        lines.append(f"Tasks **missing updates** (no comment in last 24h): **{summary.tasks_missing_updates}**")
        lines.append("")

        # By Assignee
        lines.append("---")
        lines.append("")
        lines.append("## Compliance by Assignee")
        lines.append("")
        lines.append("| Assignee | Total Tasks | Issues | Compliance |")
        lines.append("|----------|-------------|--------|------------|")
        for assignee, data in summary.by_assignee.items():
            total = data["total"]
            issues = data["issues"]
            compliant = total - issues
            rate = (compliant / total * 100) if total > 0 else 100
            lines.append(f"| **{assignee}** | {total} | {issues} | {rate:.0f}% |")
        lines.append("")

        # Detailed Tables
        lines.append("---")
        lines.append("")
        lines.append("## Detailed Findings")
        lines.append("")

        # Missing Epic
        missing_epic = [t for t in results if t.missing_epic]
        lines.append(f"### Missing Epic ({len(missing_epic)} tasks)")
        lines.append("")
        if missing_epic:
            lines.append("| Task | Assignee | Sprint | Progress | Link |")
            lines.append("|------|----------|--------|----------|------|")
            for t in missing_epic:
                name = t.name.replace("|", "-")[:50]
                lines.append(f"| {name} | {t.assignee} | {t.sprint or 'None'} | {t.progress or 'None'} | [Open]({t.url}) |")
        else:
            lines.append("[OK] All tasks have Epic assigned")
        lines.append("")

        # Missing Sprint
        missing_sprint = [t for t in results if t.missing_sprint]
        lines.append(f"### Missing Sprint ({len(missing_sprint)} tasks)")
        lines.append("")
        if missing_sprint:
            lines.append("| Task | Assignee | Epic | Progress | Link |")
            lines.append("|------|----------|------|----------|------|")
            for t in missing_sprint:
                name = t.name.replace("|", "-")[:50]
                lines.append(f"| {name} | {t.assignee} | {t.epic or 'None'} | {t.progress or 'None'} | [Open]({t.url}) |")
        else:
            lines.append("[OK] All tasks have Sprint assigned")
        lines.append("")

        # Missing Type
        missing_type = [t for t in results if t.missing_type]
        lines.append(f"### Missing Type ({len(missing_type)} tasks)")
        lines.append("")
        if missing_type:
            lines.append("| Task | Assignee | Progress | Link |")
            lines.append("|------|----------|----------|------|")
            for t in missing_type:
                name = t.name.replace("|", "-")[:50]
                lines.append(f"| {name} | {t.assignee} | {t.progress or 'None'} | [Open]({t.url}) |")
        else:
            lines.append("[OK] All tasks have Type assigned")
        lines.append("")

        # Missing Story Points
        missing_points = [t for t in results if t.missing_points]
        lines.append(f"### Missing Story Points ({len(missing_points)} tasks)")
        lines.append("")
        if missing_points:
            lines.append("| Task | Assignee | Type | Progress | Link |")
            lines.append("|------|----------|------|----------|------|")
            for t in missing_points:
                name = t.name.replace("|", "-")[:50]
                lines.append(f"| {name} | {t.assignee} | {t.task_type or 'None'} | {t.progress or 'None'} | [Open]({t.url}) |")
        else:
            lines.append("[OK] All tasks have Story Points assigned")
        lines.append("")

        # Missing Severity
        missing_severity = [t for t in results if t.missing_severity]
        lines.append(f"### Missing Severity ({len(missing_severity)} tasks)")
        lines.append("")
        if missing_severity:
            lines.append("| Task | Assignee | Type | Progress | Link |")
            lines.append("|------|----------|------|----------|------|")
            for t in missing_severity:
                name = t.name.replace("|", "-")[:50]
                lines.append(f"| {name} | {t.assignee} | {t.task_type or 'None'} | {t.progress or 'None'} | [Open]({t.url}) |")
        else:
            lines.append("[OK] All tasks have Severity assigned")
        lines.append("")

        # Missing Due Date
        missing_due = [t for t in results if t.missing_due_date]
        lines.append(f"### Missing Due Date ({len(missing_due)} tasks)")
        lines.append("")
        if missing_due:
            lines.append("| Task | Assignee | Sprint | Progress | Link |")
            lines.append("|------|----------|--------|----------|------|")
            for t in missing_due:
                name = t.name.replace("|", "-")[:50]
                lines.append(f"| {name} | {t.assignee} | {t.sprint or 'None'} | {t.progress or 'None'} | [Open]({t.url}) |")
        else:
            lines.append("[OK] All tasks have Due Date set")
        lines.append("")

        # Missing Description/ACs
        missing_desc = [t for t in results if t.missing_description]
        lines.append(f"### Missing Description/ACs ({len(missing_desc)} tasks)")
        lines.append(f"*(Tasks with less than {self.config.min_description_length} characters in description)*")
        lines.append("")
        if missing_desc:
            lines.append("| Task | Assignee | Chars | Progress | Link |")
            lines.append("|------|----------|-------|----------|------|")
            for t in missing_desc:
                name = t.name.replace("|", "-")[:50]
                lines.append(f"| {name} | {t.assignee} | {t.description_length} | {t.progress or 'None'} | [Open]({t.url}) |")
        else:
            lines.append("[OK] All tasks have adequate descriptions")
        lines.append("")

        # Missing Daily Updates
        missing_updates = [t for t in results if t.missing_daily_update]
        lines.append(f"### Missing Daily Updates ({len(missing_updates)} tasks)")
        lines.append("*(Active tasks with no comment in the last 24 hours)*")
        lines.append("")
        if missing_updates:
            lines.append("| Task | Assignee | Status | Last Update | Hours Ago | Link |")
            lines.append("|------|----------|--------|-------------|-----------|------|")
            for t in missing_updates:
                name = t.name.replace("|", "-")[:40]
                last_update = t.last_comment_date[:10] if t.last_comment_date else "Never"
                hours = f"{t.hours_since_update:.0f}h" if t.hours_since_update else "N/A"
                lines.append(f"| {name} | {t.assignee} | {t.progress} | {last_update} | {hours} | [Open]({t.url}) |")
        else:
            lines.append("[OK] All active tasks have recent updates")
        lines.append("")

        # Action Items
        lines.append("---")
        lines.append("")
        lines.append("## Recommended Actions")
        lines.append("")

        # Top offenders
        top_assignees = list(summary.by_assignee.items())[:5]
        for i, (assignee, data) in enumerate(top_assignees, 1):
            if data["issues"] > 0:
                lines.append(f"{i}. **{assignee}** - {data['issues']} tickets need attention")

        lines.append("")
        lines.append("### Priority Actions:")
        lines.append("")
        if summary.tasks_missing_updates > 0:
            lines.append(f"1. [CRITICAL] **{summary.tasks_missing_updates} active tickets** need daily status updates")
        if summary.missing_due_date > 0:
            lines.append(f"2. [HIGH] **{summary.missing_due_date} tickets** need due dates assigned")
        if summary.missing_description > 0:
            lines.append(f"3. [HIGH] **{summary.missing_description} tickets** need proper descriptions with ACs")
        if summary.missing_points > 0:
            lines.append(f"4. [MEDIUM] **{summary.missing_points} tickets** need story points estimated")

        lines.append("")

        return "\n".join(lines)


class HTMLReportGenerator:
    """Generates HTML compliance reports."""

    def __init__(self, config: Config):
        self.config = config

    def generate(self, results: list[TaskCompliance], summary: ReportSummary) -> str:
        """Generate HTML report."""
        css = """
        <style>
            * { box-sizing: border-box; }
            body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 0; padding: 20px; background: #f0f2f5; }
            .container { max-width: 1400px; margin: 0 auto; }
            .card { background: white; border-radius: 12px; padding: 24px; margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
            h1 { color: #1a1a2e; margin-top: 0; }
            h2 { color: #16213e; border-bottom: 2px solid #e8e8e8; padding-bottom: 10px; margin-top: 0; }
            h3 { color: #0f3460; margin-top: 24px; }
            .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 24px; }
            .stat-card { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 12px; text-align: center; }
            .stat-card.warning { background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); }
            .stat-card.success { background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); }
            .stat-card h3 { margin: 0; font-size: 36px; color: white; }
            .stat-card p { margin: 8px 0 0 0; opacity: 0.9; }
            table { width: 100%; border-collapse: collapse; margin: 16px 0; font-size: 14px; }
            th { background: #f8f9fa; color: #495057; font-weight: 600; text-align: left; padding: 12px; border-bottom: 2px solid #dee2e6; }
            td { padding: 10px 12px; border-bottom: 1px solid #e9ecef; }
            tr:hover { background: #f8f9fa; }
            a { color: #667eea; text-decoration: none; }
            a:hover { text-decoration: underline; }
            .badge { display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: 500; }
            .badge-danger { background: #fee2e2; color: #991b1b; }
            .badge-warning { background: #fef3c7; color: #92400e; }
            .badge-success { background: #d1fae5; color: #065f46; }
            .progress-bar { height: 8px; background: #e9ecef; border-radius: 4px; overflow: hidden; }
            .progress-bar-fill { height: 100%; background: linear-gradient(90deg, #667eea, #764ba2); transition: width 0.3s; }
            .timestamp { color: #6b7280; font-size: 14px; }
            .section-header { display: flex; justify-content: space-between; align-items: center; }
            .count-badge { background: #667eea; color: white; padding: 4px 12px; border-radius: 20px; font-size: 14px; }
        </style>
        """

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Asana Compliance Report - {summary.report_date}</title>
    {css}
</head>
<body>
<div class="container">
    <div class="card">
        <h1>Asana Ticket Compliance Report</h1>
        <p class="timestamp"><strong>Report Date:</strong> {summary.report_date}</p>
        <p class="timestamp">Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>

    <div class="stats-grid">
        <div class="stat-card {'success' if summary.compliance_rate >= 80 else 'warning'}">
            <h3>{summary.compliance_rate:.0f}%</h3>
            <p>Compliance Rate</p>
        </div>
        <div class="stat-card">
            <h3>{summary.total_tasks}</h3>
            <p>Total Tasks</p>
        </div>
        <div class="stat-card {'success' if summary.compliant_tasks == summary.total_tasks else ''}">
            <h3>{summary.compliant_tasks}</h3>
            <p>Compliant Tasks</p>
        </div>
        <div class="stat-card {'warning' if summary.tasks_missing_updates > 0 else 'success'}">
            <h3>{summary.tasks_missing_updates}</h3>
            <p>Missing Daily Updates</p>
        </div>
    </div>

    <div class="card">
        <h2>Mandatory Attributes Missing</h2>
        <table>
            <tr>
                <th>Attribute</th>
                <th>Missing</th>
                <th>% of Tasks</th>
                <th>Status</th>
            </tr>"""

        attrs = [
            ("Epic", summary.missing_epic),
            ("Sprint", summary.missing_sprint),
            ("Type", summary.missing_type),
            ("Story Points", summary.missing_points),
            ("Severity", summary.missing_severity),
            ("Due Date", summary.missing_due_date),
            ("Description/ACs", summary.missing_description),
        ]

        for name, count in attrs:
            pct = (count / summary.total_tasks * 100) if summary.total_tasks > 0 else 0
            badge_class = "badge-success" if count == 0 else ("badge-danger" if pct > 30 else "badge-warning")
            status = "Complete" if count == 0 else f"{count} missing"
            html += f"""
            <tr>
                <td><strong>{name}</strong></td>
                <td>{count}</td>
                <td>{pct:.1f}%</td>
                <td><span class="badge {badge_class}">{status}</span></td>
            </tr>"""

        html += """
        </table>
    </div>

    <div class="card">
        <h2>By Assignee</h2>
        <table>
            <tr>
                <th>Assignee</th>
                <th>Total</th>
                <th>Issues</th>
                <th>Compliance</th>
            </tr>"""

        for assignee, data in summary.by_assignee.items():
            total = data["total"]
            issues = data["issues"]
            rate = ((total - issues) / total * 100) if total > 0 else 100
            badge_class = "badge-success" if rate >= 80 else ("badge-danger" if rate < 50 else "badge-warning")
            html += f"""
            <tr>
                <td><strong>{assignee}</strong></td>
                <td>{total}</td>
                <td>{issues}</td>
                <td><span class="badge {badge_class}">{rate:.0f}%</span></td>
            </tr>"""

        html += """
        </table>
    </div>"""

        # Missing Updates Section
        missing_updates = [t for t in results if t.missing_daily_update]
        if missing_updates:
            html += f"""
    <div class="card">
        <div class="section-header">
            <h2>Missing Daily Updates</h2>
            <span class="count-badge">{len(missing_updates)} tasks</span>
        </div>
        <table>
            <tr>
                <th>Task</th>
                <th>Assignee</th>
                <th>Status</th>
                <th>Last Update</th>
                <th>Link</th>
            </tr>"""

            for t in missing_updates:
                name = t.name[:50] if len(t.name) > 50 else t.name
                last_update = t.last_comment_date[:10] if t.last_comment_date else "Never"
                hours = f"{t.hours_since_update:.0f}h ago" if t.hours_since_update else "N/A"
                html += f"""
            <tr>
                <td>{name}</td>
                <td>{t.assignee}</td>
                <td><span class="badge badge-warning">{t.progress}</span></td>
                <td>{last_update} ({hours})</td>
                <td><a href="{t.url}" target="_blank">Open →</a></td>
            </tr>"""

            html += """
        </table>
    </div>"""

        # Missing Attributes Sections
        sections = [
            ("Missing Epic", [t for t in results if t.missing_epic]),
            ("Missing Sprint", [t for t in results if t.missing_sprint]),
            ("Missing Due Date", [t for t in results if t.missing_due_date]),
            ("Missing Description/ACs", [t for t in results if t.missing_description]),
        ]

        for title, tasks in sections:
            if tasks:
                html += f"""
    <div class="card">
        <div class="section-header">
            <h2>{title}</h2>
            <span class="count-badge">{len(tasks)} tasks</span>
        </div>
        <table>
            <tr>
                <th>Task</th>
                <th>Assignee</th>
                <th>Progress</th>
                <th>Link</th>
            </tr>"""

                for t in tasks[:20]:  # Limit to 20 per section
                    name = t.name[:50] if len(t.name) > 50 else t.name
                    html += f"""
            <tr>
                <td>{name}</td>
                <td>{t.assignee}</td>
                <td>{t.progress or 'None'}</td>
                <td><a href="{t.url}" target="_blank">Open →</a></td>
            </tr>"""

                if len(tasks) > 20:
                    html += f"""
            <tr><td colspan="4" style="text-align:center; color:#6b7280;">... and {len(tasks) - 20} more</td></tr>"""

                html += """
        </table>
    </div>"""

        html += """
</div>
</body>
</html>"""

        return html


class JSONReportGenerator:
    """Generates JSON compliance reports."""

    def __init__(self, config: Config):
        self.config = config

    def generate(self, results: list[TaskCompliance], summary: ReportSummary) -> str:
        """Generate JSON report."""
        report = {
            "report_date": summary.report_date,
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_tasks": summary.total_tasks,
                "compliant_tasks": summary.compliant_tasks,
                "compliance_rate": round(summary.compliance_rate, 2),
                "mandatory_missing": {
                    "epic": summary.missing_epic,
                    "sprint": summary.missing_sprint,
                    "type": summary.missing_type,
                    "story_points": summary.missing_points,
                    "severity": summary.missing_severity,
                    "due_date": summary.missing_due_date,
                    "description": summary.missing_description,
                },
                "daily_updates": {
                    "tasks_needing_updates": summary.tasks_needing_updates,
                    "tasks_missing_updates": summary.tasks_missing_updates,
                },
                "by_assignee": summary.by_assignee,
            },
            "tasks": {
                "missing_epic": [asdict(t) for t in results if t.missing_epic],
                "missing_sprint": [asdict(t) for t in results if t.missing_sprint],
                "missing_type": [asdict(t) for t in results if t.missing_type],
                "missing_points": [asdict(t) for t in results if t.missing_points],
                "missing_severity": [asdict(t) for t in results if t.missing_severity],
                "missing_due_date": [asdict(t) for t in results if t.missing_due_date],
                "missing_description": [asdict(t) for t in results if t.missing_description],
                "missing_daily_update": [asdict(t) for t in results if t.missing_daily_update],
            },
            "all_tasks": [asdict(t) for t in results],
        }

        return json.dumps(report, indent=2, default=str)


class ExcelReportGenerator:
    """Generates Excel compliance reports with multiple sheets."""

    def __init__(self, config: Config):
        self.config = config

        # Styles
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        self.danger_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        self.success_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        self.todo_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # Light blue for To Do
        self.active_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")  # Light yellow for active
        self.title_font = Font(bold=True, size=14, color="1A1A2E")
        self.subtitle_font = Font(bold=True, size=12, color="16213E")
        self.link_font = Font(color="667EEA", underline="single")
        self.thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

    def _style_header_row(self, ws, row: int, num_cols: int):
        """Apply header styling to a row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = min(cell_length, 50)  # Cap at 50 chars
                except (TypeError, AttributeError):
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = max(adjusted_width, 12)

    def _add_task_row(self, ws, row: int, task: TaskCompliance, columns: list[str]):
        """Add a task row with data and styling."""
        col_data = {
            'Task Name': task.name[:60] if len(task.name) > 60 else task.name,
            'Assignee': task.assignee,
            'Progress': task.progress or 'None',
            'Status': task.status_label,
            'Sprint': task.sprint or 'None',
            'Epic': task.epic or 'None',
            'Type': task.task_type or 'None',
            'Story Points': task.story_points or 'None',
            'Severity': task.severity or 'None',
            'Due Date': task.due_on or 'None',
            'Created': task.created_at[:10] if task.created_at else 'N/A',
            'Description Chars': task.description_length,
            'Last Comment By': task.last_comment_author or 'N/A',
            'Total Comments': task.total_comments,
            'Last Update': task.last_comment_date[:10] if task.last_comment_date else 'Never',
            'Hours Since Update': f"{task.hours_since_update:.0f}" if task.hours_since_update else 'N/A',
            'Compliance Score': f"{task.compliance_score}%",
            'Missing Fields': ', '.join(task.mandatory_missing) if task.mandatory_missing else 'None',
            'Rule Violations': ', '.join(task.rule_violations) if task.rule_violations else 'None',
            'Link': task.url,
            # Quick Wins columns
            'Days Until Due': task.days_until_due if task.days_until_due is not None else 'No date',
            'Task Age (Days)': task.task_age_days,
            'Overdue': 'Yes' if task.is_overdue else 'No',
            'Days Overdue': abs(task.days_until_due) if task.is_overdue and task.days_until_due is not None else 'N/A',
        }

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col_idx, value=col_data.get(col_name, ''))
            cell.border = self.thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Add hyperlink for Link column
            if col_name == 'Link':
                cell.hyperlink = task.url
                cell.font = self.link_font
                cell.value = "Open Task"

            # Color coding for compliance score
            if col_name == 'Compliance Score':
                score = task.compliance_score
                if score >= 80:
                    cell.fill = self.success_fill
                elif score >= 50:
                    cell.fill = self.warning_fill
                else:
                    cell.fill = self.danger_fill

            # Color coding for Progress/Status column
            if col_name in ('Progress', 'Status'):
                if task.is_todo:
                    cell.fill = self.todo_fill
                elif task.needs_daily_update:
                    cell.fill = self.active_fill

            # Color coding for overdue/due soon (Quick Wins)
            if col_name == 'Overdue' and task.is_overdue:
                cell.fill = self.danger_fill

            if col_name == 'Days Until Due' and task.days_until_due is not None:
                if task.days_until_due < 0:
                    cell.fill = self.danger_fill
                elif task.days_until_due <= 3:
                    cell.fill = self.warning_fill

            if col_name == 'Days Overdue' and task.is_overdue:
                cell.fill = self.danger_fill

    def generate(self, results: list[TaskCompliance], summary: ReportSummary) -> Workbook:
        """Generate Excel workbook with multiple sheets."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel reports. Install with: pip install openpyxl")

        wb = Workbook()

        # ===== Sheet 1: Summary =====
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Title
        ws_summary.cell(row=1, column=1, value="Asana Ticket Compliance Report").font = self.title_font
        ws_summary.cell(row=2, column=1, value=f"Report Date: {summary.report_date}").font = self.subtitle_font
        ws_summary.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # Overall stats
        ws_summary.cell(row=5, column=1, value="Overall Compliance").font = self.subtitle_font
        stats = [
            ("Total Tasks", summary.total_tasks),
            ("Compliant Tasks", summary.compliant_tasks),
            ("Compliance Rate", f"{summary.compliance_rate:.1f}%"),
        ]
        for idx, (label, value) in enumerate(stats, 6):
            ws_summary.cell(row=idx, column=1, value=label)
            ws_summary.cell(row=idx, column=2, value=value)

        # Mandatory attributes missing
        ws_summary.cell(row=10, column=1, value="Mandatory Attributes Missing/Invalid").font = self.subtitle_font
        attrs = [
            ("Epic", summary.missing_epic),
            ("Sprint", summary.missing_sprint),
            ("Type", summary.missing_type),
            ("Story Points (Missing)", summary.missing_points),
            ("Story Points (Invalid - not Fibonacci)", summary.invalid_points),
            ("Severity", summary.missing_severity),
            ("Due Date", summary.missing_due_date),
            ("Description/ACs", summary.missing_description),
            ("Rule Violations (Epic/Bug with Points)", getattr(summary, 'rule_violations', 0)),
        ]
        headers = ["Attribute", "Count", "% of Tasks"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=11, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        for idx, (attr, count) in enumerate(attrs, 12):
            pct = (count / summary.total_tasks * 100) if summary.total_tasks > 0 else 0
            ws_summary.cell(row=idx, column=1, value=attr)
            ws_summary.cell(row=idx, column=2, value=count)
            cell_pct = ws_summary.cell(row=idx, column=3, value=f"{pct:.1f}%")
            if pct > 30:
                cell_pct.fill = self.danger_fill
            elif pct > 10:
                cell_pct.fill = self.warning_fill
            else:
                cell_pct.fill = self.success_fill

        # Task Status Breakdown
        ws_summary.cell(row=21, column=1, value="Task Status Breakdown").font = self.subtitle_font
        ws_summary.cell(row=22, column=1, value="To Do Tasks")
        cell_todo = ws_summary.cell(row=22, column=2, value=summary.tasks_todo)
        cell_todo.fill = self.todo_fill
        ws_summary.cell(row=23, column=1, value="Active Tasks (In Progress/Review/QA)")
        cell_active = ws_summary.cell(row=23, column=2, value=summary.tasks_active)
        cell_active.fill = self.active_fill
        ws_summary.cell(row=24, column=1, value="Tasks Missing Daily Updates")
        cell_missing = ws_summary.cell(row=24, column=2, value=summary.tasks_missing_updates)
        if summary.tasks_missing_updates > 0:
            cell_missing.fill = self.danger_fill

        # By Assignee
        ws_summary.cell(row=27, column=1, value="Compliance by Assignee").font = self.subtitle_font
        assignee_headers = ["Assignee", "Total Tasks", "Compliant", "Issues", "Compliance Rate"]
        for col, header in enumerate(assignee_headers, 1):
            cell = ws_summary.cell(row=28, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        row = 29
        for assignee, data in summary.by_assignee.items():
            total = data["total"]
            issues = data["issues"]
            compliant = total - issues
            rate = (compliant / total * 100) if total > 0 else 100
            ws_summary.cell(row=row, column=1, value=assignee)
            ws_summary.cell(row=row, column=2, value=total)
            ws_summary.cell(row=row, column=3, value=compliant)
            ws_summary.cell(row=row, column=4, value=issues)
            cell_rate = ws_summary.cell(row=row, column=5, value=f"{rate:.0f}%")
            if rate >= 80:
                cell_rate.fill = self.success_fill
            elif rate >= 50:
                cell_rate.fill = self.warning_fill
            else:
                cell_rate.fill = self.danger_fill
            row += 1

        self._auto_adjust_columns(ws_summary)

        # ===== Sheet 2: All Tasks =====
        ws_all = wb.create_sheet("All Tasks")
        all_columns = ['Task Name', 'Assignee', 'Status', 'Sprint', 'Epic', 'Type',
                       'Story Points', 'Severity', 'Due Date', 'Days Until Due',
                       'Overdue', 'Task Age (Days)', 'Created',
                       'Description Chars', 'Last Comment By', 'Total Comments',
                       'Hours Since Update', 'Compliance Score', 'Missing Fields',
                       'Rule Violations', 'Link']

        for col, header in enumerate(all_columns, 1):
            ws_all.cell(row=1, column=col, value=header)
        self._style_header_row(ws_all, 1, len(all_columns))

        for row_idx, task in enumerate(results, 2):
            self._add_task_row(ws_all, row_idx, task, all_columns)

        self._auto_adjust_columns(ws_all)
        ws_all.freeze_panes = 'A2'  # Freeze header row

        # ===== Sheet 3: To Do Tasks =====
        todo_tasks = [t for t in results if t.is_todo]
        ws_todo = wb.create_sheet("To Do")
        todo_columns = ['Task Name', 'Assignee', 'Sprint', 'Epic', 'Type',
                        'Story Points', 'Due Date', 'Compliance Score', 'Missing Fields', 'Link']

        for col, header in enumerate(todo_columns, 1):
            ws_todo.cell(row=1, column=col, value=header)
        self._style_header_row(ws_todo, 1, len(todo_columns))

        for row_idx, task in enumerate(todo_tasks, 2):
            self._add_task_row(ws_todo, row_idx, task, todo_columns)

        self._auto_adjust_columns(ws_todo)
        ws_todo.freeze_panes = 'A2'

        # ===== Sheet 4: Missing Daily Updates =====
        missing_updates = [t for t in results if t.missing_daily_update]
        ws_updates = wb.create_sheet("Missing Daily Updates")
        update_columns = ['Task Name', 'Assignee', 'Progress', 'Last Update', 'Hours Since Update', 'Link']

        for col, header in enumerate(update_columns, 1):
            ws_updates.cell(row=1, column=col, value=header)
        self._style_header_row(ws_updates, 1, len(update_columns))

        for row_idx, task in enumerate(missing_updates, 2):
            self._add_task_row(ws_updates, row_idx, task, update_columns)

        self._auto_adjust_columns(ws_updates)
        ws_updates.freeze_panes = 'A2'

        # ===== Sheet 4: Missing Epic =====
        missing_epic = [t for t in results if t.missing_epic]
        ws_epic = wb.create_sheet("Missing Epic")
        epic_columns = ['Task Name', 'Assignee', 'Progress', 'Sprint', 'Due Date', 'Link']

        for col, header in enumerate(epic_columns, 1):
            ws_epic.cell(row=1, column=col, value=header)
        self._style_header_row(ws_epic, 1, len(epic_columns))

        for row_idx, task in enumerate(missing_epic, 2):
            self._add_task_row(ws_epic, row_idx, task, epic_columns)

        self._auto_adjust_columns(ws_epic)
        ws_epic.freeze_panes = 'A2'

        # ===== Sheet 5: Missing Sprint =====
        missing_sprint = [t for t in results if t.missing_sprint]
        ws_sprint = wb.create_sheet("Missing Sprint")
        sprint_columns = ['Task Name', 'Assignee', 'Progress', 'Epic', 'Due Date', 'Link']

        for col, header in enumerate(sprint_columns, 1):
            ws_sprint.cell(row=1, column=col, value=header)
        self._style_header_row(ws_sprint, 1, len(sprint_columns))

        for row_idx, task in enumerate(missing_sprint, 2):
            self._add_task_row(ws_sprint, row_idx, task, sprint_columns)

        self._auto_adjust_columns(ws_sprint)
        ws_sprint.freeze_panes = 'A2'

        # ===== Sheet 6: Missing Due Date =====
        missing_due = [t for t in results if t.missing_due_date]
        ws_due = wb.create_sheet("Missing Due Date")
        due_columns = ['Task Name', 'Assignee', 'Progress', 'Sprint', 'Epic', 'Link']

        for col, header in enumerate(due_columns, 1):
            ws_due.cell(row=1, column=col, value=header)
        self._style_header_row(ws_due, 1, len(due_columns))

        for row_idx, task in enumerate(missing_due, 2):
            self._add_task_row(ws_due, row_idx, task, due_columns)

        self._auto_adjust_columns(ws_due)
        ws_due.freeze_panes = 'A2'

        # ===== Sheet 7: Missing Description/ACs =====
        missing_desc = [t for t in results if t.missing_description]
        ws_desc = wb.create_sheet("Missing Description")
        desc_columns = ['Task Name', 'Assignee', 'Progress', 'Description Chars', 'Link']

        for col, header in enumerate(desc_columns, 1):
            ws_desc.cell(row=1, column=col, value=header)
        self._style_header_row(ws_desc, 1, len(desc_columns))

        for row_idx, task in enumerate(missing_desc, 2):
            self._add_task_row(ws_desc, row_idx, task, desc_columns)

        self._auto_adjust_columns(ws_desc)
        ws_desc.freeze_panes = 'A2'

        # ===== Sheet 8: Missing Story Points =====
        missing_points = [t for t in results if t.missing_points]
        ws_points = wb.create_sheet("Missing Story Points")
        points_columns = ['Task Name', 'Assignee', 'Progress', 'Type', 'Sprint', 'Link']

        for col, header in enumerate(points_columns, 1):
            ws_points.cell(row=1, column=col, value=header)
        self._style_header_row(ws_points, 1, len(points_columns))

        for row_idx, task in enumerate(missing_points, 2):
            self._add_task_row(ws_points, row_idx, task, points_columns)

        self._auto_adjust_columns(ws_points)
        ws_points.freeze_panes = 'A2'

        # ===== Sheet 9: Invalid Story Points (non-Fibonacci) =====
        invalid_points_tasks = [t for t in results if t.invalid_points]
        ws_invalid = wb.create_sheet("Invalid Story Points")
        invalid_columns = ['Task Name', 'Assignee', 'Story Points', 'Progress', 'Type', 'Sprint', 'Link']

        for col, header in enumerate(invalid_columns, 1):
            ws_invalid.cell(row=1, column=col, value=header)
        self._style_header_row(ws_invalid, 1, len(invalid_columns))

        for row_idx, task in enumerate(invalid_points_tasks, 2):
            self._add_task_row(ws_invalid, row_idx, task, invalid_columns)

        self._auto_adjust_columns(ws_invalid)
        ws_invalid.freeze_panes = 'A2'

        # ===== Sheet 10: Missing Severity =====
        missing_severity = [t for t in results if t.missing_severity]
        ws_severity = wb.create_sheet("Missing Severity")
        severity_columns = ['Task Name', 'Assignee', 'Progress', 'Type', 'Sprint', 'Link']

        for col, header in enumerate(severity_columns, 1):
            ws_severity.cell(row=1, column=col, value=header)
        self._style_header_row(ws_severity, 1, len(severity_columns))

        for row_idx, task in enumerate(missing_severity, 2):
            self._add_task_row(ws_severity, row_idx, task, severity_columns)

        self._auto_adjust_columns(ws_severity)
        ws_severity.freeze_panes = 'A2'

        # ===== Sheet 11: Missing Type =====
        missing_type = [t for t in results if t.missing_type]
        ws_type = wb.create_sheet("Missing Type")
        type_columns = ['Task Name', 'Assignee', 'Progress', 'Sprint', 'Epic', 'Link']

        for col, header in enumerate(type_columns, 1):
            ws_type.cell(row=1, column=col, value=header)
        self._style_header_row(ws_type, 1, len(type_columns))

        for row_idx, task in enumerate(missing_type, 2):
            self._add_task_row(ws_type, row_idx, task, type_columns)

        self._auto_adjust_columns(ws_type)
        ws_type.freeze_panes = 'A2'

        # ===== Sheet 12: Rule Violations =====
        rule_violations = [t for t in results if t.rule_violations]
        ws_violations = wb.create_sheet("Rule Violations")
        violations_columns = ['Task Name', 'Assignee', 'Type', 'Story Points', 'Progress', 'Rule Violations', 'Link']

        for col, header in enumerate(violations_columns, 1):
            ws_violations.cell(row=1, column=col, value=header)
        self._style_header_row(ws_violations, 1, len(violations_columns))

        for row_idx, task in enumerate(rule_violations, 2):
            self._add_task_row(ws_violations, row_idx, task, violations_columns)

        self._auto_adjust_columns(ws_violations)
        ws_violations.freeze_panes = 'A2'

        # ===== Sheet 13: Overdue Tasks (Quick Wins) =====
        overdue_tasks = [t for t in results if t.is_overdue]
        # Sort by most overdue first (most negative days_until_due)
        overdue_tasks.sort(key=lambda t: t.days_until_due if t.days_until_due is not None else 0)
        ws_overdue = wb.create_sheet("Overdue Tasks")
        overdue_columns = ['Task Name', 'Assignee', 'Due Date', 'Days Overdue',
                          'Story Points', 'Progress', 'Sprint', 'Link']

        for col, header in enumerate(overdue_columns, 1):
            ws_overdue.cell(row=1, column=col, value=header)
        self._style_header_row(ws_overdue, 1, len(overdue_columns))

        for row_idx, task in enumerate(overdue_tasks, 2):
            self._add_task_row(ws_overdue, row_idx, task, overdue_columns)

        self._auto_adjust_columns(ws_overdue)
        ws_overdue.freeze_panes = 'A2'

        # ===== Sheet 14: Due This Week (Quick Wins) =====
        due_soon_tasks = [
            t for t in results
            if t.days_until_due is not None
            and 0 <= t.days_until_due <= 7
            and t.progress != "Done"
        ]
        # Sort by due date (soonest first)
        due_soon_tasks.sort(key=lambda t: t.days_until_due if t.days_until_due is not None else 999)
        ws_due_soon = wb.create_sheet("Due This Week")
        due_soon_columns = ['Task Name', 'Assignee', 'Due Date', 'Days Until Due',
                           'Story Points', 'Progress', 'Sprint', 'Link']

        for col, header in enumerate(due_soon_columns, 1):
            ws_due_soon.cell(row=1, column=col, value=header)
        self._style_header_row(ws_due_soon, 1, len(due_soon_columns))

        for row_idx, task in enumerate(due_soon_tasks, 2):
            self._add_task_row(ws_due_soon, row_idx, task, due_soon_columns)

        self._auto_adjust_columns(ws_due_soon)
        ws_due_soon.freeze_panes = 'A2'

        return wb

    def _is_invalid_story_points(self, task: TaskCompliance) -> tuple[bool, str]:
        """Check if a task has invalid story points and return reason.

        Returns:
            tuple of (is_invalid, reason)
        """
        if not task.story_points:
            return False, ""

        try:
            points = float(task.story_points)
        except (ValueError, TypeError):
            return True, "Non-numeric value"

        # Bug or Epic with story points
        if task.task_type in self.config.types_without_points and points > 0:
            return True, f"{task.task_type} should not have points"

        # Non-Fibonacci number
        if points != int(points) or int(points) not in self.config.valid_story_points:
            return True, f"Non-Fibonacci ({task.story_points})"

        return False, ""

    def generate_with_completed(
        self,
        results: list[TaskCompliance],
        completed_results: list[TaskCompliance],
        summary: ReportSummary
    ) -> 'Workbook':
        """Generate Excel workbook including completed tasks for invalid points analysis."""
        # First generate the standard report
        wb = self.generate(results, summary)

        # Now add sheet for ALL invalid story points (including completed tasks)
        all_tasks = results + completed_results

        invalid_tasks = []
        for task in all_tasks:
            is_invalid, reason = self._is_invalid_story_points(task)
            if is_invalid:
                invalid_tasks.append((task, reason))

        # Sort by assignee then by points
        invalid_tasks.sort(key=lambda x: (x[0].assignee or "ZZZ", -(float(x[0].story_points or 0))))

        ws_invalid = wb.create_sheet("Invalid Story Points")
        invalid_columns = ['Task Name', 'Assignee', 'Type', 'Story Points', 'Issue',
                          'Progress', 'Sprint', 'Link']

        for col, header in enumerate(invalid_columns, 1):
            ws_invalid.cell(row=1, column=col, value=header)
        self._style_header_row(ws_invalid, 1, len(invalid_columns))

        for row_idx, (task, reason) in enumerate(invalid_tasks, 2):
            # Custom row handling to include the reason
            col_data = {
                'Task Name': task.name[:60] if len(task.name) > 60 else task.name,
                'Assignee': task.assignee or 'Unassigned',
                'Type': task.task_type or 'None',
                'Story Points': task.story_points or 'None',
                'Issue': reason,
                'Progress': task.progress or 'None',
                'Sprint': task.sprint or 'None',
                'Link': task.url,
            }

            for col_idx, col_name in enumerate(invalid_columns, 1):
                cell = ws_invalid.cell(row=row_idx, column=col_idx, value=col_data.get(col_name, ''))
                cell.border = self.thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # Highlight the issue column in red
                if col_name == 'Issue':
                    cell.fill = self.danger_fill

                # Add hyperlink for Link column
                if col_name == 'Link':
                    cell.hyperlink = task.url
                    cell.font = self.link_font
                    cell.value = "Open Task"

        self._auto_adjust_columns(ws_invalid)
        ws_invalid.freeze_panes = 'A2'

        # Add summary row at the top
        if invalid_tasks:
            # Calculate totals by assignee
            assignee_invalid_points = {}
            for task, _ in invalid_tasks:
                assignee = task.assignee or "Unassigned"
                try:
                    points = float(task.story_points) if task.story_points else 0
                except (ValueError, TypeError):
                    points = 0
                assignee_invalid_points[assignee] = assignee_invalid_points.get(assignee, 0) + points

            # Add a summary sheet for invalid points by assignee
            ws_invalid_summary = wb.create_sheet("Invalid Points Summary")
            summary_columns = ['Assignee', 'Invalid Points', 'Task Count']

            for col, header in enumerate(summary_columns, 1):
                ws_invalid_summary.cell(row=1, column=col, value=header)
            self._style_header_row(ws_invalid_summary, 1, len(summary_columns))

            # Count tasks per assignee
            assignee_task_count = {}
            for task, _ in invalid_tasks:
                assignee = task.assignee or "Unassigned"
                assignee_task_count[assignee] = assignee_task_count.get(assignee, 0) + 1

            row_idx = 2
            for assignee in sorted(assignee_invalid_points.keys(), key=lambda a: -assignee_invalid_points[a]):
                ws_invalid_summary.cell(row=row_idx, column=1, value=assignee)
                cell_points = ws_invalid_summary.cell(row=row_idx, column=2, value=assignee_invalid_points[assignee])
                cell_points.fill = self.danger_fill
                ws_invalid_summary.cell(row=row_idx, column=3, value=assignee_task_count[assignee])
                row_idx += 1

            # Add total row
            total_invalid_points = sum(assignee_invalid_points.values())
            total_invalid_tasks = len(invalid_tasks)
            ws_invalid_summary.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
            ws_invalid_summary.cell(row=row_idx, column=2, value=total_invalid_points).font = Font(bold=True)
            ws_invalid_summary.cell(row=row_idx, column=3, value=total_invalid_tasks).font = Font(bold=True)

            self._auto_adjust_columns(ws_invalid_summary)

        return wb


# =============================================================================
# Main Application
# =============================================================================

class AsanaComplianceReporter:
    """Main application."""

    def __init__(self, access_token: str, config: Optional[Config] = None):
        self.config = config or Config()
        self.client = AsanaClient(access_token, self.config)
        self.analyzer = ComplianceAnalyzer(self.config, self.client)

        self.generators = {
            'markdown': MarkdownReportGenerator(self.config),
            'html': HTMLReportGenerator(self.config),
            'json': JSONReportGenerator(self.config),
            'excel': ExcelReportGenerator(self.config),
        }

        # Store results for Excel generation (since Excel returns workbook, not string)
        self._last_results: list[TaskCompliance] = []
        self._last_summary: Optional[ReportSummary] = None

    def run(self, output_format: str = 'markdown', fetch_comments: bool = True):
        """Run the compliance report.

        Returns:
            For text formats (markdown, html, json): tuple[str, ReportSummary]
            For excel: tuple[Workbook, ReportSummary]
        """
        print("Fetching tasks from Asana...")
        tasks = self.client.get_tasks(completed=False)
        print(f"   Found {len(tasks)} incomplete tasks")

        print("Analyzing compliance...")
        results = self.analyzer.analyze_all(tasks, fetch_comments=fetch_comments)
        print(f"   Analyzed {len(results)} tasks (excluding Done)")

        print("Generating summary...")
        summary = self.analyzer.generate_summary(results)

        # Store for potential reuse
        self._last_results = results
        self._last_summary = summary

        print(f"Generating {output_format} report...")
        generator = self.generators.get(output_format, self.generators['markdown'])
        report = generator.generate(results, summary)

        return report, summary

    def save_report(self, report, output_format: str, custom_path: Optional[str] = None) -> Path:
        """Save report to file.

        Args:
            report: The report content (str for text formats, Workbook for excel)
            output_format: The format of the report
            custom_path: Optional custom file path for the report (date will be appended)
        """
        date_str = datetime.now().strftime('%Y-%m-%d')

        if custom_path:
            # Insert date before file extension
            custom_path_obj = Path(custom_path)
            stem = custom_path_obj.stem
            ext = custom_path_obj.suffix
            filename_with_date = f"{stem}_{date_str}{ext}"
            filepath = custom_path_obj.parent / filename_with_date
            filepath.parent.mkdir(parents=True, exist_ok=True)
        else:
            output_dir = Path(self.config.output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            extensions = {'markdown': 'md', 'html': 'html', 'json': 'json', 'excel': 'xlsx'}
            ext = extensions.get(output_format, 'md')
            filename = f"compliance_report_{date_str}.{ext}"
            filepath = output_dir / filename

        if output_format == 'excel':
            # Excel workbook - save using openpyxl
            report.save(filepath)
        else:
            # Text-based formats
            filepath.write_text(report)

        print(f"Report saved: {filepath}")
        return filepath


def print_console_summary(summary: ReportSummary):
    """Print summary to console."""
    if RICH_AVAILABLE:
        console = Console()

        table = Table(title=f"Compliance Summary - {summary.report_date}")
        table.add_column("Metric", style="cyan")
        table.add_column("Value", style="magenta")

        table.add_row("Report Date", summary.report_date)
        table.add_row("Total Tasks", str(summary.total_tasks))
        table.add_row("Compliant", str(summary.compliant_tasks))
        table.add_row("Compliance Rate", f"{summary.compliance_rate:.1f}%")
        table.add_row("─" * 20, "─" * 10)
        table.add_row("Missing Epic", str(summary.missing_epic))
        table.add_row("Missing Sprint", str(summary.missing_sprint))
        table.add_row("Missing Type", str(summary.missing_type))
        table.add_row("Missing Points", str(summary.missing_points))
        table.add_row("Missing Severity", str(summary.missing_severity))
        table.add_row("Missing Due Date", str(summary.missing_due_date))
        table.add_row("Missing Description", str(summary.missing_description))
        table.add_row("─" * 20, "─" * 10)
        table.add_row("Missing Daily Updates", str(summary.tasks_missing_updates))

        console.print(table)
    else:
        print("\n" + "=" * 50)
        print(f"COMPLIANCE SUMMARY - {summary.report_date}")
        print("=" * 50)
        print(f"Report Date: {summary.report_date}")
        print(f"Total Tasks: {summary.total_tasks}")
        print(f"Compliant: {summary.compliant_tasks}")
        print(f"Compliance Rate: {summary.compliance_rate:.1f}%")
        print("-" * 50)
        print(f"Missing Epic: {summary.missing_epic}")
        print(f"Missing Sprint: {summary.missing_sprint}")
        print(f"Missing Type: {summary.missing_type}")
        print(f"Missing Points: {summary.missing_points}")
        print(f"Missing Severity: {summary.missing_severity}")
        print(f"Missing Due Date: {summary.missing_due_date}")
        print(f"Missing Description: {summary.missing_description}")
        print("-" * 50)
        print(f"Missing Daily Updates: {summary.tasks_missing_updates}")
        print("=" * 50)


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Generate Asana ticket compliance report"
    )
    parser.add_argument('--output-dir', '-o', default='./reports')
    parser.add_argument('--format', '-f', choices=['markdown', 'html', 'json', 'excel', 'all'], default='markdown')
    parser.add_argument('--excel-path', '-e', help='Custom path for Excel output (e.g., /path/to/report.xlsx)')
    parser.add_argument('--token', '-t', help='Asana access token')
    parser.add_argument('--no-save', action='store_true')
    parser.add_argument('--no-comments', action='store_true', help='Skip fetching comments (faster)')
    parser.add_argument('--quiet', '-q', action='store_true')

    args = parser.parse_args()

    access_token = args.token or os.environ.get('ASANA_ACCESS_TOKEN')
    if not access_token:
        print("Error: ASANA_ACCESS_TOKEN required")
        sys.exit(1)

    # Check openpyxl availability for Excel format
    if (args.format == 'excel' or args.format == 'all') and not OPENPYXL_AVAILABLE:
        print("Error: openpyxl is required for Excel reports. Install with: pip install openpyxl")
        sys.exit(1)

    config = Config(output_dir=args.output_dir)
    reporter = AsanaComplianceReporter(access_token, config)

    formats = ['markdown', 'html', 'json', 'excel'] if args.format == 'all' else [args.format]

    for fmt in formats:
        report, summary = reporter.run(
            output_format=fmt,
            fetch_comments=not args.no_comments
        )

        if args.no_save:
            if fmt != 'excel':
                print(report)
            else:
                print("Note: Excel format cannot be printed to console, skipping...")
        else:
            # Use custom path for Excel if provided
            custom_path = args.excel_path if fmt == 'excel' and args.excel_path else None
            reporter.save_report(report, fmt, custom_path=custom_path)

    if not args.quiet:
        print_console_summary(summary)


if __name__ == '__main__':
    main()
