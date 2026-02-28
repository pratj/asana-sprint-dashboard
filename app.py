"""
Streamlit Web App for Asana Sprint Dashboard
==============================================
A focused dashboard for sprint compliance and burndown tracking.

Run locally:
    cd scripts && streamlit run app.py
"""
from __future__ import annotations

import io
import os
import re
from datetime import datetime, timedelta
from typing import Optional

import streamlit as st
import pandas as pd

# Import Plotly for interactive charts
try:
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

# Import the core report logic
from asana_daily_report import (
    Config,
    AsanaComplianceReporter,
    TaskCompliance,
    ReportSummary,
    MarkdownReportGenerator,
    JSONReportGenerator,
    OPENPYXL_AVAILABLE,
)

# =============================================================================
# Page Configuration
# =============================================================================

st.set_page_config(
    page_title="SourceHub - Sprint Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Neumorphism Design System CSS
st.markdown("""
<style>
    /* =================================================================
       DESIGN TOKENS - Neumorphism (Soft UI) System
       ================================================================= */
    :root {
        /* Base Colors */
        --nm-bg: #E4E8EC;
        --nm-surface: #E4E8EC;
        --nm-shadow-dark: #A3B1C6;
        --nm-shadow-light: #FFFFFF;

        /* Semantic Accents (Muted) */
        --nm-primary: #6B7FD7;
        --nm-success: #5B9A8B;
        --nm-warning: #D4A574;
        --nm-error: #C9736D;
        --nm-info: #5A9AA8;

        /* Text Colors */
        --nm-text-primary: #2D3748;
        --nm-text-secondary: #5A6778;
        --nm-text-muted: #8896A4;

        /* Shadow Patterns */
        --nm-shadow-raised: 6px 6px 12px #A3B1C6, -6px -6px 12px #FFFFFF;
        --nm-shadow-inset: inset 3px 3px 6px #A3B1C6, inset -3px -3px 6px #FFFFFF;
        --nm-shadow-pressed: inset 2px 2px 5px #A3B1C6, inset -2px -2px 5px #FFFFFF;
        --nm-shadow-hover: 10px 10px 20px #A3B1C6, -10px -10px 20px #FFFFFF;
    }

    /* =================================================================
       GLOBAL STYLES
       ================================================================= */
    .stApp {
        background: var(--nm-bg) !important;
    }

    [data-testid="stSidebar"] {
        background: #D8DCE2 !important;
    }

    [data-testid="stSidebar"] [data-testid="stMarkdown"] {
        color: var(--nm-text-primary);
    }

    /* =================================================================
       METRIC CARDS - Neumorphic Style
       ================================================================= */
    .nm-card {
        background: var(--nm-bg);
        border-radius: 16px;
        padding: 24px 20px 20px 20px;
        text-align: center;
        margin-bottom: 1rem;
        box-shadow: var(--nm-shadow-raised);
        position: relative;
        overflow: hidden;
        transition: box-shadow 0.25s ease;
    }

    .nm-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: var(--nm-primary);
        border-radius: 16px 16px 0 0;
    }

    .nm-card:hover {
        box-shadow: var(--nm-shadow-hover);
    }

    .nm-card--success::before { background: var(--nm-success); }
    .nm-card--warning::before { background: var(--nm-error); }
    .nm-card--info::before { background: var(--nm-info); }

    .nm-card-value {
        font-size: 2.25rem;
        font-weight: 700;
        color: var(--nm-text-primary);
        margin: 0;
        line-height: 1.2;
    }

    .nm-card-label {
        font-size: 0.9rem;
        color: var(--nm-text-secondary);
        margin-top: 8px;
        font-weight: 500;
    }

    /* =================================================================
       ALERT SECTIONS - Neumorphic Style with Colored Backgrounds
       ================================================================= */
    .nm-alert {
        background: var(--nm-bg);
        border-radius: 16px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        box-shadow: var(--nm-shadow-raised);
        border-left: 5px solid var(--nm-info);
        position: relative;
        overflow: hidden;
    }

    /* Critical/Error Alert - Soft rose/coral tint */
    .nm-alert--error {
        background: linear-gradient(135deg, #F0E4E4 0%, #E8DCDC 100%);
        border-left-color: var(--nm-error);
        box-shadow:
            6px 6px 12px rgba(163, 145, 145, 0.5),
            -6px -6px 12px rgba(255, 255, 255, 0.8),
            inset 0 1px 0 rgba(255, 255, 255, 0.6);
    }

    .nm-alert--error::before {
        content: '';
        position: absolute;
        top: 0;
        right: 0;
        width: 100px;
        height: 100px;
        background: radial-gradient(circle at top right, rgba(201, 115, 109, 0.15), transparent 70%);
        pointer-events: none;
    }

    .nm-alert--error h3 {
        color: #8B4C47;
    }

    .nm-alert--error p {
        color: #6B5A58;
    }

    /* Warning/Amber Alert - Soft warm amber tint */
    .nm-alert--warning {
        background: linear-gradient(135deg, #F2EBE0 0%, #EAE2D6 100%);
        border-left-color: var(--nm-warning);
        box-shadow:
            6px 6px 12px rgba(163, 155, 140, 0.5),
            -6px -6px 12px rgba(255, 255, 255, 0.8),
            inset 0 1px 0 rgba(255, 255, 255, 0.6);
    }

    .nm-alert--warning::before {
        content: '';
        position: absolute;
        top: 0;
        right: 0;
        width: 100px;
        height: 100px;
        background: radial-gradient(circle at top right, rgba(212, 165, 116, 0.15), transparent 70%);
        pointer-events: none;
    }

    .nm-alert--warning h3 {
        color: #7A6340;
    }

    .nm-alert--warning p {
        color: #6B6055;
    }

    .nm-alert h3 {
        color: var(--nm-text-primary);
        margin: 0 0 0.5rem 0;
        font-weight: 600;
        font-size: 1.1rem;
    }

    .nm-alert p {
        color: var(--nm-text-secondary);
        margin: 0 0 1rem 0;
        font-size: 0.9rem;
    }

    /* =================================================================
       COMPLIANCE DETAILS SECTION - Soft blue/purple tint
       ================================================================= */
    .nm-section-compliance {
        background: linear-gradient(135deg, #E4E8F0 0%, #DCE2EC 100%);
        border-radius: 16px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        box-shadow:
            6px 6px 12px rgba(140, 155, 180, 0.4),
            -6px -6px 12px rgba(255, 255, 255, 0.8),
            inset 0 1px 0 rgba(255, 255, 255, 0.6);
        border-left: 5px solid var(--nm-primary);
        position: relative;
        overflow: hidden;
    }

    .nm-section-compliance::before {
        content: '';
        position: absolute;
        top: 0;
        right: 0;
        width: 120px;
        height: 120px;
        background: radial-gradient(circle at top right, rgba(107, 127, 215, 0.12), transparent 70%);
        pointer-events: none;
    }

    .nm-section-compliance h3 {
        color: #4A5580;
        margin: 0 0 0.5rem 0;
        font-weight: 600;
        font-size: 1.1rem;
    }

    .nm-section-compliance p {
        color: #5A6778;
        margin: 0;
        font-size: 0.9rem;
    }

    /* =================================================================
       BUTTONS - Neumorphic Style
       ================================================================= */
    .stButton > button {
        background: var(--nm-bg) !important;
        border: none !important;
        border-radius: 10px !important;
        box-shadow: var(--nm-shadow-raised) !important;
        color: var(--nm-text-primary) !important;
        font-weight: 500 !important;
        transition: all 0.15s ease !important;
    }

    .stButton > button:hover {
        box-shadow: var(--nm-shadow-hover) !important;
        color: var(--nm-primary) !important;
    }

    .stButton > button:active {
        box-shadow: var(--nm-shadow-pressed) !important;
    }

    .stButton > button[kind="primary"] {
        background: var(--nm-bg) !important;
        color: var(--nm-primary) !important;
    }

    .stButton > button[kind="primary"]::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        height: 3px;
        background: var(--nm-primary);
        border-radius: 10px 10px 0 0;
    }

    /* =================================================================
       INPUTS - Neumorphic Inset Style
       ================================================================= */
    .stTextInput > div > div > input,
    .stSelectbox > div > div,
    .stMultiSelect > div > div,
    .stNumberInput > div > div > input {
        background: var(--nm-bg) !important;
        border: none !important;
        border-radius: 8px !important;
        box-shadow: var(--nm-shadow-inset) !important;
        color: var(--nm-text-primary) !important;
    }

    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {
        box-shadow: var(--nm-shadow-inset), 0 0 0 3px rgba(107, 127, 215, 0.3) !important;
    }

    /* Dropdown/Select cursor pointer */
    .stSelectbox > div > div,
    .stSelectbox [data-baseweb="select"],
    .stSelectbox [data-baseweb="select"] > div,
    .stMultiSelect > div > div,
    .stMultiSelect [data-baseweb="select"],
    .stMultiSelect [data-baseweb="select"] > div {
        cursor: pointer !important;
    }

    /* =================================================================
       EXPANDERS - Clean borderless style
       ================================================================= */
    div[data-testid="stExpander"] {
        background: var(--nm-bg) !important;
        border: none !important;
        border-radius: 12px !important;
        box-shadow: none !important;
        overflow: hidden;
    }

    div[data-testid="stExpander"] > details {
        border: none !important;
    }

    div[data-testid="stExpander"] > details > summary {
        background: transparent !important;
        color: var(--nm-text-primary) !important;
        font-weight: 500;
        border: none !important;
    }

    div[data-testid="stExpander"] > details[open] > summary {
        border: none !important;
        border-bottom: none !important;
    }

    /* Remove any outline/border on expander focus */
    div[data-testid="stExpander"] *:focus {
        outline: none !important;
        box-shadow: none !important;
    }

    /* =================================================================
       SEVERITY-COLORED EXPANDER WRAPPERS
       ================================================================= */

    /* Critical/Red severity - soft coral/rose */
    .nm-expander-red {
        background: linear-gradient(135deg, #F0E4E4 0%, #E8DCDC 100%);
        border-radius: 14px;
        padding: 4px;
        margin-bottom: 12px;
        box-shadow:
            5px 5px 10px rgba(163, 145, 145, 0.4),
            -5px -5px 10px rgba(255, 255, 255, 0.7),
            inset 0 1px 0 rgba(255, 255, 255, 0.5);
        border-left: 4px solid var(--nm-error);
    }

    .nm-expander-red div[data-testid="stExpander"] {
        background: transparent !important;
    }

    .nm-expander-red div[data-testid="stExpander"] > details > summary {
        color: #8B4C47 !important;
    }

    /* Warning/Orange severity - soft peach/orange */
    .nm-expander-orange {
        background: linear-gradient(135deg, #F5EBE0 0%, #EDE3D6 100%);
        border-radius: 14px;
        padding: 4px;
        margin-bottom: 12px;
        box-shadow:
            5px 5px 10px rgba(170, 155, 140, 0.4),
            -5px -5px 10px rgba(255, 255, 255, 0.7),
            inset 0 1px 0 rgba(255, 255, 255, 0.5);
        border-left: 4px solid #D4885C;
    }

    .nm-expander-orange div[data-testid="stExpander"] {
        background: transparent !important;
    }

    .nm-expander-orange div[data-testid="stExpander"] > details > summary {
        color: #8B5A3C !important;
    }

    /* Caution/Yellow severity - soft cream/yellow */
    .nm-expander-yellow {
        background: linear-gradient(135deg, #F5F0E0 0%, #EDE8D4 100%);
        border-radius: 14px;
        padding: 4px;
        margin-bottom: 12px;
        box-shadow:
            5px 5px 10px rgba(170, 165, 140, 0.4),
            -5px -5px 10px rgba(255, 255, 255, 0.7),
            inset 0 1px 0 rgba(255, 255, 255, 0.5);
        border-left: 4px solid #C9A84C;
    }

    .nm-expander-yellow div[data-testid="stExpander"] {
        background: transparent !important;
    }

    .nm-expander-yellow div[data-testid="stExpander"] > details > summary {
        color: #7A6830 !important;
    }

    /* =================================================================
       DATA ROWS - Neumorphic Style
       ================================================================= */
    .nm-data-row {
        background: var(--nm-bg);
        border-radius: 12px;
        padding: 16px;
        margin-bottom: 12px;
        box-shadow: var(--nm-shadow-raised);
        transition: box-shadow 0.25s ease;
    }

    .nm-data-row:hover {
        box-shadow: var(--nm-shadow-hover);
    }

    /* =================================================================
       SIDEBAR SECTIONS
       ================================================================= */
    .nm-sidebar-section {
        background: var(--nm-bg);
        border-radius: 12px;
        padding: 16px;
        margin-bottom: 16px;
        box-shadow: var(--nm-shadow-raised);
    }

    /* =================================================================
       LOADING & UTILITY STYLES
       ================================================================= */
    .loading-text {
        font-size: 1.1rem;
        color: var(--nm-primary);
        padding: 1rem;
    }

    /* Status indicators */
    [data-testid="stStatus"] {
        background: var(--nm-bg) !important;
        border-radius: 12px !important;
        box-shadow: var(--nm-shadow-raised) !important;
        border: none !important;
    }

    /* Metrics */
    [data-testid="stMetric"] {
        background: var(--nm-bg);
        border-radius: 12px;
        padding: 16px;
        box-shadow: var(--nm-shadow-raised);
    }

    [data-testid="stMetric"] label {
        color: var(--nm-text-secondary) !important;
    }

    [data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: var(--nm-text-primary) !important;
    }

    /* Dividers */
    hr {
        border-color: rgba(163, 177, 198, 0.3) !important;
    }

    /* Links */
    a {
        color: var(--nm-primary) !important;
    }

    a:hover {
        color: var(--nm-info) !important;
    }

    /* Captions */
    .stCaption, [data-testid="stCaptionContainer"] {
        color: var(--nm-text-muted) !important;
    }

    /* Headers */
    h1, h2, h3 {
        color: var(--nm-text-primary) !important;
    }

    /* Dialogs/Modals */
    [data-testid="stModal"] > div {
        background: var(--nm-bg) !important;
        border-radius: 16px !important;
        box-shadow: 12px 12px 24px #A3B1C6, -12px -12px 24px #FFFFFF !important;
    }

    /* Download buttons */
    .stDownloadButton > button {
        background: var(--nm-bg) !important;
        border: none !important;
        border-radius: 10px !important;
        box-shadow: var(--nm-shadow-raised) !important;
        color: var(--nm-text-primary) !important;
    }

    .stDownloadButton > button:hover {
        box-shadow: var(--nm-shadow-hover) !important;
        color: var(--nm-primary) !important;
    }

    /* Checkbox and Radio */
    .stCheckbox > label > span,
    .stRadio > label > span {
        color: var(--nm-text-primary) !important;
    }

    /* Info/Warning/Error boxes */
    .stAlert {
        background: var(--nm-bg) !important;
        border-radius: 12px !important;
        box-shadow: var(--nm-shadow-raised) !important;
        border-left: 4px solid var(--nm-info) !important;
    }

    /* Text area */
    .stTextArea > div > div > textarea {
        background: var(--nm-bg) !important;
        border: none !important;
        border-radius: 8px !important;
        box-shadow: var(--nm-shadow-inset) !important;
        color: var(--nm-text-primary) !important;
    }

    /* Dataframe - remove outer border, keep internal grid */
    [data-testid="stDataFrame"] {
        background: var(--nm-bg) !important;
        border-radius: 0 !important;
        box-shadow: none !important;
        border: none !important;
        overflow: visible;
    }

    [data-testid="stDataFrame"] > div {
        border: none !important;
        box-shadow: none !important;
    }

    /* Remove outer border from table container */
    [data-testid="stDataFrame"] iframe {
        border: none !important;
    }

    /* Style the table inside dataframe */
    [data-testid="stDataFrame"] table {
        border-collapse: collapse !important;
        border: none !important;
    }

    [data-testid="stDataFrame"] th,
    [data-testid="stDataFrame"] td {
        border-left: none !important;
        border-right: none !important;
        border-top: 1px solid rgba(163, 177, 198, 0.3) !important;
        border-bottom: 1px solid rgba(163, 177, 198, 0.3) !important;
    }

    [data-testid="stDataFrame"] tr:first-child th,
    [data-testid="stDataFrame"] tr:first-child td {
        border-top: none !important;
    }

    [data-testid="stDataFrame"] tr:last-child th,
    [data-testid="stDataFrame"] tr:last-child td {
        border-bottom: none !important;
    }

    /* Accessibility - Focus states */
    *:focus-visible {
        outline: 3px solid var(--nm-primary) !important;
        outline-offset: 2px;
    }

    /* Reduced motion preference */
    @media (prefers-reduced-motion: reduce) {
        *, *::before, *::after {
            transition: none !important;
            animation: none !important;
        }
    }

    /* =================================================================
       SPRINT PROGRESS BAR - Neumorphic Style (Quick Wins)
       ================================================================= */
    .nm-progress-container {
        background: var(--nm-bg);
        border-radius: 20px;
        padding: 24px;
        margin-bottom: 1.5rem;
        box-shadow: var(--nm-shadow-raised);
    }

    .nm-progress-bar-outer {
        background: var(--nm-bg);
        border-radius: 12px;
        height: 24px;
        box-shadow: var(--nm-shadow-inset);
        overflow: hidden;
        position: relative;
    }

    .nm-progress-bar-inner {
        height: 100%;
        border-radius: 12px;
        background: linear-gradient(90deg, var(--nm-primary) 0%, var(--nm-success) 100%);
        box-shadow: 0 2px 8px rgba(107, 127, 215, 0.4);
        transition: width 0.6s ease;
    }

    .nm-progress-text {
        position: absolute;
        top: 50%;
        left: 50%;
        transform: translate(-50%, -50%);
        font-weight: 600;
        font-size: 0.85rem;
        color: var(--nm-text-primary);
        text-shadow: 0 1px 2px rgba(255,255,255,0.8);
    }

    .nm-progress-stats {
        display: flex;
        justify-content: space-between;
        margin-top: 12px;
        font-size: 0.9rem;
        color: var(--nm-text-secondary);
    }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# Session State
# =============================================================================

def init_session_state():
    """Initialize session state variables."""
    defaults = {
        "authenticated": False,
        "auth_failed": False,
        "results": None,
        "completed_results": None,
        "summary": None,
        "config": None,
        "reporter": None,
        "report_generated": False,
        "is_generating": False,
        "selected_task_gid": None,
        "selected_task_url": None,
        "selected_task_name": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


# =============================================================================
# Authentication
# =============================================================================

def get_app_passcode() -> Optional[str]:
    """Get the app passcode from secrets or environment."""
    try:
        if "APP_PASSCODE" in st.secrets:
            return st.secrets["APP_PASSCODE"]
    except FileNotFoundError:
        pass
    return os.environ.get("APP_PASSCODE")


def check_passcode(entered_passcode: str) -> bool:
    """Check if the entered passcode is correct."""
    correct_passcode = get_app_passcode()
    if not correct_passcode:
        # No passcode configured - allow access
        return True
    return entered_passcode == correct_passcode


def render_login_screen():
    """Render a beautiful neumorphic login screen."""
    # Center the login form
    st.markdown("""
    <style>
        /* Login page specific styles */
        .login-container {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            min-height: 70vh;
            padding: 20px;
        }

        .login-card {
            background: var(--nm-bg, #E4E8EC);
            border-radius: 24px;
            padding: 48px 40px;
            box-shadow:
                12px 12px 24px #A3B1C6,
                -12px -12px 24px #FFFFFF;
            text-align: center;
            max-width: 400px;
            width: 100%;
        }

        .login-logo {
            font-size: 3.5rem;
            margin-bottom: 8px;
        }

        .login-title {
            font-size: 1.8rem;
            font-weight: 700;
            color: #2D3748;
            margin: 0 0 8px 0;
        }

        .login-subtitle {
            font-size: 0.95rem;
            color: #5A6778;
            margin: 0 0 32px 0;
        }

        .login-error {
            background: linear-gradient(135deg, #F0E4E4 0%, #E8DCDC 100%);
            border-radius: 12px;
            padding: 12px 16px;
            margin-bottom: 20px;
            border-left: 4px solid #C9736D;
        }

        .login-error p {
            color: #8B4C47;
            margin: 0;
            font-size: 0.9rem;
        }

        .login-footer {
            margin-top: 24px;
            font-size: 0.8rem;
            color: #8896A4;
        }

        /* Style the input field */
        .login-card .stTextInput > div > div > input {
            text-align: center;
            font-size: 1.2rem;
            letter-spacing: 8px;
            padding: 16px !important;
        }

        /* Style the button */
        .login-card .stButton > button {
            width: 100%;
            padding: 12px 24px !important;
            font-size: 1rem !important;
            font-weight: 600 !important;
            margin-top: 8px;
        }
    </style>
    """, unsafe_allow_html=True)

    # Create centered layout
    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        st.markdown("""
        <div class="login-container">
            <div class="login-card">
                <div class="login-logo">🔐</div>
                <h1 class="login-title">Sprint Dashboard</h1>
                <p class="login-subtitle">Enter passcode to continue</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Show error message if authentication failed
        if st.session_state.get("auth_failed"):
            st.markdown("""
            <div class="login-error">
                <p>Incorrect passcode. Please try again.</p>
            </div>
            """, unsafe_allow_html=True)

        # Use a form to ensure atomic submission of passcode
        with st.form("login_form", clear_on_submit=False):
            # Passcode input
            passcode = st.text_input(
                "Passcode",
                type="password",
                placeholder="••••••",
                label_visibility="collapsed",
                key="passcode_input"
            )

            # Login button - form submission ensures passcode value is committed
            submitted = st.form_submit_button("Unlock", type="primary", use_container_width=True)

            if submitted:
                if check_passcode(passcode):
                    st.session_state["authenticated"] = True
                    st.session_state["auth_failed"] = False
                    st.rerun()
                else:
                    st.session_state["auth_failed"] = True
                    st.rerun()

        # Footer
        st.markdown("""
        <div class="login-footer">
            SourceHub Development Team
        </div>
        """, unsafe_allow_html=True)


# =============================================================================
# Sidebar
# =============================================================================

def render_sidebar():
    """Render sidebar with configuration."""
    with st.sidebar.expander("Configuration", expanded=False):
        # Check for token in secrets or environment (secure sources)
        default_token = ""
        token_is_secure = False
        try:
            if "ASANA_ACCESS_TOKEN" in st.secrets:
                default_token = st.secrets["ASANA_ACCESS_TOKEN"]
                token_is_secure = True
        except FileNotFoundError:
            pass

        if not default_token and os.environ.get("ASANA_ACCESS_TOKEN"):
            default_token = os.environ.get("ASANA_ACCESS_TOKEN", "")
            token_is_secure = True

        # Only show token input if NOT securely configured
        if token_is_secure:
            # Token is securely configured via secrets/env - hide input
            token = default_token
        else:
            # No secure token - show input for local development only
            st.subheader("Authentication")
            token = st.text_input(
                "Asana Access Token",
                value="",
                type="password",
                help="Your Asana Personal Access Token"
            )
            st.caption(
                "[Get token from Asana](https://app.asana.com/0/developer-console)"
            )
            st.markdown("---")

        st.subheader("Options")

        fetch_comments = st.checkbox(
            "Fetch Comments",
            value=True,
            help="Check for daily updates (slower but more accurate)"
        )

        fetch_completed = st.checkbox(
            "Fetch Completed Tasks",
            value=True,
            help="Include completed tasks for burndown calculation"
        )

        min_description_length = st.number_input(
            "Min Description Length",
            min_value=50,
            max_value=500,
            value=100,
            step=25,
        )

        hours_without_update = st.number_input(
            "Hours Without Update",
            min_value=12,
            max_value=72,
            value=24,
            step=6,
        )

    return {
        "token": token,
        "fetch_comments": fetch_comments,
        "fetch_completed": fetch_completed,
        "min_description_length": min_description_length,
        "hours_without_update": hours_without_update,
    }


def render_dashboard_filters(
    results: list[TaskCompliance],
    completed_results: Optional[list[TaskCompliance]],
    analyzer
) -> dict:
    """Render filter controls on the dashboard (horizontal layout)."""
    st.subheader("Filters")

    col1, col2, col3, col4 = st.columns([2, 2, 2, 1])

    with col1:
        # Sprint filter - combine active and completed tasks to get all sprints with data
        all_tasks_for_sprints = results + (completed_results or [])
        sprints = analyzer.get_unique_sprints(all_tasks_for_sprints)

        # Default to the last sprint (most recent) if available
        default_index = len(sprints) if sprints else 0

        selected_sprint = st.selectbox(
            "Sprint",
            ["All"] + sprints,
            index=default_index,
            help="Filter by sprint (showing only sprints with data)",
            key="filter_sprint"
        )

    with col2:
        # Assignee filter - also from all tasks
        assignees = analyzer.get_unique_assignees(all_tasks_for_sprints)
        selected_assignees = st.multiselect(
            "Assignees",
            assignees,
            default=[],
            help="Filter by assignee (empty = all)",
            key="filter_assignees"
        )

    with col3:
        # Status filter
        statuses = analyzer.get_unique_statuses(results)
        selected_statuses = st.multiselect(
            "Status",
            statuses,
            default=[],
            help="Filter by status (empty = all)",
            key="filter_statuses"
        )

    with col4:
        st.write("")  # Spacing
        st.write("")  # Align with other fields
        if st.button("Refresh Data", type="secondary", use_container_width=True):
            st.session_state["report_generated"] = False
            st.rerun()

    # Completion Analytics Date Range Filter
    st.subheader("Completion Date Range")
    col_start, col_end = st.columns(2)
    with col_start:
        completion_start = st.date_input(
            "From",
            value=datetime.now().date() - timedelta(days=14),
            help="Start date for completion analytics",
            key="completion_date_start"
        )
    with col_end:
        completion_end = st.date_input(
            "To",
            value=datetime.now().date(),
            help="End date for completion analytics",
            key="completion_date_end"
        )

    # Validate date range
    if completion_start > completion_end:
        st.error("Start date must be before or equal to end date")
        completion_start = completion_end - timedelta(days=14)

    return {
        "sprint": selected_sprint if selected_sprint != "All" else None,
        "assignees": selected_assignees if selected_assignees else None,
        "statuses": selected_statuses if selected_statuses else None,
        "completion_start": completion_start,
        "completion_end": completion_end,
    }


# =============================================================================
# Metric Cards
# =============================================================================

def render_metric_cards(summary: ReportSummary, metrics: dict):
    """Render summary metric cards with neumorphic design."""
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        compliance_class = "nm-card--success" if summary.compliance_rate >= 80 else "nm-card--warning"
        st.markdown(f"""
        <div class="nm-card {compliance_class}">
            <div class="nm-card-value">{summary.compliance_rate:.0f}%</div>
            <div class="nm-card-label">Compliance Rate</div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
        <div class="nm-card">
            <div class="nm-card-value">{summary.total_tasks}</div>
            <div class="nm-card-label">Total Tasks</div>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        st.markdown(f"""
        <div class="nm-card nm-card--info">
            <div class="nm-card-value">{metrics.get('total_points', 0):.0f}</div>
            <div class="nm-card-label">Story Points</div>
        </div>
        """, unsafe_allow_html=True)

    with col4:
        updates_class = "nm-card--warning" if summary.tasks_missing_updates > 0 else "nm-card--success"
        st.markdown(f"""
        <div class="nm-card {updates_class}">
            <div class="nm-card-value">{summary.tasks_missing_updates}</div>
            <div class="nm-card-label">Missing Updates</div>
        </div>
        """, unsafe_allow_html=True)


# =============================================================================
# Asana Task Viewer (Modal Dialog)
# =============================================================================

@st.dialog("Task Details", width="large")
def show_task_dialog(task_gid: str, task_url: str, task_name: str, reporter):
    """Show task details in a modal dialog."""
    # Header with link to Asana
    col1, col2 = st.columns([4, 1])
    with col1:
        st.subheader(task_name)
    with col2:
        st.link_button("Open in Asana", task_url, use_container_width=True)

    st.divider()

    # Fetch full task details from API
    try:
        with st.spinner("Loading task details..."):
            task_details = reporter.client.tasks_api.get_task(
                task_gid,
                opts={"opt_fields": "name,notes,assignee.name,due_on,completed,created_at,modified_at,custom_fields,custom_fields.name,custom_fields.display_value,permalink_url"}
            )
            task = task_details.to_dict() if hasattr(task_details, 'to_dict') else dict(task_details)

        # Display task details in columns
        col1, col2 = st.columns(2)

        with col1:
            assignee = task.get('assignee', {})
            assignee_name = assignee.get('name', 'Unassigned') if assignee else 'Unassigned'
            st.markdown(f"**Assignee:** {assignee_name}")
            st.markdown(f"**Due Date:** {task.get('due_on') or 'Not set'}")
            st.markdown(f"**Status:** {'Completed' if task.get('completed') else 'In Progress'}")

        with col2:
            st.markdown(f"**Created:** {task.get('created_at', '')[:10] if task.get('created_at') else 'N/A'}")
            st.markdown(f"**Modified:** {task.get('modified_at', '')[:10] if task.get('modified_at') else 'N/A'}")

        # Custom fields
        custom_fields = task.get('custom_fields', []) or []
        if custom_fields:
            st.divider()
            st.markdown("**Custom Fields:**")
            cf_cols = st.columns(3)
            for i, cf in enumerate(custom_fields):
                if cf and cf.get('display_value'):
                    cf_cols[i % 3].markdown(f"**{cf.get('name')}:** {cf.get('display_value')}")

        # Description
        st.divider()
        notes = task.get('notes', '')
        if notes:
            st.markdown("**Description:**")
            st.text_area("", value=notes, height=200, disabled=True, key="dialog_task_notes", label_visibility="collapsed")
        else:
            st.warning("No description provided")

        # Fetch and display recent comments
        st.divider()
        st.markdown("**Recent Comments:**")
        try:
            comments = reporter.client.get_task_comments(task_gid, limit=5)
            if comments:
                for comment in comments[:5]:
                    author = comment.get('created_by', {}).get('name', 'Unknown')
                    text = comment.get('text', '')
                    date = comment.get('created_at', '')[:10] if comment.get('created_at') else ''
                    if text:
                        st.markdown(f"**{author}** ({date})")
                        st.markdown(f"> {text[:500]}{'...' if len(text) > 500 else ''}")
                        st.write("")
            else:
                st.info("No comments yet")
        except Exception:
            st.info("Could not load comments")

    except Exception as e:
        st.error(f"Could not load task details: {e}")
        st.link_button("Open in Asana instead", task_url)


def open_task_viewer(task_gid: str, task_url: str, task_name: str):
    """Store task info in session state to trigger dialog."""
    st.session_state["selected_task_gid"] = task_gid
    st.session_state["selected_task_url"] = task_url
    st.session_state["selected_task_name"] = task_name


# =============================================================================
# Burndown Chart
# =============================================================================

def task_in_sprint(task: TaskCompliance, sprint: str) -> bool:
    """Check if a task belongs to a sprint (handles comma-separated sprint values)."""
    if not task.sprint:
        return False
    # Sprint field can be comma-separated like "Manali, London"
    task_sprints = [s.strip() for s in task.sprint.split(",")]
    return sprint in task_sprints


def classify_sprint_task(task: TaskCompliance, sprint: str) -> str:
    """Classify a task's relationship to the selected sprint.

    Returns:
        "current_only" - task is tagged with only the selected sprint
        "spillover"    - task is tagged with selected sprint AND other sprints
        "carry_in"     - task is NOT tagged with selected sprint
    """
    if not task.sprint:
        return "carry_in"
    task_sprints = [s.strip() for s in task.sprint.split(",")]
    if sprint not in task_sprints:
        return "carry_in"
    if len(task_sprints) > 1:
        return "spillover"
    return "current_only"


def render_burndown_chart(
    results: list[TaskCompliance],
    completed_results: Optional[list[TaskCompliance]] = None,
    selected_sprint: Optional[str] = None,
    target_sprint_points: Optional[int] = None,
    all_results: Optional[list[TaskCompliance]] = None,
):
    """Render sprint burndown chart with actual progress line."""
    if not PLOTLY_AVAILABLE:
        st.warning("Plotly is required for charts. Install with: pip install plotly")
        return

    # Determine which sprint to show
    if selected_sprint:
        sprint = selected_sprint
        # Filter tasks that contain this sprint (handles comma-separated values)
        sprint_tasks = [t for t in results if task_in_sprint(t, sprint)]
        completed_sprint_tasks = [t for t in (completed_results or []) if task_in_sprint(t, sprint)]
    else:
        sprint = "All Sprints"
        sprint_tasks = results
        completed_sprint_tasks = completed_results or []

    if not sprint_tasks and not completed_sprint_tasks:
        st.info("No tasks found for burndown chart")
        return

    # Exclude QA testers from burndown calculations
    QA_TESTERS = ("Dimpi Gogoi", "Vinay")
    def _is_qa_tester(assignee):
        if not assignee:
            return False
        return any(assignee.startswith(qa) for qa in QA_TESTERS)

    sprint_tasks = [t for t in sprint_tasks if not _is_qa_tester(t.assignee)]
    completed_sprint_tasks = [t for t in completed_sprint_tasks if not _is_qa_tester(t.assignee)]

    # Classify spillover tasks (tagged with current sprint + other sprints)
    spillover_gids = set()
    spillover_points = 0
    if selected_sprint:
        for t in list(sprint_tasks) + list(completed_sprint_tasks):
            if t.gid not in spillover_gids and classify_sprint_task(t, selected_sprint) == "spillover":
                spillover_gids.add(t.gid)
                try:
                    spillover_points += float(t.story_points) if t.story_points else 0
                except (ValueError, TypeError):
                    pass

    # Separate tasks by completion status
    # Tasks in Review, QA, Ready to Ship, or Done are considered "completed" for burndown purposes
    completed_statuses = ("Review", "QA", "Ready to Ship", "Done")
    done_tasks = [t for t in sprint_tasks if t.progress in completed_statuses]
    active_tasks = [t for t in sprint_tasks if t.progress not in completed_statuses]

    # Calculate total points (all tasks in sprint)
    total_points = 0
    completed_points = 0
    completion_dates = {}  # date -> {points, developers, tasks}

    def _record_completion(date_key, pts, dev_name, task_obj, target=None):
        """Helper to record a completion into a date-keyed dict (default: completion_dates)."""
        dest = target if target is not None else completion_dates
        if date_key not in dest:
            dest[date_key] = {"points": 0, "developers": {}, "tasks": []}
        dest[date_key]["points"] += pts
        dest[date_key]["developers"][dev_name] = (
            dest[date_key]["developers"].get(dev_name, 0) + pts
        )
        dest[date_key]["tasks"].append(task_obj)

    # Process active (not done) tasks
    for task in active_tasks:
        try:
            points = float(task.story_points) if task.story_points else 0
        except (ValueError, TypeError):
            points = 0
        total_points += points

    # Process "Done" tasks from incomplete list
    for task in done_tasks:
        try:
            points = float(task.story_points) if task.story_points else 0
        except (ValueError, TypeError):
            points = 0
        total_points += points
        completed_points += points

        # Use status_changed_at (when progress field changed), fallback to modified_at, then due_on
        if points > 0:
            dev_name = task.assignee or "Unassigned"
            status_date = getattr(task, 'status_changed_at', None)
            modified_date = getattr(task, 'modified_at', None)
            if status_date:
                _record_completion(status_date, points, dev_name, task)
            elif modified_date:
                _record_completion(modified_date[:10], points, dev_name, task)
            elif task.due_on:
                _record_completion(task.due_on, points, dev_name, task)

    # Process truly completed tasks from Asana
    for task in completed_sprint_tasks:
        try:
            points = float(task.story_points) if task.story_points else 0
        except (ValueError, TypeError):
            points = 0
        total_points += points
        completed_points += points

        dev_name = task.assignee or "Unassigned"
        status_date = getattr(task, 'status_changed_at', None)
        modified_date = getattr(task, 'modified_at', None)
        # Use completed_at, then status_changed_at, then modified_at, then due_on
        if points > 0 and task.completed_at:
            _record_completion(task.completed_at[:10], points, dev_name, task)
        elif points > 0 and status_date:
            _record_completion(status_date, points, dev_name, task)
        elif points > 0 and modified_date:
            _record_completion(modified_date[:10], points, dev_name, task)
        elif points > 0 and task.due_on:
            _record_completion(task.due_on, points, dev_name, task)

    if total_points == 0:
        st.info("No story points found for this sprint")
        return

    # Determine sprint boundaries (2-week sprints, 10 working days)
    today = datetime.now()

    # Use anchor date to compute current sprint start (2-week cadence)
    anchor_str = os.environ.get("SPRINT_ANCHOR_DATE", "2026-02-03")
    try:
        anchor = datetime.strptime(anchor_str, "%Y-%m-%d")
        days_since_anchor = (today - anchor).days
        current_sprint_offset = (days_since_anchor // 14) * 14
        sprint_start = (anchor + timedelta(days=current_sprint_offset)).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
    except ValueError:
        # Fallback: most recent Monday on or before today
        days_since_monday = today.weekday()
        sprint_start = (today - timedelta(days=days_since_monday)).replace(
            hour=0, minute=0, second=0, microsecond=0
        )

    # Build list of 10 working days (skip Saturday=5, Sunday=6)
    sprint_working_dates = []
    current_date = sprint_start
    while len(sprint_working_dates) < 10:
        if current_date.weekday() < 5:  # Mon-Fri
            sprint_working_dates.append(current_date)
        current_date += timedelta(days=1)

    sprint_days = len(sprint_working_dates)  # 10
    sprint_end = sprint_working_dates[-1]

    # Also accumulate any weekend completions into the next working day
    # so points completed on Sat/Sun aren't lost from the chart
    all_calendar_dates = []
    d = sprint_start
    while d <= sprint_end:
        all_calendar_dates.append(d)
        d += timedelta(days=1)

    # Map each calendar date to its nearest working day (for weekend rollup)
    def _next_working_day(dt):
        while dt.weekday() >= 5:
            dt += timedelta(days=1)
        return dt

    # Roll weekend completion_dates into the following Monday
    weekend_keys = [k for k in completion_dates if datetime.strptime(k, "%Y-%m-%d").weekday() >= 5]
    for wk in weekend_keys:
        wd = _next_working_day(datetime.strptime(wk, "%Y-%m-%d")).strftime("%Y-%m-%d")
        if wd not in completion_dates:
            completion_dates[wd] = {"points": 0, "developers": {}, "tasks": []}
        completion_dates[wd]["points"] += completion_dates[wk]["points"]
        for dev, pts in completion_dates[wk]["developers"].items():
            completion_dates[wd]["developers"][dev] = completion_dates[wd]["developers"].get(dev, 0) + pts
        completion_dates[wd].setdefault("tasks", []).extend(completion_dates[wk].get("tasks", []))
        del completion_dates[wk]

    # Detect carry-in tasks: completed during this sprint window but NOT tagged
    # with the current sprint.  These represent real work done but invisible in
    # the main burndown because the task belongs to an older sprint.
    carry_in_tasks = []
    carry_in_points = 0
    carry_in_completion_dates = {}  # date -> {points, developers, tasks}
    sprint_start_str = sprint_start.strftime("%Y-%m-%d")
    sprint_end_str = sprint_end.strftime("%Y-%m-%d")

    if selected_sprint and all_results is not None:
        # Combine unfiltered active + completed to find tasks outside this sprint
        seen_gids = {t.gid for t in list(sprint_tasks) + list(completed_sprint_tasks)}
        candidate_pool = list(all_results or []) + list(completed_results or [])
        for t in candidate_pool:
            if t.gid in seen_gids:
                continue
            if _is_qa_tester(t.assignee):
                continue
            # Must not be tagged with current sprint
            if task_in_sprint(t, selected_sprint):
                continue
            # Must have a completion indicator within the sprint window
            comp_date = None
            if t.completed_at:
                comp_date = t.completed_at[:10]
            elif t.progress in ("Review", "QA", "Ready to Ship", "Done"):
                status_dt = getattr(t, 'status_changed_at', None)
                mod_dt = getattr(t, 'modified_at', None)
                comp_date = status_dt or (mod_dt[:10] if mod_dt else None)
            if not comp_date:
                continue
            if comp_date < sprint_start_str or comp_date > sprint_end_str:
                continue
            seen_gids.add(t.gid)
            try:
                pts = float(t.story_points) if t.story_points else 0
            except (ValueError, TypeError):
                pts = 0
            carry_in_tasks.append(t)
            carry_in_points += pts
            if pts > 0:
                dev_name = t.assignee or "Unassigned"
                _record_completion(comp_date, pts, dev_name, t, target=carry_in_completion_dates)

    # Generate sprint-day-based series (x-axis = Sprint Day 1, 2, 3, …, 10)
    sprint_day_nums = []
    real_dates = []
    ideal_line = []
    actual_line = []
    hover_texts = []

    ideal_total = target_sprint_points if target_sprint_points else total_points
    daily_decrement = ideal_total / sprint_days
    remaining = total_points

    for day_num, working_date in enumerate(sprint_working_dates):
        date_str = working_date.strftime("%Y-%m-%d")
        sprint_day = day_num + 1
        sprint_day_nums.append(sprint_day)
        real_dates.append(date_str)

        # Ideal burndown (uses target if set, otherwise total from tasks)
        ideal_remaining = max(0, ideal_total - (daily_decrement * day_num))
        ideal_line.append(round(ideal_remaining))

        # Actual burndown - subtract completed points up to this date
        if date_str in completion_dates:
            remaining -= completion_dates[date_str]["points"]

        # Only show actual line up to today
        if working_date <= today:
            actual_val = max(0, remaining)
            actual_line.append(round(actual_val))
            # Build hover text with developer breakdown
            if date_str in completion_dates:
                devs = completion_dates[date_str]["developers"]
                dev_lines = [f"{name}: {pts:.0f} pts" for name, pts in sorted(devs.items())]
                hover_texts.append(
                    f"Day {sprint_day} ({date_str})<br>"
                    f"Remaining: {actual_val:.0f} pts<br>" + "<br>".join(dev_lines)
                )
            else:
                hover_texts.append(f"Day {sprint_day} ({date_str})<br>Remaining: {actual_val:.0f} pts")
        else:
            actual_line.append(None)
            hover_texts.append(None)

    # Create chart
    fig = go.Figure()

    # Neumorphism color palette for charts
    nm_primary = '#6B7FD7'      # Muted blue-purple
    nm_success = '#5B9A8B'      # Sage green
    nm_error = '#C9736D'        # Muted coral
    nm_text_primary = '#2D3748' # Dark slate
    nm_bg = '#E4E8EC'           # Soft gray background

    # Ideal burndown line
    fig.add_trace(go.Scatter(
        x=sprint_day_nums,
        y=ideal_line,
        mode='lines',
        name='Ideal Burndown',
        line=dict(color=nm_primary, dash='dash', width=2)
    ))

    # Actual burndown line
    fig.add_trace(go.Scatter(
        x=sprint_day_nums,
        y=actual_line,
        mode='lines+markers',
        name='Actual Burndown',
        line=dict(color=nm_success, width=3),
        marker=dict(size=6),
        connectgaps=False,
        hovertext=hover_texts,
        hoverinfo="text"
    ))

    # Current state marker — find which sprint day is today
    today_str = today.strftime("%Y-%m-%d")
    if today_str in real_dates:
        idx = real_dates.index(today_str)
        current_remaining = actual_line[idx] if actual_line[idx] is not None else remaining
        fig.add_trace(go.Scatter(
            x=[sprint_day_nums[idx]],
            y=[current_remaining],
            mode='markers',
            name='Today',
            marker=dict(color=nm_error, size=14, symbol='diamond'),
            showlegend=True
        ))

    # Summary annotation
    pct_complete = (completed_points / total_points * 100) if total_points > 0 else 0
    summary_text = f"Completed: {completed_points:.0f} / {total_points:.0f} pts ({pct_complete:.0f}%)"
    if target_sprint_points:
        summary_text += f" | Target: {ideal_total:.0f} pts"
    fig.add_annotation(
        x=0.02, y=0.98,
        xref="paper", yref="paper",
        text=summary_text,
        showarrow=False,
        font=dict(size=14, color=nm_success),
        bgcolor="rgba(228,232,236,0.95)",
        borderpad=6
    )

    # Date range for subtitle
    start_label = sprint_start.strftime("%-m/%-d/%Y")
    end_label = sprint_end.strftime("%-m/%-d/%Y")

    fig.update_layout(
        title=dict(
            text=f"Sprint Burndown: {sprint}, {start_label} - {end_label}",
            font=dict(size=20, color=nm_text_primary)
        ),
        xaxis_title="Sprint Day",
        yaxis_title="Story Points Remaining",
        hovermode="x unified",
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        height=450,
        margin=dict(t=80),
        paper_bgcolor=nm_bg,
        plot_bgcolor=nm_bg,
        font=dict(color=nm_text_primary),
        xaxis=dict(
            dtick=1,
            gridcolor='rgba(163, 177, 198, 0.3)',
            linecolor='rgba(163, 177, 198, 0.5)',
            tickcolor='rgba(163, 177, 198, 0.5)',
        ),
        yaxis=dict(
            gridcolor='rgba(163, 177, 198, 0.3)',
            linecolor='rgba(163, 177, 198, 0.5)',
            tickcolor='rgba(163, 177, 198, 0.5)',
        ),
    )

    st.plotly_chart(fig, use_container_width=True, key="burndown_main")

    # Spillover / carry-in info banners
    if spillover_gids:
        spillover_count = len(spillover_gids)
        st.info(
            f"**Spillover tasks:** {spillover_count} task(s) worth "
            f"**{int(spillover_points)} pts** are tagged with multiple sprints. "
            f"Their points are included in this burndown (matches Asana)."
        )

    if carry_in_tasks:
        carry_in_count = len(carry_in_tasks)
        st.warning(
            f"**Carry-in tasks:** {carry_in_count} task(s) worth "
            f"**{int(carry_in_points)} pts** were completed during this sprint "
            f"but are tagged with an older sprint. These are NOT included in the "
            f"burndown above."
        )
        with st.expander(f"View {carry_in_count} carry-in task(s)"):
            ci_rows = []
            for t in carry_in_tasks:
                try:
                    pts = float(t.story_points) if t.story_points else 0
                except (ValueError, TypeError):
                    pts = 0
                comp_date = ""
                if t.completed_at:
                    comp_date = t.completed_at[:10]
                elif getattr(t, 'status_changed_at', None):
                    comp_date = t.status_changed_at  # already YYYY-MM-DD
                elif getattr(t, 'modified_at', None):
                    comp_date = t.modified_at[:10]
                ci_rows.append({
                    "Task": t.name,
                    "Assignee": t.assignee or "Unassigned",
                    "Points": int(pts),
                    "Sprint Tag": t.sprint or "",
                    "Status": t.progress or "Completed",
                    "Completed": comp_date,
                })
            st.dataframe(pd.DataFrame(ci_rows), use_container_width=True, hide_index=True)

    # Per-developer point breakdown
    all_burndown_tasks = list(sprint_tasks) + list(completed_sprint_tasks)
    dev_points = {}  # developer -> {total, completed, spillover, tasks}
    for t in all_burndown_tasks:
        try:
            pts = float(t.story_points) if t.story_points else 0
        except (ValueError, TypeError):
            pts = 0
        dev = t.assignee or "Unassigned"
        if dev not in dev_points:
            dev_points[dev] = {"total": 0, "completed": 0, "spillover": 0, "tasks": []}
        dev_points[dev]["total"] += pts
        is_done = (t.progress in completed_statuses) or (t.completed_at is not None)
        if is_done:
            dev_points[dev]["completed"] += pts
        if t.gid in spillover_gids:
            dev_points[dev]["spillover"] += pts
        dev_points[dev]["tasks"].append({"name": t.name, "points": pts, "status": t.progress or "Completed"})

    with st.expander("Points Breakdown by Developer"):
        breakdown_rows = []
        for dev, data in sorted(dev_points.items(), key=lambda x: x[1]["total"], reverse=True):
            row = {
                "Developer": dev,
                "Total Pts": int(data["total"]),
                "Completed Pts": int(data["completed"]),
                "Remaining Pts": int(data["total"] - data["completed"]),
                "# Tasks": len(data["tasks"]),
            }
            if spillover_gids:
                row["Spillover Pts"] = int(data["spillover"])
            breakdown_rows.append(row)
        if breakdown_rows:
            st.dataframe(pd.DataFrame(breakdown_rows), use_container_width=True, hide_index=True)
            st.caption(f"Sprint total: {int(total_points)} pts across {len(dev_points)} developers")

    # Return data for the full-width table (rendered outside column layout)
    pts_per_day = round(ideal_total / sprint_days, 1) if sprint_days > 0 else 0
    completed_detail = []
    for date_str in real_dates:
        if date_str in completion_dates:
            tasks = completion_dates[date_str]["tasks"]
            lines = []
            for t in tasks:
                pts = float(t.story_points) if t.story_points else 0
                marker = " [SPILLOVER]" if t.gid in spillover_gids else ""
                lines.append(f"{t.name}{marker} ({t.assignee}, {pts:.0f}pts)")
            completed_detail.append("; ".join(lines))
        else:
            completed_detail.append("")

    return {
        "sprint_day_nums": sprint_day_nums,
        "real_dates": real_dates,
        "ideal_line": ideal_line,
        "actual_line": actual_line,
        "completed_detail": completed_detail,
        "ideal_total": ideal_total,
        "sprint_days": sprint_days,
        "pts_per_day": pts_per_day,
        "sprint": sprint,
        "completion_dates": completion_dates,
        "spillover_gids": spillover_gids,
        "spillover_points": spillover_points,
        "carry_in_tasks": carry_in_tasks,
        "carry_in_points": carry_in_points,
        "carry_in_completion_dates": carry_in_completion_dates,
    }


# =============================================================================
# Burndown Excel Report Generator
# =============================================================================

_ASANA_PROFILE_RE = re.compile(
    r'https?://app\.asana\.com/\d+/\d+/profile/(\d+)'
)


def _clean_comment_text(text: str, asana_client=None) -> str:
    """Replace Asana @mention profile URLs with @Name and clean up whitespace."""

    def _replace_mention(match):
        user_gid = match.group(1)
        if asana_client:
            name = asana_client.get_user_name(user_gid)
            if name:
                return f"@{name}"
        return ""

    text = _ASANA_PROFILE_RE.sub(_replace_mention, text)
    # Collapse leftover artifacts: "cc: " with nothing after, dangling dashes, etc.
    text = re.sub(r'\bcc:\s*$', '', text, flags=re.MULTILINE)
    text = re.sub(r'\b[Cc]c:\s+(?=\s)', '', text)
    # Collapse multiple spaces / blank lines
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n{2,}', '\n', text)
    return text.strip()


def _extract_pr_links(notes: str | None, comments: list[dict]) -> str:
    """Extract GitHub PR URLs from task notes and comments."""
    pr_pattern = re.compile(r'https?://github\.com/[^\s)>\]]+/pull/\d+')
    urls = set()
    if notes:
        urls.update(pr_pattern.findall(notes))
    for comment in comments:
        text = comment.get('text', '') or ''
        urls.update(pr_pattern.findall(text))
    return "\n".join(sorted(urls))


def generate_burndown_excel_report(burndown_data: dict, asana_client, all_sprint_tasks: list | None = None) -> bytes:
    """Generate a rich multi-sheet Excel burndown report.

    Sheet 0 (Overview): Task inventory grouped by status, reconciliation, per-assignee breakdown.
    Sheet 1 (Summary): Burndown chart, stats, and daily summary table.
    Sheets 2-11: One sheet per sprint day with completed task details.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    today_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Shared styles
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    dark_blue_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    gray_fill = PatternFill(start_color="757575", end_color="757575", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=16, color="1F4E79")
    subtitle_font = Font(bold=True, size=12, color="333333")
    link_font = Font(color="1F4E79", underline="single")
    bold_center = Font(bold=True, size=11)

    # Day-type header fills
    blue_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    amber_fill = PatternFill(start_color="F57F17", end_color="F57F17", fill_type="solid")

    # Due/upcoming task styles (greyed out)
    due_task_font = Font(color="999999", size=11)
    due_task_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    due_task_link_font = Font(color="999999", underline="single")
    due_task_bold = Font(color="999999", bold=True, size=11)
    section_font = Font(color="666666", bold=True, italic=True, size=11)

    # Status color fills
    status_fills = {
        "Review": PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
        "QA": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "Ready to Ship": PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
        "Done": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    }

    def _auto_width(ws):
        for col_cells in ws.columns:
            # Skip merged cells that lack column_letter
            first = col_cells[0]
            if not hasattr(first, 'column_letter'):
                continue
            col_letter = first.column_letter
            max_len = 0
            for cell in col_cells:
                try:
                    if hasattr(cell, 'value') and cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 50))
                except (TypeError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    def _style_header(ws, row, num_cols, fill):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    # ── Sheet 0: Overview ────────────────────────────────────────────────
    sprint_name = burndown_data["sprint"]
    spillover_gids = burndown_data.get("spillover_gids", set())
    spillover_points = burndown_data.get("spillover_points", 0)
    carry_in_tasks_list = burndown_data.get("carry_in_tasks", [])
    carry_in_points = burndown_data.get("carry_in_points", 0)

    # -- Classify all sprint tasks by progress status --
    completed_statuses = ("Review", "QA", "Ready to Ship", "Done")
    status_buckets: dict[str, list] = {
        "In Progress": [],
        "Review": [],
        "QA": [],
        "Ready to Ship": [],
        "Done": [],
    }
    # Deduplicate by gid (a task can appear in both active and completed lists)
    deduped_tasks: list = []
    if all_sprint_tasks:
        seen_gids: set[str] = set()
        for t in all_sprint_tasks:
            if t.gid in seen_gids:
                continue
            seen_gids.add(t.gid)
            deduped_tasks.append(t)

    for t in deduped_tasks:
        progress = t.progress or ""
        if progress in status_buckets:
            status_buckets[progress].append(t)
        elif progress:
            status_buckets.setdefault(progress, []).append(t)

    def _task_pts(task):
        try:
            return float(task.story_points) if task.story_points else 0
        except (ValueError, TypeError):
            return 0

    # -- Per-assignee aggregation --
    assignee_data: dict[str, dict] = {}  # name -> {total, completed, remaining, spillover, count}
    for t in deduped_tasks:
        name = t.assignee or "Unassigned"
        pts = _task_pts(t)
        if name not in assignee_data:
            assignee_data[name] = {"total": 0, "completed": 0, "remaining": 0, "spillover": 0, "count": 0}
        assignee_data[name]["total"] += pts
        assignee_data[name]["count"] += 1
        if t.progress in completed_statuses:
            assignee_data[name]["completed"] += pts
        else:
            assignee_data[name]["remaining"] += pts
        if t.gid in spillover_gids:
            assignee_data[name]["spillover"] += pts

    # -- Spillover task objects (from all_sprint_tasks) --
    spillover_task_objs = [t for t in (all_sprint_tasks or []) if t.gid in spillover_gids]

    # -- Create the Overview worksheet --
    ws_overview = wb.active
    ws_overview.title = "Overview"

    # Section A: Sprint Header
    ws_overview.cell(row=1, column=1, value=f"Sprint Overview: {sprint_name}").font = title_font
    ws_overview.merge_cells("A1:G1")
    ws_overview.cell(row=2, column=1, value=f"Generated: {today_str}").font = subtitle_font
    ws_overview.merge_cells("A2:G2")

    # Section B: Reconciliation Summary
    ov_row = 4
    ws_overview.cell(row=ov_row, column=1, value="Reconciliation Summary").font = subtitle_font
    ws_overview.merge_cells(start_row=ov_row, start_column=1, end_row=ov_row, end_column=3)
    ov_row += 1

    recon_headers = ["Category", "# Tasks", "Story Points"]
    for i, h in enumerate(recon_headers):
        cell = ws_overview.cell(row=ov_row, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = dark_blue_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    ov_row += 1

    total_tasks = 0
    total_pts = 0.0
    for status_name in ("In Progress", "Review", "QA", "Ready to Ship", "Done"):
        tasks = status_buckets.get(status_name, [])
        count = len(tasks)
        pts = sum(_task_pts(t) for t in tasks)
        total_tasks += count
        total_pts += pts
        ws_overview.cell(row=ov_row, column=1, value=status_name).border = thin_border
        ws_overview.cell(row=ov_row, column=2, value=count).border = thin_border
        ws_overview.cell(row=ov_row, column=2).alignment = Alignment(horizontal='center')
        ws_overview.cell(row=ov_row, column=3, value=int(pts)).border = thin_border
        ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center')
        # Color code status
        if status_name in status_fills:
            ws_overview.cell(row=ov_row, column=1).fill = status_fills[status_name]
        ov_row += 1

    # Sprint Total row (bold)
    for col in range(1, 4):
        ws_overview.cell(row=ov_row, column=col).font = bold_center
        ws_overview.cell(row=ov_row, column=col).border = thin_border
        ws_overview.cell(row=ov_row, column=col).alignment = Alignment(horizontal='center')
    ws_overview.cell(row=ov_row, column=1, value="Sprint Total")
    ws_overview.cell(row=ov_row, column=2, value=total_tasks)
    ws_overview.cell(row=ov_row, column=3, value=int(total_pts))
    ov_row += 1

    # Spillover row
    ws_overview.cell(row=ov_row, column=1, value="Spillover (multi-sprint)").border = thin_border
    ws_overview.cell(row=ov_row, column=2, value=len(spillover_task_objs)).border = thin_border
    ws_overview.cell(row=ov_row, column=2).alignment = Alignment(horizontal='center')
    ws_overview.cell(row=ov_row, column=3, value=int(spillover_points)).border = thin_border
    ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center')
    ov_row += 1

    # Carry-in row
    ws_overview.cell(row=ov_row, column=1, value="Carry-in (older sprint, done this window)").border = thin_border
    ws_overview.cell(row=ov_row, column=2, value=len(carry_in_tasks_list)).border = thin_border
    ws_overview.cell(row=ov_row, column=2).alignment = Alignment(horizontal='center')
    ws_overview.cell(row=ov_row, column=3, value=int(carry_in_points)).border = thin_border
    ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center')
    ov_row += 1

    # Effective Total row (bold)
    effective_total = total_pts + carry_in_points
    for col in range(1, 4):
        ws_overview.cell(row=ov_row, column=col).font = bold_center
        ws_overview.cell(row=ov_row, column=col).border = thin_border
        ws_overview.cell(row=ov_row, column=col).alignment = Alignment(horizontal='center')
    ws_overview.cell(row=ov_row, column=1, value="Effective Total (Sprint + Carry-in)")
    ws_overview.cell(row=ov_row, column=2, value="")
    ws_overview.cell(row=ov_row, column=3, value=int(effective_total))
    ov_row += 2

    # Section C: Points by Assignee
    ws_overview.cell(row=ov_row, column=1, value="Points by Assignee").font = subtitle_font
    ws_overview.merge_cells(start_row=ov_row, start_column=1, end_row=ov_row, end_column=6)
    ov_row += 1

    assignee_headers = ["Assignee", "Total Pts", "Completed Pts", "Remaining Pts", "Spillover Pts", "# Tasks"]
    for i, h in enumerate(assignee_headers):
        cell = ws_overview.cell(row=ov_row, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = dark_blue_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    ov_row += 1

    for name in sorted(assignee_data.keys()):
        data = assignee_data[name]
        vals = [name, int(data["total"]), int(data["completed"]), int(data["remaining"]), int(data["spillover"]), data["count"]]
        for i, val in enumerate(vals):
            cell = ws_overview.cell(row=ov_row, column=i + 1, value=val)
            cell.border = thin_border
            if i >= 1:
                cell.alignment = Alignment(horizontal='center')
        ov_row += 1

    ov_row += 1

    # Section D: In Progress Task List
    in_progress_tasks = status_buckets.get("In Progress", [])
    ip_pts = sum(_task_pts(t) for t in in_progress_tasks)
    overview_task_headers = ["Task Name", "Assignee", "Story Points", "Type", "Epic", "Sprint Tag", "Asana Link"]

    section_header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    ws_overview.cell(
        row=ov_row, column=1,
        value=f"IN PROGRESS \u2014 {len(in_progress_tasks)} tasks, {int(ip_pts)} pts"
    ).font = Font(bold=True, size=12, color="1565C0")
    ws_overview.merge_cells(start_row=ov_row, start_column=1, end_row=ov_row, end_column=len(overview_task_headers))
    for col in range(1, len(overview_task_headers) + 1):
        ws_overview.cell(row=ov_row, column=col).fill = section_header_fill
    ov_row += 1

    for i, h in enumerate(overview_task_headers):
        ws_overview.cell(row=ov_row, column=i + 1, value=h)
    _style_header(ws_overview, ov_row, len(overview_task_headers), dark_blue_fill)
    ov_row += 1

    for t in sorted(in_progress_tasks, key=lambda x: _task_pts(x), reverse=True):
        vals = [
            t.name,
            t.assignee or "Unassigned",
            _task_pts(t),
            t.task_type or "",
            t.epic or "",
            t.sprint or "",
            t.url,
        ]
        for i, val in enumerate(vals):
            cell = ws_overview.cell(row=ov_row, column=i + 1, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        # Hyperlink for Asana Link column
        link_cell = ws_overview.cell(row=ov_row, column=len(overview_task_headers))
        link_cell.hyperlink = t.url
        link_cell.font = link_font
        link_cell.value = "Open in Asana"
        # Bold story points
        ws_overview.cell(row=ov_row, column=3).font = bold_center
        ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center', vertical='top')
        ov_row += 1

    # Subtotal row
    ws_overview.cell(row=ov_row, column=1, value="Subtotal").font = bold_center
    ws_overview.cell(row=ov_row, column=3, value=int(ip_pts)).font = bold_center
    ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center')
    for col in range(1, len(overview_task_headers) + 1):
        ws_overview.cell(row=ov_row, column=col).border = thin_border
    ov_row += 2

    # Section E: Completed Task List (Review + QA + Ready to Ship + Done)
    completed_tasks_all = []
    for s in ("Review", "QA", "Ready to Ship", "Done"):
        completed_tasks_all.extend(status_buckets.get(s, []))
    comp_pts = sum(_task_pts(t) for t in completed_tasks_all)

    completed_headers = ["Task Name", "Assignee", "Story Points", "Current Status", "Status Changed Date",
                         "Type", "Epic", "Sprint Tag", "Asana Link"]
    completed_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    ws_overview.cell(
        row=ov_row, column=1,
        value=f"COMPLETED (Review/QA/Ready to Ship/Done) \u2014 {len(completed_tasks_all)} tasks, {int(comp_pts)} pts"
    ).font = Font(bold=True, size=12, color="2E7D32")
    ws_overview.merge_cells(start_row=ov_row, start_column=1, end_row=ov_row, end_column=len(completed_headers))
    for col in range(1, len(completed_headers) + 1):
        ws_overview.cell(row=ov_row, column=col).fill = completed_fill
    ov_row += 1

    for i, h in enumerate(completed_headers):
        ws_overview.cell(row=ov_row, column=i + 1, value=h)
    _style_header(ws_overview, ov_row, len(completed_headers), green_fill)
    ov_row += 1

    for t in sorted(completed_tasks_all, key=lambda x: _task_pts(x), reverse=True):
        status_date = ""
        if getattr(t, 'status_changed_at', None):
            status_date = t.status_changed_at
        elif t.completed_at:
            status_date = t.completed_at[:10]
        vals = [
            t.name,
            t.assignee or "Unassigned",
            _task_pts(t),
            t.progress or "",
            status_date,
            t.task_type or "",
            t.epic or "",
            t.sprint or "",
            t.url,
        ]
        for i, val in enumerate(vals):
            cell = ws_overview.cell(row=ov_row, column=i + 1, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        # Color code status
        status = t.progress or ""
        if status in status_fills:
            ws_overview.cell(row=ov_row, column=4).fill = status_fills[status]
        # Hyperlink
        link_cell = ws_overview.cell(row=ov_row, column=len(completed_headers))
        link_cell.hyperlink = t.url
        link_cell.font = link_font
        link_cell.value = "Open in Asana"
        # Bold story points
        ws_overview.cell(row=ov_row, column=3).font = bold_center
        ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center', vertical='top')
        ov_row += 1

    # Subtotal row
    ws_overview.cell(row=ov_row, column=1, value="Subtotal").font = bold_center
    ws_overview.cell(row=ov_row, column=3, value=int(comp_pts)).font = bold_center
    ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center')
    for col in range(1, len(completed_headers) + 1):
        ws_overview.cell(row=ov_row, column=col).border = thin_border
    ov_row += 2

    # Section F: Spillover Tasks
    if spillover_task_objs:
        sp_pts = sum(_task_pts(t) for t in spillover_task_objs)
        spillover_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
        spillover_headers = ["Task Name", "Assignee", "Story Points", "Current Status", "Status Changed Date",
                             "Type", "Epic", "Sprint Tag", "Asana Link"]

        ws_overview.cell(
            row=ov_row, column=1,
            value=f"SPILLOVER TASKS \u2014 {len(spillover_task_objs)} tasks, {int(sp_pts)} pts"
        ).font = Font(bold=True, size=12, color="1565C0")
        ws_overview.merge_cells(start_row=ov_row, start_column=1, end_row=ov_row, end_column=len(spillover_headers))
        for col in range(1, len(spillover_headers) + 1):
            ws_overview.cell(row=ov_row, column=col).fill = spillover_fill
        ov_row += 1

        for i, h in enumerate(spillover_headers):
            ws_overview.cell(row=ov_row, column=i + 1, value=h)
        _style_header(ws_overview, ov_row, len(spillover_headers), PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"))
        ov_row += 1

        for t in sorted(spillover_task_objs, key=lambda x: _task_pts(x), reverse=True):
            status_date = ""
            if getattr(t, 'status_changed_at', None):
                status_date = t.status_changed_at
            elif t.completed_at:
                status_date = t.completed_at[:10]
            vals = [
                t.name,
                t.assignee or "Unassigned",
                _task_pts(t),
                t.progress or "",
                status_date,
                t.task_type or "",
                t.epic or "",
                t.sprint or "",
                t.url,
            ]
            for i, val in enumerate(vals):
                cell = ws_overview.cell(row=ov_row, column=i + 1, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            status = t.progress or ""
            if status in status_fills:
                ws_overview.cell(row=ov_row, column=4).fill = status_fills[status]
            link_cell = ws_overview.cell(row=ov_row, column=len(spillover_headers))
            link_cell.hyperlink = t.url
            link_cell.font = link_font
            link_cell.value = "Open in Asana"
            ws_overview.cell(row=ov_row, column=3).font = bold_center
            ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center', vertical='top')
            ov_row += 1

        # Subtotal
        ws_overview.cell(row=ov_row, column=1, value="Subtotal").font = bold_center
        ws_overview.cell(row=ov_row, column=3, value=int(sp_pts)).font = bold_center
        ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center')
        for col in range(1, len(spillover_headers) + 1):
            ws_overview.cell(row=ov_row, column=col).border = thin_border
        ov_row += 2

    # Section G: Carry-in Tasks
    if carry_in_tasks_list:
        ci_pts = sum(_task_pts(t) for t in carry_in_tasks_list)
        carryin_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        carryin_headers = ["Task Name", "Assignee", "Story Points", "Current Status", "Status Changed Date",
                           "Type", "Epic", "Sprint Tag", "Asana Link"]

        ws_overview.cell(
            row=ov_row, column=1,
            value=f"CARRY-IN TASKS \u2014 {len(carry_in_tasks_list)} tasks, {int(ci_pts)} pts"
        ).font = Font(bold=True, size=12, color="E65100")
        ws_overview.merge_cells(start_row=ov_row, start_column=1, end_row=ov_row, end_column=len(carryin_headers))
        for col in range(1, len(carryin_headers) + 1):
            ws_overview.cell(row=ov_row, column=col).fill = carryin_fill
        ov_row += 1

        for i, h in enumerate(carryin_headers):
            ws_overview.cell(row=ov_row, column=i + 1, value=h)
        _style_header(ws_overview, ov_row, len(carryin_headers), PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"))
        ov_row += 1

        for t in sorted(carry_in_tasks_list, key=lambda x: _task_pts(x), reverse=True):
            status_date = ""
            if getattr(t, 'status_changed_at', None):
                status_date = t.status_changed_at
            elif t.completed_at:
                status_date = t.completed_at[:10]
            vals = [
                t.name,
                t.assignee or "Unassigned",
                _task_pts(t),
                t.progress or "Completed",
                status_date,
                t.task_type or "",
                t.epic or "",
                t.sprint or "",
                t.url,
            ]
            for i, val in enumerate(vals):
                cell = ws_overview.cell(row=ov_row, column=i + 1, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            status = t.progress or ""
            if status in status_fills:
                ws_overview.cell(row=ov_row, column=4).fill = status_fills[status]
            link_cell = ws_overview.cell(row=ov_row, column=len(carryin_headers))
            link_cell.hyperlink = t.url
            link_cell.font = link_font
            link_cell.value = "Open in Asana"
            ws_overview.cell(row=ov_row, column=3).font = bold_center
            ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center', vertical='top')
            ov_row += 1

        # Subtotal
        ws_overview.cell(row=ov_row, column=1, value="Subtotal").font = bold_center
        ws_overview.cell(row=ov_row, column=3, value=int(ci_pts)).font = bold_center
        ws_overview.cell(row=ov_row, column=3).alignment = Alignment(horizontal='center')
        for col in range(1, len(carryin_headers) + 1):
            ws_overview.cell(row=ov_row, column=col).border = thin_border

    _auto_width(ws_overview)
    ws_overview.freeze_panes = "A6"  # Freeze through reconciliation column headers

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Summary")

    ws_summary.cell(row=1, column=1, value=f"Sprint Burndown: {sprint_name}").font = title_font
    ws_summary.cell(row=2, column=1, value=f"Generated: {today_str}").font = subtitle_font
    ws_summary.merge_cells("A1:F1")
    ws_summary.merge_cells("A2:F2")

    # Stats table at row 4
    stats_headers = ["Total Story Points", "# Days in Sprint", "Points / Day"]
    stats_values = [
        f"{burndown_data['ideal_total']:.0f}",
        str(burndown_data['sprint_days']),
        f"{burndown_data['pts_per_day']:.1f}",
    ]
    for i, h in enumerate(stats_headers):
        cell = ws_summary.cell(row=4, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = dark_blue_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for i, v in enumerate(stats_values):
        cell = ws_summary.cell(row=5, column=i + 1, value=v)
        cell.alignment = Alignment(horizontal='center')
        cell.font = bold_center
        cell.border = thin_border

    # Burndown data table starting at row 7
    bd_headers = ["Sprint Day", "Date", "Ideal", "Actual", "Completed Tasks"]
    for i, h in enumerate(bd_headers):
        cell = ws_summary.cell(row=7, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = dark_blue_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    for row_idx, (day_num, date, ideal, actual, detail) in enumerate(zip(
        burndown_data["sprint_day_nums"],
        burndown_data["real_dates"],
        burndown_data["ideal_line"],
        burndown_data["actual_line"],
        burndown_data["completed_detail"],
    ), start=8):
        ws_summary.cell(row=row_idx, column=1, value=day_num).border = thin_border
        ws_summary.cell(row=row_idx, column=2, value=date).border = thin_border
        ws_summary.cell(row=row_idx, column=3, value=ideal).border = thin_border
        actual_cell = ws_summary.cell(row=row_idx, column=4, value=actual if actual is not None else "")
        actual_cell.border = thin_border
        detail_cell = ws_summary.cell(row=row_idx, column=5, value=detail)
        detail_cell.border = thin_border
        detail_cell.alignment = Alignment(wrap_text=True)

    # Burndown line chart
    data_end_row = 7 + len(burndown_data["sprint_day_nums"])
    chart = LineChart()
    chart.title = f"Burndown: {sprint_name}"
    chart.x_axis.title = "Sprint Day"
    chart.y_axis.title = "Story Points Remaining"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    ideal_ref = Reference(ws_summary, min_col=3, min_row=7, max_row=data_end_row)
    actual_ref = Reference(ws_summary, min_col=4, min_row=7, max_row=data_end_row)
    cats = Reference(ws_summary, min_col=1, min_row=8, max_row=data_end_row)

    chart.add_data(ideal_ref, titles_from_data=True)
    chart.add_data(actual_ref, titles_from_data=True)
    chart.set_categories(cats)

    # Style chart lines
    s_ideal = chart.series[0]
    s_ideal.graphicalProperties.line.dashStyle = "dash"
    s_ideal.graphicalProperties.line.solidFill = "1F4E79"
    s_actual = chart.series[1]
    s_actual.graphicalProperties.line.solidFill = "2E7D32"

    chart_row = data_end_row + 2
    ws_summary.add_chart(chart, f"A{chart_row}")

    # Spillover & carry-in summary section below the chart
    # (spillover_gids, spillover_points, carry_in_tasks_list, carry_in_points
    #  are already extracted in the Overview sheet section above)

    # Place this section ~16 rows below chart_row (chart occupies ~14 rows)
    sci_row = chart_row + 16

    if spillover_gids or carry_in_tasks_list:
        ws_summary.cell(row=sci_row, column=1, value="Spillover & Carry-in Summary").font = subtitle_font
        ws_summary.merge_cells(start_row=sci_row, start_column=1, end_row=sci_row, end_column=5)
        sci_row += 1

        if spillover_gids:
            ws_summary.cell(
                row=sci_row, column=1,
                value=f"Spillover tasks (multi-sprint tagged): {len(spillover_gids)} tasks, {int(spillover_points)} pts"
            ).font = Font(size=11, color="1565C0")
            sci_row += 1

        if carry_in_tasks_list:
            ws_summary.cell(
                row=sci_row, column=1,
                value=f"Carry-in tasks (older sprint, completed this window): {len(carry_in_tasks_list)} tasks, {int(carry_in_points)} pts"
            ).font = Font(size=11, color="E65100")
            sci_row += 2

            # Carry-in detail table
            ci_headers = ["Task", "Assignee", "Points", "Sprint Tag", "Status", "Completed"]
            for i, h in enumerate(ci_headers):
                cell = ws_summary.cell(row=sci_row, column=i + 1, value=h)
                cell.font = header_font
                cell.fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            sci_row += 1

            for t in carry_in_tasks_list:
                try:
                    pts = float(t.story_points) if t.story_points else 0
                except (ValueError, TypeError):
                    pts = 0
                comp_date = ""
                if t.completed_at:
                    comp_date = t.completed_at[:10]
                elif getattr(t, 'status_changed_at', None):
                    comp_date = t.status_changed_at  # already YYYY-MM-DD
                elif getattr(t, 'modified_at', None):
                    comp_date = t.modified_at[:10]
                ci_vals = [
                    t.name,
                    t.assignee or "Unassigned",
                    pts,
                    t.sprint or "",
                    t.progress or "Completed",
                    comp_date,
                ]
                for i, val in enumerate(ci_vals):
                    cell = ws_summary.cell(row=sci_row, column=i + 1, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True)
                sci_row += 1

    _auto_width(ws_summary)
    ws_summary.freeze_panes = "A8"

    # ── Sheets 2-11: Per-Day Task Details ────────────────────────────────
    completion_dates = burndown_data.get("completion_dates", {})
    today_date = datetime.now().strftime("%Y-%m-%d")

    # Build due-by-date index from all sprint tasks
    due_by_date: dict[str, list] = {}
    if all_sprint_tasks:
        for task in all_sprint_tasks:
            if getattr(task, 'due_on', None):
                due_by_date.setdefault(task.due_on, []).append(task)
        # Roll weekend due dates into the following Monday
        weekend_keys = [k for k in due_by_date if datetime.strptime(k, "%Y-%m-%d").weekday() >= 5]
        for wk in weekend_keys:
            dt_wk = datetime.strptime(wk, "%Y-%m-%d")
            while dt_wk.weekday() >= 5:
                dt_wk += timedelta(days=1)
            monday = dt_wk.strftime("%Y-%m-%d")
            due_by_date.setdefault(monday, []).extend(due_by_date[wk])
            del due_by_date[wk]
    day_columns = [
        "Task Name", "Asana Link", "Assignee", "QA Assignee",
        "Story Points", "Type", "Epic", "Status",
        "Status Changed", "GitHub PR", "Comments",
    ]

    def _write_task_row(ws, row, task_obj, greyed_out=False, is_spillover=False):
        """Write a single task row. When greyed_out=True, use muted styles for due tasks."""
        # Fetch comments and extract PR links
        task_comments = []
        comments_text = ""
        pr_links = ""
        try:
            if asana_client:
                task_comments = asana_client.get_task_comments(task_obj.gid, limit=10)
                comment_lines = []
                for c in task_comments:
                    created_by = c.get('created_by') or {}
                    author = created_by.get('name')
                    if not author and asana_client and created_by.get('gid'):
                        author = asana_client.get_user_name(created_by['gid']) or 'Unknown'
                    author = author or 'Unknown'
                    raw = (c.get('text') or '')[:300]
                    text = _clean_comment_text(raw, asana_client)
                    date = (c.get('created_at') or '')[:10]
                    if text:
                        comment_lines.append(f"[{author} {date}] {text}")
                comments_text = "\n".join(comment_lines)
                pr_links = _extract_pr_links(task_obj.notes, task_comments)
        except Exception as e:
            print(f"Warning: Could not fetch comments for task {task_obj.gid}: {e}")

        pts = float(task_obj.story_points) if task_obj.story_points else 0

        task_display_name = f"[SPILLOVER] {task_obj.name}" if is_spillover else task_obj.name
        values = [
            task_display_name,                          # Task Name
            task_obj.url,                               # Asana Link
            task_obj.assignee or "Unassigned",          # Assignee
            task_obj.qa_assignee or "",                 # QA Assignee
            pts,                                        # Story Points
            task_obj.task_type or "",                   # Type
            task_obj.epic or "",                        # Epic
            task_obj.progress or "",                    # Status
            task_obj.status_changed_at or "",           # Status Changed
            pr_links,                                   # GitHub PR
            comments_text,                              # Comments
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if greyed_out:
                cell.font = due_task_font
                cell.fill = due_task_fill

        # Hyperlink for Asana Link column (col 2)
        link_cell = ws.cell(row=row, column=2)
        link_cell.hyperlink = task_obj.url
        link_cell.font = due_task_link_font if greyed_out else link_font
        link_cell.value = "Open in Asana"

        # Bold + centered for Story Points (col 5)
        pts_cell = ws.cell(row=row, column=5)
        pts_cell.font = due_task_bold if greyed_out else bold_center
        pts_cell.alignment = Alignment(horizontal='center', vertical='top')

        # Color code status cell (col 8) — skip for greyed-out rows
        if not greyed_out:
            status_cell = ws.cell(row=row, column=8)
            status = task_obj.progress or ""
            if status in status_fills:
                status_cell.fill = status_fills[status]

    for day_num, date_str in zip(burndown_data["sprint_day_nums"], burndown_data["real_dates"]):
        # Sheet name like "Day 1 (Feb 3)"
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        sheet_name = f"Day {day_num} ({dt.strftime('%b')} {dt.day})"
        ws = wb.create_sheet(title=sheet_name)

        # Classify the day
        is_today = (date_str == today_date)
        is_future = (date_str > today_date)

        # Gather completed tasks
        completed_tasks = []
        if date_str in completion_dates:
            completed_tasks = completion_dates[date_str].get("tasks", [])

        # Gather due tasks for today/future (exclude already-completed tasks)
        due_tasks = []
        if (is_today or is_future) and date_str in due_by_date:
            completed_gids = {t.gid for t in completed_tasks}
            due_tasks = [t for t in due_by_date[date_str] if t.gid not in completed_gids]

        has_completed = bool(completed_tasks)
        has_due = bool(due_tasks)
        has_any = has_completed or has_due

        # Determine header fill and label suffix
        if is_today:
            header_fill = blue_fill
            label_suffix = " (TODAY)"
        elif is_future:
            header_fill = amber_fill
            label_suffix = " (Upcoming)"
        elif has_completed:
            header_fill = green_fill
            label_suffix = ""
        else:
            header_fill = gray_fill
            label_suffix = ""

        # Header row
        header_text = f"Sprint Day {day_num} — {date_str}{label_suffix}"
        ws.cell(row=1, column=1, value=header_text).font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(day_columns))
        for col in range(1, len(day_columns) + 1):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF", size=14)

        if not has_any:
            ws.cell(row=3, column=1, value="No tasks completed or due this day").font = Font(
                italic=True, size=12, color="757575"
            )
            ws.freeze_panes = "A3"
            continue

        # Column headers at row 3
        _style_header(ws, 3, len(day_columns), dark_blue_fill)
        for i, h in enumerate(day_columns):
            ws.cell(row=3, column=i + 1, value=h)
        _style_header(ws, 3, len(day_columns), dark_blue_fill)

        current_row = 4

        # Write completed tasks (normal styling)
        for task_obj in completed_tasks:
            _write_task_row(ws, current_row, task_obj, greyed_out=False,
                            is_spillover=task_obj.gid in spillover_gids)
            current_row += 1

        # Separator row if both groups exist
        if has_completed and has_due:
            sep_cell = ws.cell(row=current_row, column=1, value="--- Due/Upcoming Tasks ---")
            sep_cell.font = section_font
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(day_columns))
            current_row += 1

        # Write due tasks (greyed out)
        for task_obj in due_tasks:
            _write_task_row(ws, current_row, task_obj, greyed_out=True,
                            is_spillover=task_obj.gid in spillover_gids)
            current_row += 1

        _auto_width(ws)
        ws.freeze_panes = "A4"

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# =============================================================================
# Quick Wins - Sprint Progress Bar
# =============================================================================

def render_sprint_progress_bar(
    results: list[TaskCompliance],
    completed_results: Optional[list[TaskCompliance]] = None,
    selected_sprint: Optional[str] = None
):
    """Render beautiful neumorphic sprint progress bar with accurate completion data."""
    # Filter by sprint if selected
    if selected_sprint:
        sprint_tasks = [t for t in results if task_in_sprint(t, selected_sprint)]
        completed_sprint_tasks = [t for t in (completed_results or []) if task_in_sprint(t, selected_sprint)]
    else:
        sprint_tasks = results
        completed_sprint_tasks = completed_results or []

    # Calculate total and completed points (same logic as burndown)
    # Tasks in Review, QA, Ready to Ship, or Done are considered "completed" for progress tracking
    completed_statuses = ("Review", "QA", "Ready to Ship", "Done")
    total_points = 0
    completed_points = 0

    # Active tasks
    for task in sprint_tasks:
        try:
            points = float(task.story_points) if task.story_points else 0
        except (ValueError, TypeError):
            points = 0
        total_points += points
        # Count Review, QA, Done status tasks as completed
        if task.progress in completed_statuses:
            completed_points += points

    # Completed tasks from Asana (truly completed)
    for task in completed_sprint_tasks:
        try:
            points = float(task.story_points) if task.story_points else 0
        except (ValueError, TypeError):
            points = 0
        total_points += points
        completed_points += points

    pct = (completed_points / total_points * 100) if total_points > 0 else 0

    st.markdown(f"""
    <div class="nm-progress-container">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px;">
            <h3 style="margin: 0; color: var(--nm-text-primary);">Sprint Progress</h3>
            <span style="font-size: 1.5rem; font-weight: 700; color: var(--nm-primary);">{pct:.0f}%</span>
        </div>
        <div class="nm-progress-bar-outer">
            <div class="nm-progress-bar-inner" style="width: {pct}%;"></div>
            <div class="nm-progress-text">{completed_points:.0f} / {total_points:.0f} pts</div>
        </div>
        <div class="nm-progress-stats">
            <span>Completed: {completed_points:.0f} pts</span>
            <span>Remaining: {total_points - completed_points:.0f} pts</span>
        </div>
    </div>
    """, unsafe_allow_html=True)


# =============================================================================
# Quick Wins - Overdue Tasks Alert
# =============================================================================

def render_overdue_alert_section(results: list[TaskCompliance]):
    """Render red alert for overdue tasks."""
    overdue_tasks = [t for t in results if getattr(t, 'is_overdue', False)]

    if not overdue_tasks:
        return

    # Sort by most overdue first (most negative days_until_due)
    overdue_tasks.sort(key=lambda t: getattr(t, 'days_until_due', 0) or 0)

    total_overdue_points = sum(
        float(t.story_points) if t.story_points else 0
        for t in overdue_tasks
    )

    st.markdown(f"""
    <div class="nm-alert nm-alert--error">
        <h3>Overdue Tasks ({len(overdue_tasks)})</h3>
        <p>{total_overdue_points:.0f} story points are past due date</p>
    </div>
    """, unsafe_allow_html=True)

    # Create header row
    header_cols = st.columns([3, 1.5, 1.5, 1, 1, 1])
    headers = ["Task Name", "Assignee", "Due Date", "Days Overdue", "Points", "Actions"]
    for i, header in enumerate(headers):
        header_cols[i].markdown(f"**{header}**")

    # Create data rows (sorted by most overdue first)
    for idx, task in enumerate(overdue_tasks):
        row_cols = st.columns([3, 1.5, 1.5, 1, 1, 1])

        # Task name (truncated)
        task_name = task.name[:35] + "..." if len(task.name) > 35 else task.name
        row_cols[0].write(task_name)

        # Assignee
        row_cols[1].write(task.assignee or "Unassigned")

        # Due Date
        row_cols[2].write(task.due_on or "-")

        # Days Overdue
        days_until = getattr(task, 'days_until_due', None)
        days_overdue = abs(days_until) if days_until is not None and days_until < 0 else 0
        row_cols[3].write(f"{days_overdue}d")

        # Points
        row_cols[4].write(task.story_points or "-")

        # Action buttons
        btn_col1, btn_col2 = row_cols[5].columns(2)
        if btn_col1.button("👁", key=f"overdue_view_{idx}", help="View in app"):
            st.session_state["selected_task_gid"] = task.gid
            st.session_state["selected_task_url"] = task.url
            st.session_state["selected_task_name"] = task.name
            st.rerun()
        btn_col2.link_button("🔗", task.url, help="Open in Asana")

    st.markdown("---")


# =============================================================================
# Quick Wins - Due This Week Alert
# =============================================================================

def render_due_this_week_section(results: list[TaskCompliance]):
    """Render amber alert for tasks due within 7 days."""
    due_soon = [
        t for t in results
        if getattr(t, 'days_until_due', None) is not None
        and 0 <= t.days_until_due <= 7
        and t.progress != "Done"
    ]

    if not due_soon:
        return

    # Sort by due date ascending (soonest first)
    due_soon.sort(key=lambda t: getattr(t, 'days_until_due', 999) or 999)

    total_due_points = sum(
        float(t.story_points) if t.story_points else 0
        for t in due_soon
    )

    st.markdown(f"""
    <div class="nm-alert nm-alert--warning">
        <h3>Due This Week ({len(due_soon)})</h3>
        <p>{total_due_points:.0f} story points due in the next 7 days</p>
    </div>
    """, unsafe_allow_html=True)

    # Create header row
    header_cols = st.columns([3, 1.5, 1.5, 1, 1, 1])
    headers = ["Task Name", "Assignee", "Due Date", "Days Left", "Points", "Actions"]
    for i, header in enumerate(headers):
        header_cols[i].markdown(f"**{header}**")

    # Create data rows
    for idx, task in enumerate(due_soon):
        row_cols = st.columns([3, 1.5, 1.5, 1, 1, 1])

        # Task name (truncated)
        task_name = task.name[:35] + "..." if len(task.name) > 35 else task.name
        row_cols[0].write(task_name)

        # Assignee
        row_cols[1].write(task.assignee or "Unassigned")

        # Due Date
        row_cols[2].write(task.due_on or "-")

        # Days Left
        days_left = getattr(task, 'days_until_due', None)
        if days_left == 0:
            row_cols[3].write("Today")
        elif days_left == 1:
            row_cols[3].write("Tomorrow")
        else:
            row_cols[3].write(f"{days_left}d")

        # Points
        row_cols[4].write(task.story_points or "-")

        # Action buttons
        btn_col1, btn_col2 = row_cols[5].columns(2)
        if btn_col1.button("👁", key=f"due_soon_view_{idx}", help="View in app"):
            st.session_state["selected_task_gid"] = task.gid
            st.session_state["selected_task_url"] = task.url
            st.session_state["selected_task_name"] = task.name
            st.rerun()
        btn_col2.link_button("🔗", task.url, help="Open in Asana")

    st.markdown("---")


# =============================================================================
# Quick Wins - Points by Assignee Chart (Stacked Bar with Invalid Detection)
# =============================================================================

# Valid Fibonacci story points
VALID_FIBONACCI_POINTS = (0, 1, 2, 3, 5, 8, 13)
# Types that should NOT have story points
TYPES_WITHOUT_POINTS = ("Epic", "Bug")


def is_invalid_story_points(task: TaskCompliance) -> bool:
    """Check if a task has invalid story points (Bug/Epic with points or non-Fibonacci)."""
    if not task.story_points:
        return False

    try:
        points = float(task.story_points)
    except (ValueError, TypeError):
        return True  # Non-numeric is invalid

    # Bug or Epic with story points
    if task.task_type in TYPES_WITHOUT_POINTS and points > 0:
        return True

    # Non-Fibonacci number
    if points != int(points) or int(points) not in VALID_FIBONACCI_POINTS:
        return True

    return False


def render_points_by_assignee_chart(
    results: list[TaskCompliance],
    completed_results: Optional[list[TaskCompliance]] = None,
    selected_sprint: Optional[str] = None
):
    """Render stacked horizontal bar chart showing completed vs remaining vs invalid points per assignee."""
    if not PLOTLY_AVAILABLE:
        st.warning("Plotly is required for charts. Install with: pip install plotly")
        return

    # Filter by sprint if selected
    if selected_sprint:
        sprint_tasks = [t for t in results if task_in_sprint(t, selected_sprint)]
        completed_sprint_tasks = [t for t in (completed_results or []) if task_in_sprint(t, selected_sprint)]
    else:
        sprint_tasks = results
        completed_sprint_tasks = completed_results or []

    # Calculate points per assignee (completed vs remaining vs invalid)
    assignee_completed = {}
    assignee_remaining = {}
    assignee_invalid = {}

    def process_task(task, is_completed_task=False):
        """Process a single task and categorize its points."""
        try:
            points = float(task.story_points) if task.story_points else 0
        except (ValueError, TypeError):
            points = 0

        if points == 0:
            return

        assignee = task.assignee or "Unassigned"

        # Initialize assignee if not seen
        if assignee not in assignee_completed:
            assignee_completed[assignee] = 0
            assignee_remaining[assignee] = 0
            assignee_invalid[assignee] = 0

        # Check if invalid (Bug/Epic with points OR non-Fibonacci)
        if is_invalid_story_points(task):
            assignee_invalid[assignee] += points
        elif is_completed_task or task.progress == "Done":
            assignee_completed[assignee] += points
        else:
            assignee_remaining[assignee] += points

    # Process active tasks
    for task in sprint_tasks:
        process_task(task, is_completed_task=False)

    # Process completed tasks from Asana
    for task in completed_sprint_tasks:
        process_task(task, is_completed_task=True)

    # Get all assignees and sort by total points
    all_assignees = set(assignee_completed.keys()) | set(assignee_remaining.keys()) | set(assignee_invalid.keys())
    if not all_assignees:
        st.info("No story points data for assignees")
        return

    assignee_totals = {
        a: assignee_completed.get(a, 0) + assignee_remaining.get(a, 0) + assignee_invalid.get(a, 0)
        for a in all_assignees
    }
    sorted_assignees = sorted(all_assignees, key=lambda a: assignee_totals[a], reverse=True)

    # Prepare data for chart
    completed_values = [assignee_completed.get(a, 0) for a in sorted_assignees]
    remaining_values = [assignee_remaining.get(a, 0) for a in sorted_assignees]
    invalid_values = [assignee_invalid.get(a, 0) for a in sorted_assignees]

    # Neumorphic colors
    nm_success = '#5B9A8B'  # Completed - green
    nm_primary = '#6B7FD7'  # Remaining - blue
    nm_error = '#C9736D'    # Invalid - red

    fig = go.Figure()

    # Completed bar
    fig.add_trace(go.Bar(
        y=sorted_assignees,
        x=completed_values,
        name='Completed',
        orientation='h',
        marker=dict(color=nm_success),
        text=[f'{v:.0f}' if v > 0 else '' for v in completed_values],
        textposition='inside',
        hovertemplate='%{y}<br>Completed: %{x:.0f} pts<extra></extra>'
    ))

    # Remaining bar
    fig.add_trace(go.Bar(
        y=sorted_assignees,
        x=remaining_values,
        name='Remaining',
        orientation='h',
        marker=dict(color=nm_primary),
        text=[f'{v:.0f}' if v > 0 else '' for v in remaining_values],
        textposition='inside',
        hovertemplate='%{y}<br>Remaining: %{x:.0f} pts<extra></extra>'
    ))

    # Invalid bar (Bug/Epic with points or non-Fibonacci)
    total_invalid = sum(invalid_values)
    if total_invalid > 0:
        fig.add_trace(go.Bar(
            y=sorted_assignees,
            x=invalid_values,
            name='Invalid',
            orientation='h',
            marker=dict(color=nm_error, pattern=dict(shape="x", size=6)),
            text=[f'{v:.0f}' if v > 0 else '' for v in invalid_values],
            textposition='inside',
            hovertemplate='%{y}<br>Invalid: %{x:.0f} pts<br>(Bug/Epic or non-Fibonacci)<extra></extra>'
        ))

    # Calculate totals for title
    total_completed = sum(completed_values)
    total_remaining = sum(remaining_values)
    total_all = total_completed + total_remaining + total_invalid

    title_text = f"Workload by Assignee ({total_completed:.0f}/{total_all:.0f} pts done)"
    if total_invalid > 0:
        title_text += f" | {total_invalid:.0f} invalid"

    fig.update_layout(
        title=dict(
            text=title_text,
            font=dict(size=16, color='#2D3748')
        ),
        barmode='stack',
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        height=max(300, len(sorted_assignees) * 40 + 100),
        paper_bgcolor='#E4E8EC',
        plot_bgcolor='#E4E8EC',
        margin=dict(t=60, b=40, l=120, r=20),
        xaxis=dict(
            title="Story Points",
            gridcolor='rgba(163, 177, 198, 0.3)',
        ),
        yaxis=dict(
            title="",
            autorange="reversed",  # Highest at top
        ),
    )

    st.plotly_chart(fig, use_container_width=True, key="points_by_assignee")

    # Show warning if there are invalid points
    if total_invalid > 0:
        st.warning(f"**{total_invalid:.0f} invalid story points detected** - Bug/Epic with points or non-Fibonacci values")


# =============================================================================
# Bug Count by Assignee Chart
# =============================================================================

def render_bug_count_chart(
    results: list[TaskCompliance],
    completed_results: Optional[list[TaskCompliance]] = None,
    selected_sprint: Optional[str] = None
) -> None:
    """Render horizontal bar chart showing bug count per assignee."""
    if not PLOTLY_AVAILABLE:
        st.warning("Plotly is required for charts. Install with: pip install plotly")
        return

    # Filter for bugs only from active results
    active_bugs = [t for t in results if t.task_type == "Bug"]

    # Filter completed results for the selected sprint
    if selected_sprint and selected_sprint != "All":
        completed_sprint_bugs = [
            t for t in (completed_results or [])
            if t.task_type == "Bug" and task_in_sprint(t, selected_sprint)
        ]
    else:
        completed_sprint_bugs = [t for t in (completed_results or []) if t.task_type == "Bug"]

    # Count bugs per assignee
    assignee_active_bugs: dict[str, int] = {}
    assignee_completed_bugs: dict[str, int] = {}

    for task in active_bugs:
        assignee = task.assignee or "Unassigned"
        if task.progress == "Done":
            assignee_completed_bugs[assignee] = assignee_completed_bugs.get(assignee, 0) + 1
        else:
            assignee_active_bugs[assignee] = assignee_active_bugs.get(assignee, 0) + 1

    for task in completed_sprint_bugs:
        assignee = task.assignee or "Unassigned"
        assignee_completed_bugs[assignee] = assignee_completed_bugs.get(assignee, 0) + 1

    # Get all assignees
    all_assignees = set(assignee_active_bugs.keys()) | set(assignee_completed_bugs.keys())

    if not all_assignees:
        st.info("No bugs found in this sprint")
        return

    # Calculate totals
    total_active = sum(assignee_active_bugs.values())
    total_completed = sum(assignee_completed_bugs.values())
    total_all = total_active + total_completed

    # Sort assignees by total bug count (descending)
    assignee_totals = {
        a: assignee_active_bugs.get(a, 0) + assignee_completed_bugs.get(a, 0)
        for a in all_assignees
    }
    sorted_assignees = sorted(all_assignees, key=lambda a: assignee_totals[a], reverse=True)

    # Prepare data for chart
    active_values = [assignee_active_bugs.get(a, 0) for a in sorted_assignees]
    completed_values = [assignee_completed_bugs.get(a, 0) for a in sorted_assignees]

    # Neumorphic colors
    nm_error = '#C9736D'    # Active bugs - coral/red
    nm_success = '#5B9A8B'  # Completed bugs - green

    fig = go.Figure()

    # Completed bugs bar (green)
    fig.add_trace(go.Bar(
        y=sorted_assignees,
        x=completed_values,
        name='Completed',
        orientation='h',
        marker=dict(color=nm_success),
        text=[f'{v}' if v > 0 else '' for v in completed_values],
        textposition='inside',
        hovertemplate='%{y}<br>Completed: %{x} bugs<extra></extra>'
    ))

    # Active bugs bar (red)
    fig.add_trace(go.Bar(
        y=sorted_assignees,
        x=active_values,
        name='Active',
        orientation='h',
        marker=dict(color=nm_error),
        text=[f'{v}' if v > 0 else '' for v in active_values],
        textposition='inside',
        hovertemplate='%{y}<br>Active: %{x} bugs<extra></extra>'
    ))

    title_text = f"Bugs by Assignee ({total_completed}/{total_all} resolved)"

    fig.update_layout(
        title=dict(
            text=title_text,
            font=dict(size=16, color='#2D3748')
        ),
        barmode='stack',
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        height=max(300, len(sorted_assignees) * 40 + 100),
        paper_bgcolor='#E4E8EC',
        plot_bgcolor='#E4E8EC',
        margin=dict(t=60, b=40, l=120, r=20),
        xaxis=dict(
            title="Bug Count",
            gridcolor='rgba(163, 177, 198, 0.3)',
            dtick=1,  # Integer tick marks for counts
        ),
        yaxis=dict(
            title="",
            autorange="reversed",  # Highest at top
        ),
    )

    st.plotly_chart(fig, use_container_width=True, key="bugs_by_assignee")


# =============================================================================
# Completion Analytics - Tasks Completed by Team
# =============================================================================

def render_team_completion_chart(
    completed_results: Optional[list[TaskCompliance]],
    filters: dict,
    selected_sprint: Optional[str] = None
) -> dict:
    """
    Render daily completion chart for the team.
    Returns daily_completions dict for consistency with individual chart.
    """
    if not PLOTLY_AVAILABLE:
        st.warning("Plotly is required for charts. Install with: pip install plotly")
        return {}

    # Get date range from filters
    completion_start = filters.get("completion_start", datetime.now().date() - timedelta(days=14))
    completion_end = filters.get("completion_end", datetime.now().date())

    # Filter completed tasks by sprint and date range
    filtered_tasks = []
    for task in (completed_results or []):
        # Skip if no completed_at date
        if not task.completed_at:
            continue

        # Parse completion date
        try:
            completed_date = datetime.strptime(task.completed_at[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue

        # Check date range
        if not (completion_start <= completed_date <= completion_end):
            continue

        # Check sprint filter
        if selected_sprint and not task_in_sprint(task, selected_sprint):
            continue

        filtered_tasks.append(task)

    if not filtered_tasks:
        st.info("No completed tasks found in the selected date range")
        return {}

    # Group tasks by completion date
    daily_completions = {}
    daily_points = {}

    for task in filtered_tasks:
        date_str = task.completed_at[:10]  # YYYY-MM-DD

        if date_str not in daily_completions:
            daily_completions[date_str] = []
            daily_points[date_str] = 0

        daily_completions[date_str].append(task)
        try:
            points = float(task.story_points) if task.story_points else 0
        except (ValueError, TypeError):
            points = 0
        daily_points[date_str] += points

    # Sort dates
    sorted_dates = sorted(daily_completions.keys())
    task_counts = [len(daily_completions[d]) for d in sorted_dates]
    point_values = [daily_points[d] for d in sorted_dates]

    # Neumorphic colors
    nm_primary = '#6B7FD7'
    nm_success = '#5B9A8B'
    nm_text_primary = '#2D3748'
    nm_bg = '#E4E8EC'

    # Create figure with secondary y-axis
    fig = go.Figure()

    # Task count bars
    fig.add_trace(go.Bar(
        x=sorted_dates,
        y=task_counts,
        name='Tasks Completed',
        marker=dict(color=nm_primary),
        text=task_counts,
        textposition='auto',
        hovertemplate='%{x}<br>Tasks: %{y}<extra></extra>'
    ))

    # Story points line
    fig.add_trace(go.Scatter(
        x=sorted_dates,
        y=point_values,
        mode='lines+markers',
        name='Story Points',
        line=dict(color=nm_success, width=3),
        marker=dict(size=8),
        yaxis='y2',
        hovertemplate='%{x}<br>Points: %{y:.0f}<extra></extra>'
    ))

    # Calculate totals
    total_tasks = sum(task_counts)
    total_points = sum(point_values)

    fig.update_layout(
        title=dict(
            text=f"Tasks Completed by Team ({total_tasks} tasks, {total_points:.0f} pts)",
            font=dict(size=16, color=nm_text_primary)
        ),
        xaxis=dict(
            title="Date",
            gridcolor='rgba(163, 177, 198, 0.3)',
        ),
        yaxis=dict(
            title="Tasks Completed",
            gridcolor='rgba(163, 177, 198, 0.3)',
            rangemode='tozero',
        ),
        yaxis2=dict(
            title="Story Points",
            overlaying='y',
            side='right',
            rangemode='tozero',
            gridcolor='rgba(163, 177, 198, 0.1)',
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        height=400,
        paper_bgcolor=nm_bg,
        plot_bgcolor=nm_bg,
        margin=dict(t=60, b=40, l=60, r=60),
        hovermode="x unified",
    )

    st.plotly_chart(fig, use_container_width=True, key="team_completion_chart")

    return daily_completions


# =============================================================================
# Completion Analytics - Tasks Completed by Individuals
# =============================================================================

def render_individual_completion_chart(
    completed_results: Optional[list[TaskCompliance]],
    filters: dict,
    selected_sprint: Optional[str] = None,
    team_data: Optional[dict] = None
) -> None:
    """
    Render daily completion chart grouped by individual assignees.
    Uses same filtered data as team chart to ensure totals match.
    """
    if not PLOTLY_AVAILABLE:
        st.warning("Plotly is required for charts. Install with: pip install plotly")
        return

    # Use team_data if provided (ensures consistency), otherwise filter again
    if team_data:
        daily_completions = team_data
    else:
        # Get date range from filters
        completion_start = filters.get("completion_start", datetime.now().date() - timedelta(days=14))
        completion_end = filters.get("completion_end", datetime.now().date())

        # Filter completed tasks by sprint and date range
        daily_completions = {}
        for task in (completed_results or []):
            if not task.completed_at:
                continue
            try:
                completed_date = datetime.strptime(task.completed_at[:10], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            if not (completion_start <= completed_date <= completion_end):
                continue
            if selected_sprint and not task_in_sprint(task, selected_sprint):
                continue

            date_str = task.completed_at[:10]
            if date_str not in daily_completions:
                daily_completions[date_str] = []
            daily_completions[date_str].append(task)

    if not daily_completions:
        st.info("No completed tasks found in the selected date range")
        return

    # Get all unique assignees and dates
    all_assignees = set()
    for tasks in daily_completions.values():
        for task in tasks:
            all_assignees.add(task.assignee or "Unassigned")

    sorted_dates = sorted(daily_completions.keys())
    sorted_assignees = sorted(all_assignees)

    # Build data matrix: assignee -> date -> count
    assignee_date_counts = {a: {d: 0 for d in sorted_dates} for a in sorted_assignees}
    for date_str, tasks in daily_completions.items():
        for task in tasks:
            assignee = task.assignee or "Unassigned"
            assignee_date_counts[assignee][date_str] += 1

    # Neumorphic color palette for assignees
    assignee_colors = [
        '#6B7FD7', '#5B9A8B', '#C9736D', '#D4A574', '#5A9AA8',
        '#9B7ED9', '#7DB87D', '#E88E8E', '#F0B86E', '#6EC8D7',
        '#C97BAF', '#85BB65', '#E8A07A', '#7B9FD4', '#D49A6A',
    ]

    nm_text_primary = '#2D3748'
    nm_bg = '#E4E8EC'

    fig = go.Figure()

    # Add stacked bars for each assignee
    for i, assignee in enumerate(sorted_assignees):
        counts = [assignee_date_counts[assignee][d] for d in sorted_dates]
        color = assignee_colors[i % len(assignee_colors)]

        fig.add_trace(go.Bar(
            x=sorted_dates,
            y=counts,
            name=assignee,
            marker=dict(color=color),
            hovertemplate=f'{assignee}<br>%{{x}}<br>Tasks: %{{y}}<extra></extra>'
        ))

    # Calculate totals for verification
    total_individual = sum(
        sum(assignee_date_counts[a][d] for d in sorted_dates)
        for a in sorted_assignees
    )
    total_team = sum(len(tasks) for tasks in daily_completions.values())

    title_text = f"Tasks Completed by Individual ({total_individual} tasks)"

    fig.update_layout(
        title=dict(
            text=title_text,
            font=dict(size=16, color=nm_text_primary)
        ),
        barmode='stack',
        xaxis=dict(
            title="Date",
            gridcolor='rgba(163, 177, 198, 0.3)',
        ),
        yaxis=dict(
            title="Tasks Completed",
            gridcolor='rgba(163, 177, 198, 0.3)',
            rangemode='tozero',
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            font=dict(size=10)
        ),
        height=400,
        paper_bgcolor=nm_bg,
        plot_bgcolor=nm_bg,
        margin=dict(t=80, b=40, l=60, r=20),
        hovermode="x unified",
    )

    st.plotly_chart(fig, use_container_width=True, key="individual_completion_chart")

    # Verification message
    if total_individual == total_team:
        st.success(f"Totals verified: {total_individual} tasks match between team and individual charts")
    else:
        st.warning(f"Total mismatch: Team={total_team}, Individual={total_individual}")


# =============================================================================
# Quick Wins - Invalid Story Points Section
# =============================================================================

def get_invalid_reason(task: TaskCompliance) -> Optional[str]:
    """Get the reason why a task has invalid story points."""
    if not task.story_points:
        return None

    try:
        points = float(task.story_points)
    except (ValueError, TypeError):
        return "Non-numeric value"

    # Bug or Epic with story points
    if task.task_type in TYPES_WITHOUT_POINTS and points > 0:
        return f"{task.task_type} should not have points"

    # Non-Fibonacci number
    if points != int(points) or int(points) not in VALID_FIBONACCI_POINTS:
        return f"Non-Fibonacci value ({task.story_points})"

    return None


def render_invalid_story_points_section(
    results: list[TaskCompliance],
    completed_results: Optional[list[TaskCompliance]] = None,
    filters: Optional[dict] = None
):
    """Render section showing all tasks with invalid story points (including completed)."""
    filters = filters or {}
    selected_sprint = filters.get("sprint")
    selected_assignees = filters.get("assignees")
    selected_statuses = filters.get("statuses")

    # DEBUG: Show what filters are being applied
    st.caption(f"DEBUG - Filters: sprint={selected_sprint}, assignees={selected_assignees}, statuses={selected_statuses}")

    # results is already filtered, just use it directly
    filtered_active_tasks = results

    # Apply all filters to completed_results
    filtered_completed_tasks = completed_results or []
    before_filter_count = len(filtered_completed_tasks)
    if selected_sprint and selected_sprint != "All":
        filtered_completed_tasks = [t for t in filtered_completed_tasks if task_in_sprint(t, selected_sprint)]
    if selected_assignees and len(selected_assignees) > 0:
        filtered_completed_tasks = [t for t in filtered_completed_tasks if t.assignee in selected_assignees]
    if selected_statuses and len(selected_statuses) > 0:
        filtered_completed_tasks = [t for t in filtered_completed_tasks if t.progress in selected_statuses]

    st.caption(f"DEBUG - Completed tasks: {before_filter_count} -> {len(filtered_completed_tasks)} after filter")

    # Combine all tasks and find invalid ones
    all_tasks = filtered_active_tasks + filtered_completed_tasks
    invalid_tasks = []

    for task in all_tasks:
        reason = get_invalid_reason(task)
        if reason:
            invalid_tasks.append((task, reason))

    if not invalid_tasks:
        return

    # Sort by assignee, then by points descending
    invalid_tasks.sort(key=lambda x: (x[0].assignee or "ZZZ", -(float(x[0].story_points or 0))))

    # Calculate total invalid points
    total_invalid_points = sum(
        float(t.story_points) if t.story_points else 0
        for t, _ in invalid_tasks
    )

    # Group by reason type for summary
    bugs_with_points = sum(1 for _, r in invalid_tasks if "Bug" in r)
    epics_with_points = sum(1 for _, r in invalid_tasks if "Epic" in r)
    non_fibonacci = sum(1 for _, r in invalid_tasks if "Fibonacci" in r)

    # Build summary text
    summary_parts = []
    if bugs_with_points:
        summary_parts.append(f"{bugs_with_points} Bugs with points")
    if epics_with_points:
        summary_parts.append(f"{epics_with_points} Epics with points")
    if non_fibonacci:
        summary_parts.append(f"{non_fibonacci} non-Fibonacci values")

    st.markdown(f"""
    <div class="nm-alert nm-alert--error">
        <h3>Invalid Story Points ({len(invalid_tasks)} tasks)</h3>
        <p>{total_invalid_points:.0f} points are invalid: {', '.join(summary_parts)}</p>
    </div>
    """, unsafe_allow_html=True)

    # Create header row
    header_cols = st.columns([2.5, 1.2, 0.8, 0.8, 2, 0.8, 0.8])
    headers = ["Task Name", "Assignee", "Type", "Points", "Issue", "Status", "Actions"]
    for i, header in enumerate(headers):
        header_cols[i].markdown(f"**{header}**")

    # Create data rows
    for idx, (task, reason) in enumerate(invalid_tasks):
        row_cols = st.columns([2.5, 1.2, 0.8, 0.8, 2, 0.8, 0.8])

        # Task name (truncated)
        task_name = task.name[:30] + "..." if len(task.name) > 30 else task.name
        row_cols[0].write(task_name)

        # Assignee
        row_cols[1].write(task.assignee or "Unassigned")

        # Type
        row_cols[2].write(task.task_type or "-")

        # Points
        row_cols[3].write(task.story_points or "-")

        # Issue reason (highlighted)
        row_cols[4].markdown(f"**:red[{reason}]**")

        # Status (show if completed)
        status = task.progress or "Done"
        if task in filtered_completed_tasks:
            status = "Completed"
        row_cols[5].write(status)

        # Action buttons
        btn_col1, btn_col2 = row_cols[6].columns(2)
        if btn_col1.button("👁", key=f"invalid_view_{idx}", help="View in app"):
            st.session_state["selected_task_gid"] = task.gid
            st.session_state["selected_task_url"] = task.url
            st.session_state["selected_task_name"] = task.name
            st.rerun()
        btn_col2.link_button("🔗", task.url, help="Open in Asana")

    st.markdown("---")


# =============================================================================
# Alert Sections
# =============================================================================

def get_missing_fields(task: TaskCompliance) -> list[str]:
    """Get list of missing mandatory fields for a task."""
    missing = []
    if task.missing_epic:
        missing.append("Epic")
    if task.missing_sprint:
        missing.append("Sprint")
    if task.missing_type:
        missing.append("Type")
    if task.missing_points:
        missing.append("Story Points")
    if task.invalid_points:
        missing.append("Invalid Points")
    if task.missing_severity:
        missing.append("Severity")
    if task.missing_due_date:
        missing.append("Due Date")
    if task.missing_description:
        missing.append("Description/ACs")
    return missing


def get_all_issues(task: TaskCompliance) -> list[str]:
    """Get list of all compliance issues including rule violations."""
    issues = get_missing_fields(task)
    # Use getattr for backward compatibility with cached TaskCompliance objects
    rule_violations = getattr(task, 'rule_violations', [])
    if rule_violations:
        issues.extend(rule_violations)
    return issues


def render_red_alert_section(results: list[TaskCompliance]):
    """Render red alert for Review/QA tasks with issues."""
    # Filter: Review or QA with any compliance issue (including rule violations)
    red_tasks = [
        t for t in results
        if t.progress in ("Review", "QA")
        and (t.mandatory_count > 0 or t.missing_daily_update or getattr(t, 'rule_violations', []))
    ]

    if not red_tasks:
        return  # Don't show section if no issues

    st.markdown("""
    <div class="nm-alert nm-alert--error">
        <h3>🔴 Critical - Review/QA Tasks Need Attention</h3>
        <p>These tasks are in final stages but have issues that may block release</p>
    </div>
    """, unsafe_allow_html=True)

    # Create header row
    header_cols = st.columns([3, 1.5, 1, 2, 1, 1])
    headers = ["Task Name", "Assignee", "Status", "Issues", "Hours Since Update", "Actions"]
    for i, header in enumerate(headers):
        header_cols[i].markdown(f"**{header}**")

    # Create data rows
    for idx, task in enumerate(red_tasks):
        row_cols = st.columns([3, 1.5, 1, 2, 1, 1])

        # Task name (truncated)
        task_name = task.name[:35] + "..." if len(task.name) > 35 else task.name
        row_cols[0].write(task_name)

        # Assignee
        row_cols[1].write(task.assignee or "Unassigned")

        # Status
        row_cols[2].write(task.progress or "-")

        # Issues
        issues = []
        if task.missing_daily_update:
            issues.append("No daily update")
        missing = get_missing_fields(task)
        if missing:
            issues.append(f"Missing: {', '.join(missing[:2])}" + ("..." if len(missing) > 2 else ""))
        task_rule_violations = getattr(task, 'rule_violations', [])
        if task_rule_violations:
            issues.append(f"Rules: {', '.join(task_rule_violations[:1])}" + ("..." if len(task_rule_violations) > 1 else ""))
        row_cols[3].write("; ".join(issues) if issues else "-")

        # Hours since update
        hours = "-"
        if task.hours_since_update is not None:
            hours = f"{task.hours_since_update:.0f}h"
        row_cols[4].write(hours)

        # Action buttons
        btn_col1, btn_col2 = row_cols[5].columns(2)
        if btn_col1.button("👁", key=f"red_view_{idx}", help="View in app"):
            st.session_state["selected_task_gid"] = task.gid
            st.session_state["selected_task_url"] = task.url
            st.session_state["selected_task_name"] = task.name
            st.rerun()
        btn_col2.link_button("🔗",task.url, help="Open in Asana")

    st.markdown("---")


def render_amber_alert_section(results: list[TaskCompliance]):
    """Render amber alert for To Do/In Progress tasks missing details or with rule violations."""
    # Filter: To Do or In Progress with missing mandatory fields or rule violations
    amber_tasks = [
        t for t in results
        if t.progress in ("To Do", "In Progress")
        and (t.mandatory_count > 0 or getattr(t, 'rule_violations', []))
    ]

    if not amber_tasks:
        return  # Don't show section if no issues

    st.markdown("""
    <div class="nm-alert nm-alert--warning">
        <h3>⚠️ Action Required - Tasks Need Attention</h3>
        <p>These tasks in To Do/In Progress have missing fields or rule violations</p>
    </div>
    """, unsafe_allow_html=True)

    # Create header row
    header_cols = st.columns([3, 1.5, 1, 3, 1])
    headers = ["Task Name", "Assignee", "Status", "Issues", "Actions"]
    for i, header in enumerate(headers):
        header_cols[i].markdown(f"**{header}**")

    # Create data rows
    for idx, task in enumerate(amber_tasks):
        row_cols = st.columns([3, 1.5, 1, 3, 1])

        # Task name (truncated)
        task_name = task.name[:35] + "..." if len(task.name) > 35 else task.name
        row_cols[0].write(task_name)

        # Assignee
        row_cols[1].write(task.assignee or "Unassigned")

        # Status
        row_cols[2].write(task.progress or "-")

        # Issues (missing fields + rule violations)
        all_issues = get_all_issues(task)
        row_cols[3].write(", ".join(all_issues) if all_issues else "-")

        # Action buttons
        btn_col1, btn_col2 = row_cols[4].columns(2)
        if btn_col1.button("👁", key=f"amber_view_{idx}", help="View in app"):
            st.session_state["selected_task_gid"] = task.gid
            st.session_state["selected_task_url"] = task.url
            st.session_state["selected_task_name"] = task.name
            st.rerun()
        btn_col2.link_button("🔗",task.url, help="Open in Asana")

    st.markdown("---")


# =============================================================================
# Compliance Tables
# =============================================================================

def render_attributes_summary(summary: ReportSummary):
    """Render mandatory attributes summary."""
    st.subheader("Mandatory Attributes Missing/Invalid")

    # Use getattr for backward compatibility with cached summaries
    rule_violations_count = getattr(summary, 'rule_violations', 0)

    attrs = [
        ("Epic", summary.missing_epic, "🟠"),
        ("Sprint", summary.missing_sprint, "🟠"),
        ("Type", summary.missing_type, "🟠"),
        ("Story Points", summary.missing_points, "🟡"),
        ("Invalid Points", summary.invalid_points, "🟡"),
        ("Severity", summary.missing_severity, "🟡"),
        ("Due Date", summary.missing_due_date, "🟡"),
        ("Description", summary.missing_description, "🟡"),
        ("Rule Violations", rule_violations_count, "🔴"),
    ]

    cols = st.columns(3)
    for i, (name, count, icon) in enumerate(attrs):
        pct = (count / summary.total_tasks * 100) if summary.total_tasks > 0 else 0
        with cols[i % 3]:
            delta_color = "off" if count == 0 else "inverse"
            label = f"{icon} {name}" if count > 0 else f"✅ {name}"
            st.metric(label=label, value=count, delta=f"{pct:.1f}%", delta_color=delta_color)


def render_assignee_table(summary: ReportSummary):
    """Render compliance by assignee."""
    st.subheader("Compliance by Assignee")

    if not summary.by_assignee:
        st.info("No assignee data")
        return

    data = []
    for assignee, info in summary.by_assignee.items():
        total = info["total"]
        issues = info["issues"]
        compliant = total - issues
        rate = (compliant / total * 100) if total > 0 else 100
        data.append({
            "Assignee": assignee,
            "Tasks": total,
            "Compliant": compliant,
            "Issues": issues,
            "Compliance": f"{rate:.0f}%"
        })

    st.dataframe(data, use_container_width=True, hide_index=True)


def render_task_table(tasks: list[TaskCompliance], title: str, columns: list[str], table_key: str = ""):
    """Render a task table with expander and view buttons."""
    if not tasks:
        return

    with st.expander(f"{title} ({len(tasks)} tasks)", expanded=False):
        # Create header row
        header_cols = st.columns([3, 2, 1, 1, 1, 1])
        col_names = ["Task", "Assignee", "Progress", "Sprint", "Due Date", "Actions"]
        for i, col_name in enumerate(col_names):
            if i < len(columns) or col_name == "Actions":
                header_cols[i].markdown(f"**{col_name}**")

        # Create data rows with view buttons
        for idx, t in enumerate(tasks):
            row_cols = st.columns([3, 2, 1, 1, 1, 1])

            task_name = t.name[:40] + "..." if len(t.name) > 40 else t.name
            row_cols[0].write(task_name)
            row_cols[1].write(t.assignee or "Unassigned")
            row_cols[2].write(t.progress or "-")
            row_cols[3].write(t.sprint or "-")
            row_cols[4].write(t.due_on or "-")

            # Action buttons
            btn_col1, btn_col2 = row_cols[5].columns(2)
            if btn_col1.button("👁", key=f"view_{table_key}_{idx}", help="View in app"):
                st.session_state["selected_task_gid"] = t.gid
                st.session_state["selected_task_url"] = t.url
                st.session_state["selected_task_name"] = t.name
                st.rerun()
            btn_col2.link_button("🔗",t.url, help="Open in Asana")


def render_rule_violations_table(tasks: list[TaskCompliance], table_key: str = "rule_violations"):
    """Render a table for tasks with rule violations showing Type and Story Points."""
    if not tasks:
        return

    with st.expander(f"🔴 Rule Violations - Epics/Bugs with Story Points ({len(tasks)} tasks)", expanded=False):
        # Create header row
        header_cols = st.columns([3, 1.5, 1, 1, 2, 1])
        col_names = ["Task", "Assignee", "Type", "Points", "Violation", "Actions"]
        for i, col_name in enumerate(col_names):
            header_cols[i].markdown(f"**{col_name}**")

        # Create data rows with view buttons
        for idx, t in enumerate(tasks):
            row_cols = st.columns([3, 1.5, 1, 1, 2, 1])

            task_name = t.name[:40] + "..." if len(t.name) > 40 else t.name
            row_cols[0].write(task_name)
            row_cols[1].write(t.assignee or "Unassigned")
            row_cols[2].write(t.task_type or "-")
            row_cols[3].write(t.story_points or "-")
            violations = getattr(t, 'rule_violations', [])
            row_cols[4].write(", ".join(violations) if violations else "-")

            # Action buttons
            btn_col1, btn_col2 = row_cols[5].columns(2)
            if btn_col1.button("👁", key=f"view_{table_key}_{idx}", help="View in app"):
                st.session_state["selected_task_gid"] = t.gid
                st.session_state["selected_task_url"] = t.url
                st.session_state["selected_task_name"] = t.name
                st.rerun()
            btn_col2.link_button("🔗",t.url, help="Open in Asana")


def render_compliance_details(results: list[TaskCompliance]):
    """Render detailed compliance findings."""
    st.markdown("""
    <div class="nm-section-compliance">
        <h3>📋 Compliance Details</h3>
        <p style="color: #5A6778; margin: 0; font-size: 0.9rem;">Detailed breakdown of tasks with missing or invalid fields</p>
    </div>
    """, unsafe_allow_html=True)

    # Rule Violations (Critical - should be addressed first)
    rule_violations = [t for t in results if getattr(t, 'rule_violations', [])]
    if rule_violations:
        render_rule_violations_table(rule_violations)

    # Missing Daily Updates (Critical)
    missing_updates = [t for t in results if t.missing_daily_update]
    if missing_updates:
        render_task_table(missing_updates, "🔴 Missing Daily Updates", ["Task", "Assignee", "Progress"], "updates")

    # Missing Epic
    missing_epic = [t for t in results if t.missing_epic]
    if missing_epic:
        render_task_table(missing_epic, "🟠 Missing Epic", ["Task", "Assignee", "Progress"], "epic")

    # Missing Sprint
    missing_sprint = [t for t in results if t.missing_sprint]
    if missing_sprint:
        render_task_table(missing_sprint, "🟠 Missing Sprint", ["Task", "Assignee", "Progress"], "sprint")

    # Missing Type
    missing_type = [t for t in results if t.missing_type]
    if missing_type:
        render_task_table(missing_type, "🟠 Missing Type", ["Task", "Assignee", "Progress"], "type")

    # Missing Story Points
    missing_points = [t for t in results if t.missing_points]
    if missing_points:
        render_task_table(missing_points, "🟡 Missing Story Points", ["Task", "Assignee", "Progress"], "points")

    # Invalid Story Points (non-Fibonacci)
    invalid_points = [t for t in results if t.invalid_points]
    if invalid_points:
        render_task_table(invalid_points, "🟡 Invalid Story Points (non-Fibonacci)", ["Task", "Assignee", "Progress"], "invalid_points")

    # Missing Severity
    missing_severity = [t for t in results if t.missing_severity]
    if missing_severity:
        render_task_table(missing_severity, "🟡 Missing Severity", ["Task", "Assignee", "Progress"], "severity")

    # Missing Due Date
    missing_due = [t for t in results if t.missing_due_date]
    if missing_due:
        render_task_table(missing_due, "🟡 Missing Due Date", ["Task", "Assignee", "Sprint"], "due")

    # Missing Description
    missing_desc = [t for t in results if t.missing_description]
    if missing_desc:
        render_task_table(missing_desc, "🟡 Missing Description/ACs", ["Task", "Assignee", "Progress"], "desc")

    # Show message if all compliant
    all_issues = (rule_violations + missing_updates + missing_epic + missing_sprint + missing_type +
                  missing_points + invalid_points + missing_severity + missing_due + missing_desc)
    if not all_issues:
        st.success("All tasks are fully compliant! No missing fields or rule violations.")


# =============================================================================
# Download Buttons
# =============================================================================

def render_download_buttons(
    results: list[TaskCompliance],
    summary: ReportSummary,
    config: Config,
    completed_results: Optional[list[TaskCompliance]] = None,
    filters: Optional[dict] = None
):
    """Render download buttons."""
    st.subheader("Download Report")

    # Apply filters to completed_results for Excel report
    filters = filters or {}
    filtered_completed = completed_results or []
    selected_sprint = filters.get("sprint")
    selected_assignees = filters.get("assignees")
    selected_statuses = filters.get("statuses")

    if selected_sprint and selected_sprint != "All":
        filtered_completed = [t for t in filtered_completed if task_in_sprint(t, selected_sprint)]
    if selected_assignees and len(selected_assignees) > 0:
        filtered_completed = [t for t in filtered_completed if t.assignee in selected_assignees]
    if selected_statuses and len(selected_statuses) > 0:
        filtered_completed = [t for t in filtered_completed if t.progress in selected_statuses]

    col1, col2, col3 = st.columns(3)

    with col1:
        md_generator = MarkdownReportGenerator(config)
        markdown_content = md_generator.generate(results, summary)
        st.download_button(
            label="Download Markdown",
            data=markdown_content,
            file_name=f"compliance_{summary.report_date}.md",
            mime="text/markdown",
        )

    with col2:
        json_generator = JSONReportGenerator(config)
        json_content = json_generator.generate(results, summary)
        st.download_button(
            label="Download JSON",
            data=json_content,
            file_name=f"compliance_{summary.report_date}.json",
            mime="application/json",
        )

    with col3:
        if OPENPYXL_AVAILABLE:
            from asana_daily_report import ExcelReportGenerator
            excel_generator = ExcelReportGenerator(config)
            # Use generate_with_completed to include invalid points analysis
            if filtered_completed:
                workbook = excel_generator.generate_with_completed(results, filtered_completed, summary)
            else:
                workbook = excel_generator.generate(results, summary)
            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            st.download_button(
                label="Download Excel",
                data=buffer,
                file_name=f"compliance_{summary.report_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.button("Download Excel", disabled=True, help="Requires openpyxl")


# =============================================================================
# Main App
# =============================================================================

def render_homepage():
    """Render the landing page before report generation."""
    # Hero section with logo and title
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        # Logo
        logo_path = os.path.join(os.path.dirname(__file__), "assets", "Text-Logo_SourceHub.png")
        if os.path.exists(logo_path):
            st.image(logo_path, width=280)

        st.markdown("""
        <div style="text-align: center; padding: 10px 20px 20px 20px;">
            <h1 style="font-size: 2.5rem; font-weight: 700; color: #2D3748; margin: 0; letter-spacing: -1px;">
                Sprint Dashboard
            </h1>
            <p style="font-size: 1rem; color: #5A6778; margin-top: 8px;">
                Development Team Compliance & Burndown Tracking
            </p>
        </div>
        """, unsafe_allow_html=True)

    # Feature cards
    st.markdown("""
    <div style="display: flex; justify-content: center; gap: 20px; flex-wrap: wrap; padding: 30px 20px;">
        <div style="background: linear-gradient(135deg, #E4E8F0 0%, #DCE2EC 100%);
                    border-radius: 16px; padding: 24px; width: 200px; text-align: center;
                    box-shadow: 6px 6px 12px #A3B1C6, -6px -6px 12px #FFFFFF;">
            <div style="font-size: 2rem; margin-bottom: 8px; color: #6B7FD7;">&#x2713;</div>
            <div style="font-weight: 600; color: #2D3748; margin-bottom: 4px;">Compliance</div>
            <div style="font-size: 0.85rem; color: #5A6778;">Track task compliance & missing fields</div>
        </div>
        <div style="background: linear-gradient(135deg, #E4F0E8 0%, #DCE8E2 100%);
                    border-radius: 16px; padding: 24px; width: 200px; text-align: center;
                    box-shadow: 6px 6px 12px #A3B1C6, -6px -6px 12px #FFFFFF;">
            <div style="font-size: 2rem; margin-bottom: 8px; color: #5B9A8B;">&#x2197;</div>
            <div style="font-weight: 600; color: #2D3748; margin-bottom: 4px;">Burndown</div>
            <div style="font-size: 0.85rem; color: #5A6778;">Visualize sprint progress & velocity</div>
        </div>
        <div style="background: linear-gradient(135deg, #F0E8E4 0%, #E8E2DC 100%);
                    border-radius: 16px; padding: 24px; width: 200px; text-align: center;
                    box-shadow: 6px 6px 12px #A3B1C6, -6px -6px 12px #FFFFFF;">
            <div style="font-size: 2rem; margin-bottom: 8px; color: #C9736D;">&#x26A0;</div>
            <div style="font-weight: 600; color: #2D3748; margin-bottom: 4px;">Alerts</div>
            <div style="font-size: 0.85rem; color: #5A6778;">Identify blockers & action items</div>
        </div>
    </div>
    """, unsafe_allow_html=True)


def main():
    """Main application."""
    init_session_state()

    # Check if passcode is required and user is not authenticated
    if get_app_passcode() and not st.session_state.get("authenticated", False):
        render_login_screen()
        return

    # Sidebar - always render for configuration
    config_options = render_sidebar()

    # PRIORITY: Check if generating - show ONLY loader, nothing else
    if st.session_state.get("is_generating", False):
        # Neumorphic loader container with status
        st.markdown("""
        <style>
            @keyframes spin {
                0% { transform: rotate(0deg); }
                100% { transform: rotate(360deg); }
            }
        </style>
        <div style="display: flex; flex-direction: column; align-items: center; justify-content: center;
                    min-height: 50vh; text-align: center;">
            <div style="background: #E4E8EC; border-radius: 20px; padding: 40px 50px;
                        box-shadow: 8px 8px 16px #A3B1C6, -8px -8px 16px #FFFFFF;">
                <div style="width: 60px; height: 60px; margin: 0 auto 20px auto;
                            border: 4px solid #E4E8EC; border-top: 4px solid #6B7FD7;
                            border-radius: 50%; animation: spin 1s linear infinite;
                            box-shadow: inset 2px 2px 4px #A3B1C6, inset -2px -2px 4px #FFFFFF;">
                </div>
                <div style="font-size: 1.2rem; color: #2D3748; font-weight: 600; margin-bottom: 8px;">
                    Generating Report
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            try:
                with st.status("Loading...", expanded=True) as status:
                    st.write("Initializing compliance reporter...")
                    config = Config(
                        min_description_length=config_options["min_description_length"],
                        hours_without_update=config_options["hours_without_update"],
                    )
                    reporter = AsanaComplianceReporter(config_options["token"], config)

                    st.write("Fetching active tasks from Asana...")
                    tasks = reporter.client.get_tasks(completed=False)
                    st.write(f"Found {len(tasks)} active tasks")

                    completed_tasks = []
                    if config_options["fetch_completed"]:
                        st.write("Fetching completed tasks from last 30 days...")
                        completed_tasks = reporter.client.get_completed_tasks(since_days=30)
                        st.write(f"Found {len(completed_tasks)} completed tasks")

                    st.write("Analyzing task compliance...")
                    results = reporter.analyzer.analyze_all(
                        tasks,
                        fetch_comments=config_options["fetch_comments"]
                    )

                    completed_results = []
                    if completed_tasks:
                        st.write("Analyzing completed tasks...")
                        completed_results = reporter.analyzer.analyze_all(
                            completed_tasks,
                            fetch_comments=False,
                            include_done=True
                        )

                    st.write("Generating summary report...")
                    summary = reporter.analyzer.generate_summary(results)

                    # Store results
                    st.session_state["results"] = results
                    st.session_state["completed_results"] = completed_results
                    st.session_state["summary"] = summary
                    st.session_state["config"] = config
                    st.session_state["reporter"] = reporter
                    st.session_state["report_generated"] = True
                    st.session_state["is_generating"] = False

                    status.update(label="Report generated!", state="complete", expanded=False)

                st.rerun()
            except Exception as e:
                st.session_state["is_generating"] = False
                error_str = str(e).lower()
                if any(x in error_str for x in ["401", "403", "unauthorized", "forbidden"]):
                    st.error("Authentication failed. Please check your access token.")
                elif "rate limit" in error_str or "429" in error_str:
                    st.error("Rate limit exceeded. Please wait and try again.")
                else:
                    st.error(f"Error generating report: {e}")
        st.stop()  # Ensure nothing else renders
        return

    # Check token
    if not config_options["token"]:
        render_homepage()
        st.markdown("""
        <div style="text-align: center; padding: 20px;">
            <div style="background: linear-gradient(135deg, #F5F0E0 0%, #EDE8D4 100%);
                        border-radius: 12px; padding: 20px; display: inline-block;
                        border-left: 4px solid #D4A574;
                        box-shadow: 4px 4px 8px #A3B1C6, -4px -4px 8px #FFFFFF;">
                <p style="color: #7A6830; margin: 0; font-size: 0.95rem;">
                    <span style="color: #D4A574;">&#x26A0;</span> Please enter your <strong>Asana Access Token</strong> in the sidebar to get started.
                </p>
                <p style="color: #5A6778; margin: 8px 0 0 0; font-size: 0.85rem;">
                    <a href="https://app.asana.com/0/developer-console" target="_blank" style="color: #6B7FD7;">
                        Get your token from Asana Developer Console &#x2192;
                    </a>
                </p>
            </div>
        </div>
        """, unsafe_allow_html=True)
        return

    # Show homepage with Generate button if report not generated
    if not st.session_state.get("report_generated"):
        render_homepage()

        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button("Generate Report", type="primary", use_container_width=True):
                st.session_state["is_generating"] = True
                st.rerun()
        return

    # Report is generated - show dashboard header
    st.markdown("""
    <div style="display: flex; align-items: center; gap: 12px; margin-bottom: 8px;">
        <h1 style="font-size: 2rem; font-weight: 700; color: #2D3748; margin: 0;">
            Sprint Dashboard
        </h1>
    </div>
    """, unsafe_allow_html=True)
    st.caption("SourceHub Development Team")

    # Report is generated - show dashboard
    results = st.session_state["results"]
    completed_results = st.session_state.get("completed_results", [])
    summary = st.session_state["summary"]
    config = st.session_state["config"]
    reporter = st.session_state["reporter"]

    # Check if task viewer dialog should be opened
    if st.session_state.get("selected_task_gid"):
        show_task_dialog(
            st.session_state["selected_task_gid"],
            st.session_state.get("selected_task_url", ""),
            st.session_state.get("selected_task_name", "Task"),
            reporter
        )
        # Clear the selection after dialog is shown
        st.session_state["selected_task_gid"] = None
        st.session_state["selected_task_url"] = None
        st.session_state["selected_task_name"] = None

    # Dashboard filters (horizontal layout)
    filters = render_dashboard_filters(results, completed_results, reporter.analyzer)

    # Apply filters
    filtered_results = reporter.analyzer.filter_results(
        results,
        sprint=filters.get("sprint"),
        assignees=filters.get("assignees"),
        statuses=filters.get("statuses"),
    )
    filtered_summary = reporter.analyzer.generate_summary(filtered_results)
    metrics = reporter.analyzer.calculate_sprint_metrics(filtered_results)

    # Report info
    st.caption(f"Report Date: {summary.report_date} | Showing: {len(filtered_results)} tasks")

    st.markdown("---")

    # Metric cards
    render_metric_cards(filtered_summary, metrics)

    st.markdown("---")

    # Sprint Progress Bar (Quick Wins)
    render_sprint_progress_bar(filtered_results, completed_results, filters.get("sprint"))

    # Charts row: Burndown and Points by Assignee side by side
    col_burndown, col_assignee = st.columns([3, 2])

    with col_burndown:
        target_sprint_points = st.number_input(
            "Target Sprint Points",
            min_value=0,
            value=0,
            step=1,
            help="Total story points the team aims to complete (ideal: 13 per developer). 0 = auto from tasks.",
            key="target_sprint_points"
        )
        # Burndown chart
        burndown_data = render_burndown_chart(
            filtered_results, completed_results, filters.get("sprint"),
            target_sprint_points=target_sprint_points if target_sprint_points > 0 else None,
            all_results=results,
        )

    with col_assignee:
        # Points by Assignee Chart (Quick Wins)
        render_points_by_assignee_chart(filtered_results, completed_results, filters.get("sprint"))

    # Burndown summary table — full page width (outside column layout)
    if burndown_data:
        bd = burndown_data
        # Expand completed tasks into separate rows so nothing is truncated
        table_rows = []
        for day_num, date, ideal, actual, detail in zip(
            bd["sprint_day_nums"],
            bd["real_dates"],
            bd["ideal_line"],
            bd["actual_line"],
            bd["completed_detail"],
        ):
            tasks = [t.strip() for t in detail.split(";") if t.strip()] if detail else []
            if tasks:
                for i, task in enumerate(tasks):
                    table_rows.append({
                        "Sprint Day": day_num if i == 0 else "",
                        "Date": date if i == 0 else "",
                        "Ideal Remaining (pts)": ideal if i == 0 else "",
                        "Actual Remaining (pts)": (actual if actual is not None else "") if i == 0 else "",
                        "Completed Task": task,
                    })
            else:
                table_rows.append({
                    "Sprint Day": day_num,
                    "Date": date,
                    "Ideal Remaining (pts)": ideal,
                    "Actual Remaining (pts)": actual if actual is not None else "",
                    "Completed Task": "",
                })
        df_table = pd.DataFrame(table_rows)
        num_rows = len(df_table)
        row_height = 35
        table_height = min(max(num_rows * row_height + 50, 200), 500)
        st.dataframe(
            df_table,
            use_container_width=True,
            hide_index=True,
            height=table_height,
            column_config={
                "Sprint Day": st.column_config.TextColumn(width="small"),
                "Date": st.column_config.TextColumn(width="small"),
                "Ideal Remaining (pts)": st.column_config.TextColumn(width="small"),
                "Actual Remaining (pts)": st.column_config.TextColumn(width="small"),
                "Completed Task": st.column_config.TextColumn(width="large"),
            },
        )

        # Summary stats row
        stat_cols = st.columns(3)
        stat_cols[0].metric("Total Story Points", f"{bd['ideal_total']:.0f}")
        stat_cols[1].metric("# Days in Sprint", bd["sprint_days"])
        stat_cols[2].metric("Points to Complete / Day", f"{bd['pts_per_day']:.1f}")

        # Carry-in / spillover metrics row
        ci_tasks = bd.get("carry_in_tasks", [])
        ci_points = bd.get("carry_in_points", 0)
        sp_gids = bd.get("spillover_gids", set())
        sp_points = bd.get("spillover_points", 0)
        if ci_tasks or sp_gids:
            ci_cols = st.columns(3)
            if sp_gids:
                ci_cols[0].metric("Spillover Tasks", f"{len(sp_gids)} ({int(sp_points)} pts)")
            if ci_tasks:
                ci_cols[1].metric("Carry-in Tasks", f"{len(ci_tasks)} ({int(ci_points)} pts)")
            if ci_tasks:
                effective = bd['ideal_total'] + ci_points
                ci_cols[2].metric("Effective Work (Sprint + Carry-in)", f"{int(effective)} pts")

        # Download burndown data (rich multi-sheet Excel report)
        col_dl1, col_dl2, col_dl3 = st.columns([2, 1, 2])
        with col_dl2:
            today_date = datetime.now().strftime("%Y-%m-%d")
            sprint_slug = bd['sprint'].replace(' ', '_')
            all_sprint_tasks = list(filtered_results) + list(completed_results or [])
            excel_bytes = generate_burndown_excel_report(bd, reporter.client, all_sprint_tasks=all_sprint_tasks)
            st.download_button(
                label="Download Burndown Data",
                data=excel_bytes,
                file_name=f"burndown_{sprint_slug}_{today_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # Bug Count by Assignee Chart
    render_bug_count_chart(filtered_results, completed_results, filters.get("sprint"))

    st.markdown("---")

    # Completion Analytics Section
    st.subheader("Completion Analytics")
    col_team, col_individual = st.columns(2)

    with col_team:
        team_data = render_team_completion_chart(completed_results, filters, filters.get("sprint"))

    with col_individual:
        render_individual_completion_chart(completed_results, filters, filters.get("sprint"), team_data)

    st.markdown("---")

    # Invalid Story Points Alert (Quick Wins) - Shows both active and completed tasks
    render_invalid_story_points_section(filtered_results, completed_results, filters)

    # Overdue Tasks Alert (Quick Wins) - Most critical first
    render_overdue_alert_section(filtered_results)

    # Due This Week Alert (Quick Wins)
    render_due_this_week_section(filtered_results)

    # Alert sections (red first - more critical, then amber)
    render_red_alert_section(filtered_results)
    render_amber_alert_section(filtered_results)

    # Compliance summary
    col1, col2 = st.columns(2)
    with col1:
        render_attributes_summary(filtered_summary)
    with col2:
        render_assignee_table(filtered_summary)

    st.markdown("---")

    # Detailed findings
    render_compliance_details(filtered_results)

    st.markdown("---")

    # Download buttons
    render_download_buttons(filtered_results, filtered_summary, config, completed_results, filters)


if __name__ == "__main__":
    main()
